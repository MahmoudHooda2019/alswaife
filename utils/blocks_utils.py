import xlsxwriter
from typing import List, Dict
import os
import openpyxl
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter

TABLE1_COLUMNS = [
    "رقم النقلة", "عدد النقلات", "التاريخ", "المحجر", "رقم الماكينة",
    "رقم البلوك", "الخامة", "الطول", "العرض", "الارتفاع",
    "م3", "الوزن", "وزن البلوك", "سعر الطن", "إجمالي سعر البلوك", "سعر النقلة"
]

TABLE2_COLUMNS = [
    "تاريخ النشر", "رقم البلوك", "النوع", "رقم الماكينة", "وقت الدخول",
    "وقت الخروج", "عدد الساعات", "الاكراميه", "السمك", "العدد",
    "الطول بعد", "الخصم", "الارتفاع", "الكميه م2",
    "سعر النشر", "إجمالي سعر النشر", "إجمالي تكلفه البلوك"
]

TABLE3_COLUMNS = [
    "تاريخ التحميل", "رقم الفاتوره", "اسم العميل",
    "رقم لبلوك", "عدد الطاولات", "ط", "ع", "اجمالى م2",
    "سعر المتر", "اجمالى سعر المبيعات", "رصيد المخزون",
    "اجمالى تكلفه النقله", "اجمالى مبيعات النقله","ربح النقله"
]
TABLE3_WIDTH = [
    15, 12, 15, 12, 12, 10, 10, 15, 15, 18, 15, 18, 18, 15
]

def export_simple_blocks_excel(rows: List[Dict]) -> str:
    """إنشاء أو تحديث ملف Excel لسجل البلوكات"""
    documents_folder = os.path.join(
        os.path.expanduser("~"), "Documents", "alswaife", "البلوكات"
    )
    os.makedirs(documents_folder, exist_ok=True)
    
    filepath = os.path.join(documents_folder, "سجل البلوكات.xlsx")
    
    if os.path.exists(filepath):
        append_to_existing_file(filepath, rows)
    else:
        create_new_excel_file(filepath, rows)
    
    return filepath


def append_to_existing_file(filepath: str, new_rows: List[Dict]):
    """إضافة صفوف جديدة إلى ملف Excel موجود - تم تصحيح ترتيب الأعمدة"""
    try:
        workbook = openpyxl.load_workbook(filepath)
        worksheet = workbook["البلوكات"]
        
        start_row = worksheet.max_row + 1
        
        # تعريف الأنماط
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        gap_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        
        # إضافة البيانات الجديدة
        for i, block_data in enumerate(new_rows):
            excel_row = start_row + i
            
            thickness_text = block_data.get("thickness_dropdown", "2سم") or "2سم"
            
            # --- الجدول الأول (بدون تغيير) ---
            table1_data = [
                block_data.get("trip_number", ""),
                block_data.get("trip_count", ""),
                block_data.get("date", ""),
                block_data.get("quarry", ""),
                block_data.get("machine_number", ""),
                block_data.get("block_number", ""),
                block_data.get("material", ""),
                "",  # الطول (معادلة) - سيتم حسابه من UI + 0.20
                "",  # العرض (معادلة)
                "",  # الارتفاع (معادلة) - سيتم حسابه من مرحلة النشر
                "",  # م3 (معادلة)
                block_data.get("weight", ""),
                "",  # وزن البلوك (معادلة)
                block_data.get("price_per_ton", ""),
                "",  # إجمالي السعر (معادلة)
                block_data.get("trip_price", "")
            ]
            
            for col_idx, value in enumerate(table1_data, start=1):
                cell = worksheet.cell(row=excel_row, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = center_alignment
            
            # Length formula: length from UI + 0.20
            # We need to put the raw length value in a cell and then reference it in a formula
            length_value = block_data.get("length", "")
            if length_value != "":
                # Put the raw length value in a temporary cell (we'll use column 50 for this)
                worksheet.cell(row=excel_row, column=50, value=length_value).border = thin_border
                # Create formula that references this cell and adds 0.20
                length_formula = f'={get_column_letter(50)}{excel_row}+0.20'
                worksheet.cell(row=excel_row, column=8, value=length_formula).border = thin_border
            else:
                worksheet.cell(row=excel_row, column=8, value="").border = thin_border
            
            # معادلات الجدول الأول (نفس كودك السابق)
            thickness_col = get_column_letter(26) # Z
            count_col = get_column_letter(28)     # AB (لاحظ أن مكان العدد في الجدول الثاني سيتغير مكانه في الكود بالأسفل، لذا يجب التأكد من العمود الصحيح لاحقاً)
            # تحديث: بناءً على الجدول الثاني الجديد، العدد سيصبح في العمود 27 وليس 28.
            # لذا سنقوم بتحديث مرجع العمود هنا ليكون صحيحاً مع التصحيح الجديد
            count_col_fixed = get_column_letter(27) # AA هو موقع العدد الجديد
            
            width_formula = f'=((VALUE(SUBSTITUTE({thickness_col}{excel_row},"سم",""))+1)*{count_col_fixed}{excel_row})'
            cell = worksheet.cell(row=excel_row, column=9, value=width_formula)
            cell.border = thin_border; cell.alignment = center_alignment
            
            # Height formula: height in publishing stage + 0.5
            # Publishing stage height is in column 31 (AE)
            publish_height_col = get_column_letter(31)
            height_formula = f'={publish_height_col}{excel_row}+0.5'
            worksheet.cell(row=excel_row, column=10, value=height_formula).border = thin_border
            
            length_col = get_column_letter(8); width_col = get_column_letter(9); height_col = get_column_letter(10)
            worksheet.cell(row=excel_row, column=11, value=f"={length_col}{excel_row}*{width_col}{excel_row}*{height_col}{excel_row}").border = thin_border
            
            m3_col = get_column_letter(11); weight_col = get_column_letter(12)
            worksheet.cell(row=excel_row, column=13, value=f"={m3_col}{excel_row}*{weight_col}{excel_row}").border = thin_border
            
            price_col = get_column_letter(14); blk_weight_col = get_column_letter(13)
            worksheet.cell(row=excel_row, column=15, value=f"={price_col}{excel_row}*{blk_weight_col}{excel_row}").border = thin_border
            
            # الفجوة
            gap_cell = worksheet.cell(row=excel_row, column=17, value="")
            gap_cell.border = thin_border; gap_cell.fill = gap_fill
            
            # --- الجدول الثاني (تم التصحيح هنا) ---
            # البداية من العمود 18
            
            # 18-26: البيانات الأساسية
            worksheet.cell(row=excel_row, column=18, value=block_data.get("date", "")).border = thin_border
            worksheet.cell(row=excel_row, column=19, value=block_data.get("block_number", "")).border = thin_border
            worksheet.cell(row=excel_row, column=20, value=block_data.get("material", "")).border = thin_border
            worksheet.cell(row=excel_row, column=21, value=block_data.get("machine_number", "")).border = thin_border
            worksheet.cell(row=excel_row, column=22, value="").border = thin_border # دخول
            worksheet.cell(row=excel_row, column=23, value="").border = thin_border # خروج
            worksheet.cell(row=excel_row, column=24, value="").border = thin_border # ساعات
            worksheet.cell(row=excel_row, column=25, value=150).border = thin_border # إكرامية
            worksheet.cell(row=excel_row, column=26, value=thickness_text).border = thin_border # السمك
            
            # 27: العدد (كان سابقاً يوضع مكانه خانة فارغة خطأ)
            worksheet.cell(row=excel_row, column=27, value=block_data.get("quantity", 1)).border = thin_border
            
            # 28: الطول (length from UI + 0.20) - إزالة هذا العمود لتتجنب التكرار
            # تم حذف هذا العمود لإصلاح المشكلة التي كانت تسبب تكرار قيمة الطول
            # سيتم فقط عرض الطول في الجدول الأول
            
            # 29: الخصم (نضعه هنا قبل المعادلة لتكون جاهزة، ترتيب الأعمدة في إكسل حسب القائمة: 10=الطول بعد، 11=الخصم)
            # حسب القائمة: الطول بعد (10) -> الخصم (11)
            # بالأرقام: الطول بعد (28) -> الخصم (29)
            worksheet.cell(row=excel_row, column=29, value=0.20).border = thin_border
            
            # 28: الطول بعد = الطول من الجدول الأول - الخصم
            # بما أننا حذفنا عمود الطول، سنستخدم قيمة ثابتة للطول بعد
            # أو يمكننا استخدام الطول من الجدول الأول
            # لنستخدم الطول من الجدول الأول (العمود 8)
            table1_length_col = get_column_letter(8) # العمود الذي يحتوي على الطول في الجدول الأول
            disc_col = get_column_letter(29) # AC
            cell = worksheet.cell(row=excel_row, column=28, value=f"={table1_length_col}{excel_row}-{disc_col}{excel_row}")
            cell.border = thin_border; cell.alignment = center_alignment

            # 30: الارتفاع
            publish_height = block_data.get("publish_height", float(block_data.get("height", 0) or 0))
            worksheet.cell(row=excel_row, column=30, value=publish_height).border = thin_border
            
            # 31: الكمية م2 = الطول بعد × الارتفاع × العدد
            # ملاحظة: عادة الحساب يكون على "الطول بعد" (الصافي). إذا كنت تريد الحساب على الطول الخام، غير العمود 28 إلى 27
            len_after_col = get_column_letter(28) # AB
            height_pub_col = get_column_letter(30) # AE
            qty_col = get_column_letter(27) # AA (العدد)
            cell = worksheet.cell(row=excel_row, column=31, value=f"={len_after_col}{excel_row}*{height_pub_col}{excel_row}*{qty_col}{excel_row}")
            cell.border = thin_border; cell.alignment = center_alignment
            
            # 32: سعر النشر
            thickness_value = thickness_text.replace("سم", "")
            if thickness_value == "2": publish_price = 125
            elif thickness_value == "3": publish_price = 145
            elif thickness_value == "4": publish_price = 155
            else: publish_price = 125
            worksheet.cell(row=excel_row, column=32, value=publish_price).border = thin_border
            
            # 33: إجمالي سعر النشر = سعر النشر × الكمية م2
            pub_price_col = get_column_letter(32) # AG
            qty_m2_col = get_column_letter(31) # AF
            cell = worksheet.cell(row=excel_row, column=33, value=f"={pub_price_col}{excel_row}*{qty_m2_col}{excel_row}")
            cell.border = thin_border; cell.alignment = center_alignment
            
            # 34: إجمالي تكلفة البلوك = إجمالي سعر النشر + الإكرامية
            tot_pub_col = get_column_letter(33) # AH
            tips_col = get_column_letter(25) # Y
            cell = worksheet.cell(row=excel_row, column=34, value=f"={tot_pub_col}{excel_row}+{tips_col}{excel_row}")
            cell.border = thin_border; cell.alignment = center_alignment
            
        workbook.save(filepath)
        
    except Exception as e:
        print(f"خطأ في إضافة البيانات: {e}")
        # في حالة الخطأ الشديد، يمكن محاولة إعادة الإنشاء، لكن بحذر
        # create_new_excel_file(filepath, new_rows) 

def create_new_excel_file(filepath: str, rows: List[Dict]):
    """إنشاء ملف Excel جديد - تم تصحيح ترتيب الأعمدة"""
    workbook = xlsxwriter.Workbook(filepath, {'constant_memory': False})
    worksheet = workbook.add_worksheet("البلوكات")
    worksheet.right_to_left()

    title_fmt = workbook.add_format({"bold": True, "font_size": 18, "align": "center", "valign": "vcenter", "bg_color": "#2E75B6", "font_color": "white", "border": 1})
    table_title_fmt = workbook.add_format({"bold": True, "font_size": 14, "align": "center", "valign": "vcenter", "bg_color": "#5B9BD5", "font_color": "white", "border": 1})
    gap_fmt = workbook.add_format({"bg_color": "#5B9BD5", "border": 1})
    header_fmt = workbook.add_format({"bold": True, "border": 1, "align": "center", "valign": "vcenter", "bg_color": "#5B9BD5", "font_color": "white", "font_size": 12})
    data_fmt = workbook.add_format({"border": 1, "align": "center", "valign": "vcenter", "font_size": 11})

    total_cols = len(TABLE1_COLUMNS) + len(TABLE2_COLUMNS) + len(TABLE3_COLUMNS)
    worksheet.merge_range(0, 0, 0, total_cols, "سجل البلوكات", title_fmt)
    worksheet.merge_range(1, 0, 1, len(TABLE1_COLUMNS) - 1, "مقاس البلوك في الأرضية", table_title_fmt)
    worksheet.merge_range(1, len(TABLE1_COLUMNS), 2, len(TABLE1_COLUMNS), "", gap_fmt)
    worksheet.merge_range(1, len(TABLE1_COLUMNS) + 1, 1, total_cols - len(TABLE3_COLUMNS) - 4 , "مرحلة النشر", table_title_fmt)
    worksheet.merge_range(1, len(TABLE1_COLUMNS) + len(TABLE2_COLUMNS) + 1, 1, total_cols, "المنصرف والمبيعات", table_title_fmt)
    
    # اجمالي الكميه م2 - Formula to sum all values in the column
    qty_m2_col = get_column_letter(total_cols - len(TABLE3_COLUMNS) - 3 + 1)  # Convert to Excel column letter (AF)
    worksheet.write_formula(1, total_cols - len(TABLE3_COLUMNS) - 3, f"=SUM({qty_m2_col}3:{qty_m2_col}1000)", table_title_fmt) # اجمالي الكميه م2
    worksheet.write(1, total_cols - len(TABLE3_COLUMNS) - 2, "", table_title_fmt)
    # اجمالي اجمالي سعر النشر - Formula to sum all values in the column
    total_price_col = get_column_letter(total_cols - len(TABLE3_COLUMNS) - 1 + 1)  # Convert to Excel column letter (AH)
    worksheet.write_formula(1, total_cols - len(TABLE3_COLUMNS) - 1, f"=SUM({total_price_col}3:{total_price_col}1000)", table_title_fmt) # اجمالي اجمالي سعر النشر
    worksheet.write(1, total_cols - len(TABLE3_COLUMNS), "", table_title_fmt)

    for idx, col in enumerate(TABLE1_COLUMNS): worksheet.write(2, idx, col, header_fmt)
    worksheet.write(2, len(TABLE1_COLUMNS), "", gap_fmt)
    for idx, col in enumerate(TABLE2_COLUMNS): worksheet.write(2, len(TABLE1_COLUMNS) + 1 + idx, col, header_fmt)
    for idx, col in enumerate(TABLE3_COLUMNS): worksheet.write(2, len(TABLE1_COLUMNS) + len(TABLE2_COLUMNS) + 1 + idx, col, header_fmt)
    worksheet.set_row(1, 30); worksheet.set_row(2, 25)

    start_row = 3
    col_offset = len(TABLE1_COLUMNS) + 1 # العمود 17 (0-indexed) أي العمود الـ 18
    
    for i, block_data in enumerate(rows):
        excel_row = start_row + i
        thickness_text = block_data.get("thickness_dropdown", "2سم") or "2سم"
        
        # --- الجدول الأول ---
        table1_values = [
            block_data.get("trip_number", ""), block_data.get("trip_count", ""), block_data.get("date", ""), 
            block_data.get("quarry", ""), block_data.get("machine_number", ""), block_data.get("block_number", ""),
            block_data.get("material", ""), "", "", "",
            "", block_data.get("weight", ""), "", block_data.get("price_per_ton", ""), "", block_data.get("trip_price", "")
        ]
        
        for col_idx, value in enumerate(table1_values):
            worksheet.write(excel_row, col_idx, value, data_fmt)
        
        # Length formula: length from UI + 0.20
        length_value = block_data.get("length", "")
        if length_value != "":
            # Put the raw length value in a temporary cell (we'll use column 50 for this)
            worksheet.write(excel_row, 49, length_value, data_fmt)  # Column 50 (0-indexed = 49)
            # Create formula that references this cell and adds 0.20
            length_formula = f'={get_column_letter(50)}{excel_row + 1}+0.20'
            worksheet.write_formula(excel_row, 7, length_formula, data_fmt)  # Column 8 (0-indexed = 7)
        
        # معادلات الجدول الأول وتصحيح مرجع العدد
        # العدد موجود الآن في col_offset + 9 (index 9 في TABLE2_COLUMNS)
        thickness_col = get_column_letter(26) # Z
        count_col_fixed = get_column_letter(col_offset + 9 + 1) # (offset + index + 1 for A1 notation) -> العمود 27 = AA
        
        width_formula = f'=((VALUE(SUBSTITUTE({thickness_col}{excel_row + 1},"سم",""))+1)*{count_col_fixed}{excel_row + 1})'
        worksheet.write_formula(excel_row, 8, width_formula, data_fmt)
        
        # Height formula: height in publishing stage + 0.5
        # Publishing stage height is in col_offset + 13
        publish_height_col = get_column_letter(col_offset + 13 + 1)
        height_formula = f'={publish_height_col}{excel_row + 1}+0.5'
        worksheet.write_formula(excel_row, 9, height_formula, data_fmt)
        
        # باقي معادلات الجدول الأول
        l_col = get_column_letter(8); w_col = get_column_letter(9); h_col = get_column_letter(10)
        worksheet.write_formula(excel_row, 10, f"={l_col}{excel_row + 1}*{w_col}{excel_row + 1}*{h_col}{excel_row + 1}", data_fmt)
        
        m3_col = get_column_letter(11); wt_col = get_column_letter(12)
        worksheet.write_formula(excel_row, 12, f"={m3_col}{excel_row + 1}*{wt_col}{excel_row + 1}", data_fmt)
        
        pr_col = get_column_letter(14); bw_col = get_column_letter(13)
        worksheet.write_formula(excel_row, 14, f"={pr_col}{excel_row + 1}*{bw_col}{excel_row + 1}", data_fmt)
            
        worksheet.write(excel_row, len(TABLE1_COLUMNS), "", gap_fmt)
        
        # --- الجدول الثاني (تم التصحيح) ---
        # 0: Date, 1: Block, 2: Material, 3: Machine
        worksheet.write(excel_row, col_offset + 0, block_data.get("date", ""), data_fmt)
        worksheet.write(excel_row, col_offset + 1, block_data.get("block_number", ""), data_fmt)
        worksheet.write(excel_row, col_offset + 2, block_data.get("material", ""), data_fmt)
        worksheet.write(excel_row, col_offset + 3, block_data.get("machine_number", ""), data_fmt)
        
        # 4,5,6: Time In, Out, Hours
        worksheet.write(excel_row, col_offset + 4, "", data_fmt)
        worksheet.write(excel_row, col_offset + 5, "", data_fmt)
        worksheet.write(excel_row, col_offset + 6, "", data_fmt)
        
        # 7: Tips
        worksheet.write(excel_row, col_offset + 7, 150, data_fmt)
        
        # 8: Thickness
        worksheet.write(excel_row, col_offset + 8, thickness_text, data_fmt)
        
        # 9: Count (تم حذف العمود الفارغ ووضع العدد هنا)
        worksheet.write(excel_row, col_offset + 9, block_data.get("quantity", 1), data_fmt)
        
        # 10: الطول (تم حذف هذا العمود لتتجنب التكرار)
        # سيتم فقط عرض الطول في الجدول الأول
        
        # 11: Discount
        worksheet.write(excel_row, col_offset + 11, 0.20, data_fmt)
        
        # 10: Length After (Formula) = Length from Table 1 - Discount
        # لنستخدم الطول من الجدول الأول (العمود 8)
        table1_length_col = get_column_letter(8)  # العمود الذي يحتوي على الطول في الجدول الأول
        disc_cell = get_column_letter(col_offset + 11 + 1)
        worksheet.write_formula(excel_row, col_offset + 10, f'={table1_length_col}{excel_row + 1}-{disc_cell}{excel_row + 1}', data_fmt)
        
        # 12: Height
        publish_height = block_data.get("publish_height", float(block_data.get("height", 0) or 0))
        worksheet.write(excel_row, col_offset + 12, publish_height, data_fmt)
        
        # 13: Qty m2 (Formula) = Length After (10) * Height (12) * Count (9)
        len_aft_cell = get_column_letter(col_offset + 10 + 1)
        h_cell = get_column_letter(col_offset + 12 + 1)
        cnt_cell = get_column_letter(col_offset + 9 + 1)
        worksheet.write_formula(excel_row, col_offset + 13, f'={len_aft_cell}{excel_row + 1}*{h_cell}{excel_row + 1}*{cnt_cell}{excel_row + 1}', data_fmt)
        
        # 14: Publish Price
        thickness_value = thickness_text.replace("سم", "")
        if thickness_value == "2": publish_price = 125
        elif thickness_value == "3": publish_price = 145
        elif thickness_value == "4": publish_price = 155
        else: publish_price = 125
        worksheet.write(excel_row, col_offset + 14, publish_price, data_fmt)
        
        # 15: Total Publish Price (Formula) = Price (14) * Qty m2 (13)
        pr_cell = get_column_letter(col_offset + 14 + 1)
        qm2_cell = get_column_letter(col_offset + 13 + 1)
        worksheet.write_formula(excel_row, col_offset + 15, f'={pr_cell}{excel_row + 1}*{qm2_cell}{excel_row + 1}', data_fmt)
        
        # 16: Total Cost (Formula) = Total Publish (15) + Tips (7)
        tot_pub_cell = get_column_letter(col_offset + 15 + 1)
        tips_cell = get_column_letter(col_offset + 7 + 1)
        worksheet.write_formula(excel_row, col_offset + 16, f'={tot_pub_cell}{excel_row + 1}+{tips_cell}{excel_row + 1}', data_fmt)

    worksheet.freeze_panes(3, 0)
    
    # تنسيق عرض الأعمدة
    for i in range(len(TABLE1_COLUMNS)): worksheet.set_column(i, i, 15) # for header of col 1
    worksheet.set_column(len(TABLE1_COLUMNS), len(TABLE1_COLUMNS), 1) # for space
    for i in range(len(TABLE2_COLUMNS)): worksheet.set_column(len(TABLE1_COLUMNS) + 1 + i, len(TABLE1_COLUMNS) + 1 + i, 15) # for header of col 2
    for i in range(len(TABLE3_COLUMNS)):
        worksheet.set_column(len(TABLE1_COLUMNS) + 1 + len(TABLE2_COLUMNS) + i, len(TABLE1_COLUMNS) + 1 + len(TABLE2_COLUMNS) + i, TABLE3_WIDTH[i]) # for header of col 3
    workbook.close()