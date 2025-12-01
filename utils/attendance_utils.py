"""
Attendance Utilities for Employee Tracking
This module provides functions to create and manage attendance Excel files.
"""

import xlsxwriter
import os
from typing import List, Dict, Optional, Tuple
from datetime import datetime


def create_or_update_attendance(filepath: str, employees_data: List[Dict]) -> Tuple[bool, Optional[str]]:
    """
    Create or update attendance Excel file with weekly schedule.
    
    Args:
        filepath (str): Path to save the Excel file
        employees_data (list): List of employee dictionaries with structure:
            {
                'name': str,
                'friday_shift1': float,
                'friday_shift2': float,
                'saturday_shift1': float,
                'saturday_shift2': float,
                'sunday_shift1': float,
                'sunday_shift2': float,
                'monday_shift1': float,
                'monday_shift2': float,
                'tuesday_shift1': float,
                'tuesday_shift2': float,
                'wednesday_shift1': float,
                'wednesday_shift2': float,
                'thursday_shift1': float,
                'thursday_shift2': float,
                'date': str,
                'advance': float,
                'price': float
            }
    
    Returns:
        tuple: (success: bool, error_message: str or None)
    """
    try:
        workbook = xlsxwriter.Workbook(filepath)
        worksheet = workbook.add_worksheet("الحضور والانصراف")
        
        # RTL + Page settings (Removed page break preview as requested)
        worksheet.right_to_left()
        # worksheet.set_pagebreak_view()  # Removed as requested
        worksheet.hide_gridlines(2)
        
        # Page settings
        worksheet.set_paper(9)  # A4
        worksheet.set_landscape()  # عرضي لاستيعاب الأعمدة الكثيرة
        worksheet.set_margins(0.3, 0.3, 0.5, 0.5)
        
        # ================
        #   FORMATS
        # ================
        
        # Header format for day names
        header_day_fmt = workbook.add_format({
            'bold': True,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'font_name': 'Arial',
            'font_size': 12,
            'bg_color': '#4472C4',
            'font_color': '#FFFFFF'
        })
        
        # Header format for shift labels
        header_shift_fmt = workbook.add_format({
            'bold': True,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'font_name': 'Arial',
            'font_size': 10,
            'bg_color': '#D9E1F2',
            'font_color': '#000000'
        })
        
        # Header format for employee name column
        header_name_fmt = workbook.add_format({
            'bold': True,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'font_name': 'Arial',
            'font_size': 12,
            'bg_color': '#4472C4',
            'font_color': '#FFFFFF'
        })
        
        # Format for employee names
        name_fmt = workbook.add_format({
            'bold': True,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'font_name': 'Arial',
            'font_size': 11,
            'bg_color': '#E7E6E6'
        })
        
        # Format for data cells
        data_fmt = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'font_name': 'Arial',
            'font_size': 10
        })
        
        # Format for total column
        total_fmt = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'font_name': 'Arial',
            'font_size': 10,
            'bold': True,
            'bg_color': '#FFF2CC',
            'num_format': '0'
        })
        
        # Format for date and advance
        special_fmt = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'font_name': 'Arial',
            'font_size': 10,
            'bg_color': '#E2EFDA'
        })
        
        # Format for weekly total row
        weekly_total_fmt = workbook.add_format({
            'bold': True,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'font_name': 'Arial',
            'font_size': 12,
            'bg_color': '#FFD966',
            'font_color': '#000000'
        })
        
        # ================
        #   HEADERS
        # ================
        
        # Days of the week starting from Friday
        days = ['الجمعة', 'السبت', 'الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس']
        
        # Row 0: Employee name header + Day names (each spans 2 columns)
        worksheet.merge_range(0, 0, 1, 0, 'اسم الموظف', header_name_fmt)
        
        col = 1
        for day in days:
            worksheet.merge_range(0, col, 0, col + 1, day, header_day_fmt)
            col += 2
        
        # Total, Date, Advance, Price headers (span both rows)
        worksheet.merge_range(0, 15, 1, 15, 'الإجمالي', header_day_fmt)
        worksheet.merge_range(0, 16, 1, 16, 'التاريخ', header_day_fmt)
        worksheet.merge_range(0, 17, 1, 17, 'السلفة', header_day_fmt)
        worksheet.merge_range(0, 18, 1, 18, 'السعر', header_day_fmt)
        
        # Row 1: Shift labels
        col = 1
        for _ in days:
            worksheet.write(1, col, 'وردية اولي', header_shift_fmt)
            worksheet.write(1, col + 1, 'وردية ثانية', header_shift_fmt)
            col += 2
        
        # ================
        #   DATA ROWS
        # ================
        
        row = 2
        for emp in employees_data:
            # Employee name
            worksheet.write(row, 0, emp.get('name', ''), name_fmt)
            
            # Attendance data for each day (14 columns total)
            shift_keys = [
                'friday_shift1', 'friday_shift2',
                'saturday_shift1', 'saturday_shift2',
                'sunday_shift1', 'sunday_shift2',
                'monday_shift1', 'monday_shift2',
                'tuesday_shift1', 'tuesday_shift2',
                'wednesday_shift1', 'wednesday_shift2',
                'thursday_shift1', 'thursday_shift2'
            ]
            
            col = 1
            for key in shift_keys:
                value = emp.get(key, 0)
                if value:
                    worksheet.write_number(row, col, float(value), data_fmt)
                else:
                    worksheet.write(row, col, '', data_fmt)
                col += 1
            
            # Total formula: SUM(B{row}:O{row}) - R{row}
            # Columns B to O are columns 1 to 14 (attendance data)
            # Column R is column 17 (advance)
            excel_row = row + 1  # Excel uses 1-based indexing
            worksheet.write_formula(row, 15, f'=SUM(B{excel_row}:O{excel_row})-R{excel_row}', total_fmt)
            
            # Date
            worksheet.write(row, 16, emp.get('date', ''), special_fmt)
            
            # Advance
            advance = emp.get('advance', 0)
            if advance:
                worksheet.write_number(row, 17, float(advance), special_fmt)
            else:
                worksheet.write(row, 17, '', special_fmt)
            
            # Price
            price = emp.get('price', 0)
            if price:
                worksheet.write_number(row, 18, float(price), special_fmt)
            else:
                worksheet.write(row, 18, '', special_fmt)
            
            row += 1
        
        # ================
        #   WEEKLY TOTAL ROW
        # ================
        
        # Add weekly total row below all employee data
        weekly_total_row = row
        data_end_row = row  # Define data_end_row here
        
        # Weekly total label
        worksheet.write(weekly_total_row, 0, 'الإجمالي الأسبوعي', weekly_total_fmt)
        
        # Calculate weekly totals for each shift column (B to O)
        for col_idx in range(1, 15):  # Columns B to O (1 to 14)
            # Column letter calculation: A=0, B=1, C=2, etc.
            col_letter = chr(65 + col_idx)  # B=1 -> B, C=2 -> C, etc.
            # Sum all values in this column from row 3 to row with data
            data_end_row = row  # Last row with employee data
            worksheet.write_formula(weekly_total_row, col_idx, f'=SUM({col_letter}3:{col_letter}{data_end_row})', weekly_total_fmt)
        
        # Weekly total for overall totals column (P)
        worksheet.write_formula(weekly_total_row, 15, f'=SUM(P3:P{data_end_row})', weekly_total_fmt)
        
        # Weekly total for advances column (R)
        worksheet.write_formula(weekly_total_row, 17, f'=SUM(R3:R{data_end_row})', weekly_total_fmt)
        
        # Weekly total for prices column (S)
        worksheet.write_formula(weekly_total_row, 18, f'=SUM(S3:S{data_end_row})', weekly_total_fmt)
        
        # ================
        #   COLUMN WIDTHS
        # ================
        
        worksheet.set_column(0, 0, 20)  # Employee name
        worksheet.set_column(1, 14, 8)  # Attendance columns
        worksheet.set_column(15, 15, 10)  # Total
        worksheet.set_column(16, 16, 12)  # Date
        worksheet.set_column(17, 17, 10)  # Advance
        worksheet.set_column(18, 18, 10)  # Price
        
        workbook.close()
        return (True, None)
        
    except PermissionError:
        return (False, "file_locked")
    except Exception as e:
        return (False, f"error: {str(e)}")


def load_attendance_data(filepath: str) -> Tuple[bool, Optional[List[Dict]], Optional[str]]:
    """
    Load attendance data from an existing Excel file.
    
    Args:
        filepath (str): Path to the Excel file
    
    Returns:
        tuple: (success: bool, data: List[Dict] or None, error_message: str or None)
    """
    if not os.path.exists(filepath):
        return (False, None, "file_not_found")
    
    try:
        import openpyxl
        
        workbook = openpyxl.load_workbook(filepath, data_only=True)
        sheet = workbook.active
        
        if sheet is None:
            return (False, None, "invalid_sheet")
        
        employees_data = []
        
        # Start reading from row 3 (index 3 in openpyxl, 1-based)
        # Stop before the weekly total row (which is the last row)
        max_row = sheet.max_row
        if max_row > 3:  # If there's data beyond headers
            max_row -= 1  # Exclude the weekly total row
            
        for row_idx in range(3, max_row + 1):
            name = sheet.cell(row=row_idx, column=1).value
            
            # Skip empty rows
            if not name:
                continue
            
            emp_data = {
                'name': str(name),
                'friday_shift1': sheet.cell(row=row_idx, column=2).value or 0,
                'friday_shift2': sheet.cell(row=row_idx, column=3).value or 0,
                'saturday_shift1': sheet.cell(row=row_idx, column=4).value or 0,
                'saturday_shift2': sheet.cell(row=row_idx, column=5).value or 0,
                'sunday_shift1': sheet.cell(row=row_idx, column=6).value or 0,
                'sunday_shift2': sheet.cell(row=row_idx, column=7).value or 0,
                'monday_shift1': sheet.cell(row=row_idx, column=8).value or 0,
                'monday_shift2': sheet.cell(row=row_idx, column=9).value or 0,
                'tuesday_shift1': sheet.cell(row=row_idx, column=10).value or 0,
                'tuesday_shift2': sheet.cell(row=row_idx, column=11).value or 0,
                'wednesday_shift1': sheet.cell(row=row_idx, column=12).value or 0,
                'wednesday_shift2': sheet.cell(row=row_idx, column=13).value or 0,
                'thursday_shift1': sheet.cell(row=row_idx, column=14).value or 0,
                'thursday_shift2': sheet.cell(row=row_idx, column=15).value or 0,
                'date': sheet.cell(row=row_idx, column=17).value or '',
                'advance': sheet.cell(row=row_idx, column=18).value or 0,
                'price': sheet.cell(row=row_idx, column=19).value or 0
            }
            
            employees_data.append(emp_data)
        
        return (True, employees_data, None)
        
    except ImportError:
        return (False, None, "openpyxl_missing")
    except Exception as e:
        return (False, None, f"error: {str(e)}")