"""
Attendance View - UI for employee attendance tracking
"""

import flet as ft
import os
from datetime import datetime
import platform
import subprocess
from utils.attendance_utils import create_or_update_attendance, load_attendance_data
from utils.purchases_utils import add_income_record, export_purchases_to_excel
from tkinter import filedialog
import tkinter as tk
from utils.utils import resource_path, is_excel_running, get_current_date
import json
from typing import Optional


class AttendanceView:
    def __init__(self, page: ft.Page):
        self.page = page
        self.page.title = "الحضور والانصراف"
        self.page.rtl = True
        
        # Data storage
        self.employees_list = self.load_employees()
        self.attendance_data = {}  # Dictionary to store attendance status
        self.current_file = None
        
        # UI components
        self.date_field: Optional[ft.TextField] = None
        self.day_field: Optional[ft.TextField] = None
        self.shift_dropdown: Optional[ft.Dropdown] = None
        self.employees_container: Optional[ft.Column] = None
        
    def load_employees(self):
        """Load employees from JSON file"""
        try:
            employees_path = resource_path(os.path.join('data', 'employees.json'))
            if os.path.exists(employees_path):
                with open(employees_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    # Strip whitespace from employee names to avoid matching issues
                    if isinstance(data, list):
                        for emp in data:
                            if 'name' in emp and isinstance(emp['name'], str):
                                emp['name'] = emp['name'].strip()
                    return data if isinstance(data, list) else []
        except:
            pass
        return []

    def build_ui(self):
        """Build the attendance tracking UI"""
        
        # Create AppBar with title and actions
        app_bar = ft.AppBar(
            leading=ft.IconButton(
                icon=ft.Icons.ARROW_BACK,
                on_click=self.go_back,
                tooltip="العودة"
            ),
            title=ft.Text(
                "الحضور والانصراف",
                size=20,
                weight=ft.FontWeight.BOLD,
                color=ft.Colors.BLUE_200
            ),
            actions=[
                ft.IconButton(
                    icon=ft.Icons.FILE_OPEN,
                    on_click=self.open_attendance_file,
                    tooltip="فتح ملف الحضور"
                ),
                ft.IconButton(
                    icon=ft.Icons.ADD,
                    on_click=self.add_new_employee,
                    tooltip="إضافة موظف جديد"
                ),
                ft.Container(
                    content=ft.IconButton(
                        icon=ft.Icons.SAVE,
                        on_click=self.save_to_excel,
                        tooltip="حفظ البيانات"
                    ),
                    margin=ft.margin.only(left=40, right=15)
                )
            ],
            bgcolor=ft.Colors.GREY_900,
        )
        
        # Date selection with enhanced styling
        self.date_field = ft.TextField(
            label="التاريخ",
            value=get_current_date('%d/%m/%Y'),
            width=200,
            text_align=ft.TextAlign.CENTER,
            on_change=self.on_date_change,
            border_radius=12,
            border_color=ft.Colors.GREEN_700,
            focused_border_color=ft.Colors.GREEN_400,
            label_style=ft.TextStyle(color=ft.Colors.GREEN_300, size=14, weight=ft.FontWeight.BOLD),
            text_style=ft.TextStyle(weight=ft.FontWeight.W_600, size=16, color=ft.Colors.WHITE),
            filled=True,
            fill_color=ft.Colors.GREY_900,
            prefix_icon=ft.Icons.CALENDAR_TODAY,
            content_padding=ft.padding.symmetric(horizontal=15, vertical=15),
            cursor_color=ft.Colors.GREEN_400,
            border_width=2
        )
        
        # Day field (automatically populated based on date) with enhanced styling
        self.day_field = ft.TextField(
            label="اليوم",
            width=200,
            text_align=ft.TextAlign.CENTER,
            disabled=True,
            border_radius=12,
            border_color=ft.Colors.GREEN_700,
            focused_border_color=ft.Colors.GREEN_400,
            label_style=ft.TextStyle(color=ft.Colors.GREEN_300, size=14, weight=ft.FontWeight.BOLD),
            text_style=ft.TextStyle(weight=ft.FontWeight.W_600, size=16, color=ft.Colors.WHITE),
            filled=True,
            fill_color=ft.Colors.GREY_900,
            prefix_icon=ft.Icons.TODAY,
            content_padding=ft.padding.symmetric(horizontal=15, vertical=15),
            border_width=2
        )
        
        # Populate day field based on current date
        self.update_day_field(self.date_field.value)
        
        self.shift_dropdown = ft.Dropdown(
            label="الوردية",
            options=[
                ft.dropdown.Option("الاولي", "الاولي"),
                ft.dropdown.Option("الثانية", "الثانية")
            ],
            width=200,
            on_change=self.on_shift_change,
            border_radius=12,
            border_color=ft.Colors.GREEN_700,
            focused_border_color=ft.Colors.GREEN_400,
            label_style=ft.TextStyle(color=ft.Colors.GREEN_300, size=14, weight=ft.FontWeight.BOLD),
            text_style=ft.TextStyle(weight=ft.FontWeight.W_600, size=16, color=ft.Colors.WHITE),
            filled=True,
            bgcolor=ft.Colors.GREY_900,
            icon=ft.Icons.WORK,
            content_padding=ft.padding.symmetric(horizontal=15, vertical=15),
            border_width=2,
            value="الاولي"
        )
        
        # Create a prominent header section for date info with gradient-like effect
        date_info_header = ft.Container(
            content=ft.Column(
                controls=[
                    ft.Row(
                        controls=[
                            ft.Container(
                                content=self.date_field,
                                shadow=ft.BoxShadow(
                                    spread_radius=1,
                                    blur_radius=10,
                                    color=ft.Colors.with_opacity(0.3, ft.Colors.GREEN_400),
                                    offset=ft.Offset(0, 2)
                                )
                            ),
                            ft.Container(
                                content=self.day_field,
                                shadow=ft.BoxShadow(
                                    spread_radius=1,
                                    blur_radius=10,
                                    color=ft.Colors.with_opacity(0.3, ft.Colors.GREEN_400),
                                    offset=ft.Offset(0, 2)
                                )
                            ),
                            ft.Container(
                                content=self.shift_dropdown,
                                shadow=ft.BoxShadow(
                                    spread_radius=1,
                                    blur_radius=10,
                                    color=ft.Colors.with_opacity(0.3, ft.Colors.GREEN_400),
                                    offset=ft.Offset(0, 2)
                                )
                            )
                        ],
                        alignment=ft.MainAxisAlignment.CENTER,
                        spacing=30
                    )
                ],
                spacing=10
            ),
            padding=ft.padding.symmetric(horizontal=30, vertical=25),
            bgcolor=ft.Colors.GREY_900,
            border=ft.border.all(2, ft.Colors.GREY_700),
            border_radius=ft.border_radius.all(16),
            margin=ft.margin.only(bottom=20, left=15, right=15, top=10),
            shadow=ft.BoxShadow(
                spread_radius=2,
                blur_radius=15,
                color=ft.Colors.with_opacity(0.5, ft.Colors.BLACK),
                offset=ft.Offset(0, 4)
            )
        )
        
        # Employees container
        self.employees_container = ft.Column(
            controls=[],
            spacing=10
        )
        
        # Load existing data for current date if available
        self.load_existing_data()
        
        # Set the AppBar
        self.page.appbar = app_bar
        
        # Main layout - Column with scroll for content below AppBar
        main_content = ft.Column(
            controls=[
                date_info_header,
                self.employees_container
            ],
            spacing=10,
            scroll=ft.ScrollMode.AUTO,
            expand=True
        )
        
        # Add the main content to the page
        self.page.add(main_content)
        self.page.update()
    
    def on_date_change(self, e):
        """Handle date change"""
        if self.date_field is not None:
            date_str = self.date_field.value
            if date_str:
                self.update_day_field(date_str)
                if (self.shift_dropdown is not None and 
                    self.shift_dropdown.value):
                    self.load_existing_data()
                else:
                    if self.employees_container is not None:
                        self.employees_container.controls.clear()
                        self.employees_container.controls.append(
                            ft.Container(
                                content=ft.Text(
                                    "الرجاء اختيار الوردية لعرض بيانات الحضور",
                                    size=18,
                                    text_align=ft.TextAlign.CENTER,
                                    color=ft.Colors.GREY_600
                                ),
                                alignment=ft.alignment.center,
                                padding=50
                            )
                        )
                self.page.update()
    
    def on_shift_change(self, e):
        """Handle shift change"""
        if (self.date_field is not None and 
            self.date_field.value):
            self.load_existing_data()
        self.page.update()
    
    def update_day_field(self, date_str):
        """Update day field based on date"""
        if self.day_field is None:
            return
            
        try:
            date_obj = datetime.strptime(date_str, '%d/%m/%Y')
            arabic_days = ['الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة', 'السبت', 'الأحد']
            day_name = arabic_days[date_obj.weekday()]
            self.day_field.value = day_name
        except:
            self.day_field.value = ""
    
    def normalize_date(self, date_str):
        """Normalize date string to dd/mm/yyyy format for comparison"""
        if not date_str:
            return ""
        
        try:
            # Try to parse different date formats
            date_str = str(date_str).strip()
            
            # Remove Arabic digits
            arabic_to_english = str.maketrans('٠١٢٣٤٥٦٧٨٩', '0123456789')
            date_str = date_str.translate(arabic_to_english)
            
            # Try different formats
            for fmt in ['%d/%m/%Y', '%Y/%m/%d', '%d-%m-%Y', '%Y-%m-%d']:
                try:
                    dt = datetime.strptime(date_str, fmt)
                    normalized = dt.strftime('%d/%m/%Y')
                    return normalized
                except:
                    continue
            
            return date_str
        except Exception as e:
            return str(date_str)
    
    def load_existing_data(self):
        """Load existing attendance data for current date and shift"""
        
        if self.employees_container is not None:
            self.employees_container.controls.clear()
        
        if not self.date_field or not self.date_field.value or not self.shift_dropdown or not self.shift_dropdown.value:
            if self.employees_container is not None:
                self.employees_container.controls.append(
                    ft.Container(
                        content=ft.Text(
                            "الرجاء اختيار التاريخ والوردية أولاً لعرض قائمة الموظفين",
                            size=18,
                            text_align=ft.TextAlign.CENTER,
                            color=ft.Colors.GREY_600
                        ),
                        alignment=ft.alignment.center,
                        padding=50
                    )
                )
            self.page.update()
            return
        
        documents_path = os.path.join(os.path.expanduser("~"), "Documents")
        alswaife_path = os.path.join(documents_path, "alswaife")
        attendance_path = os.path.join(alswaife_path, "حضور وانصراف")
        
        try:
            filename = "سجل الحضور والانصراف.xlsx"
            filepath = os.path.join(attendance_path, filename)
            
            # Normalize current date for comparison
            current_date_normalized = self.normalize_date(self.date_field.value)
            
            if os.path.exists(filepath):
                success, data, error = load_attendance_data(filepath)
                
                if success and data:
                    self.current_file = filepath
                    
                    # Filter data for the current date
                    filtered_data = []
                    for emp_record in data:
                        record_date = self.normalize_date(emp_record.get('date', ''))
                        if record_date == current_date_normalized:
                            filtered_data.append(emp_record)
                    
                    # Create a dictionary for quick lookup
                    employee_data = {emp['name']: emp for emp in filtered_data}
                    
                    # Get shift key
                    shift_key = self.get_shift_key()
                    
                    # Create employee rows with existing data
                    for emp in self.employees_list:
                        emp_name = emp['name'].strip() if isinstance(emp['name'], str) else emp['name']
                        price = emp.get('price', 0)
                        
                        is_present = False
                        emp_price = price
                        
                        # Look for this employee in the filtered data (using stripped names for comparison)
                        matched_record = None
                        for record_name, emp_record in employee_data.items():
                            if record_name.strip() == emp_name:
                                matched_record = emp_record
                                break
                        
                        if matched_record:
                            shift_value = matched_record.get(shift_key, 0)
                            if shift_key and shift_value > 0:
                                is_present = True
                            if 'price' in matched_record and matched_record['price'] != 0:
                                emp_price = matched_record['price']
                        
                        self.add_employee_row(emp_name, emp_price, is_present)
                    
                    # Load additional employees from Excel not in JSON
                    for emp_record in filtered_data:
                        emp_name = emp_record['name'].strip() if isinstance(emp_record['name'], str) else emp_record['name']
                        if not any(emp['name'].strip() == emp_name for emp in self.employees_list):
                            price = emp_record.get('price', 0)
                            is_present = shift_key and emp_record.get(shift_key, 0) > 0
                            self.add_employee_row(emp_name, price, is_present)
                else:
                    # Create employee rows without existing data
                    for emp in self.employees_list:
                        emp_name = emp['name'].strip() if isinstance(emp['name'], str) else emp['name']
                        price = emp.get('price', 0)
                        self.add_employee_row(emp_name, price, False)
            else:
                # Create employee rows without existing data
                for emp in self.employees_list:
                    emp_name = emp['name'].strip() if isinstance(emp['name'], str) else emp['name']
                    price = emp.get('price', 0)
                    self.add_employee_row(emp_name, price, False)
        except Exception as e:
            # Create employee rows without existing data
            for emp in self.employees_list:
                emp_name = emp['name'].strip() if isinstance(emp['name'], str) else emp['name']
                price = emp.get('price', 0)
                self.add_employee_row(emp_name, price, False)
        
        self.page.update()
    
    def get_shift_key(self):
        """Get the shift key based on current day and shift selection"""
        if self.day_field is None or self.shift_dropdown is None:
            return None
            
        if not self.day_field.value or not self.shift_dropdown.value:
            return None
            
        day_mapping = {
            'الاثنين': 'monday',
            'الثلاثاء': 'tuesday',
            'الأربعاء': 'wednesday',
            'الخميس': 'thursday',
            'الجمعة': 'friday',
            'السبت': 'saturday',
            'الأحد': 'sunday'
        }
        
        shift_mapping = {
            'الاولي': 'shift1',
            'الثانية': 'shift2'
        }
        
        day_key = day_mapping.get(self.day_field.value, '')
        shift_key = shift_mapping.get(self.shift_dropdown.value, '')
        
        if day_key and shift_key:
            return f"{day_key}_{shift_key}"
        return None
    
    def add_employee_row(self, name, price, is_present=False):
        """Add an employee row to the UI with improved layout"""
        shift_selected = bool(self.shift_dropdown and self.shift_dropdown.value)
        
        price_field = ft.TextField(
            value=str(price),
            width=150,
            text_align=ft.TextAlign.CENTER,
            keyboard_type=ft.KeyboardType.NUMBER,
            label="السعر",
            dense=True,
            disabled=not shift_selected,
            border_radius=10,
            border_color=ft.Colors.GREEN_700,
            focused_border_color=ft.Colors.GREEN_400,
            label_style=ft.TextStyle(color=ft.Colors.GREEN_300, size=13, weight=ft.FontWeight.BOLD),
            text_style=ft.TextStyle(weight=ft.FontWeight.W_600, size=15, color=ft.Colors.WHITE),
            filled=True,
            fill_color=ft.Colors.GREY_900,
            suffix_text="جنيه",
            suffix_style=ft.TextStyle(color=ft.Colors.GREEN_300, size=12),
            content_padding=ft.padding.symmetric(horizontal=12, vertical=12),
            cursor_color=ft.Colors.GREEN_400,
            border_width=2
        )
        
        # Check if this employee is from JSON or added manually
        is_json_employee = any(emp['name'].strip() == name.strip() for emp in self.employees_list)
        
        # Create row controls
        row_controls = [
            ft.Checkbox(
                value=is_present,
                disabled=not shift_selected,
                on_change=lambda e, n=name: self.on_attendance_change(n, e.control.value),
                scale=1.2
            ),
            ft.Text(name, size=18, weight=ft.FontWeight.BOLD, color=ft.Colors.WHITE),
            ft.Container(expand=True),
            price_field
        ]
        
        # Add delete button for all employees, but disable for JSON employees
        delete_button = ft.IconButton(
            icon=ft.Icons.DELETE,
            icon_color=ft.Colors.RED_400 if not is_json_employee else ft.Colors.GREY_600,
            tooltip="حذف الموظف" if not is_json_employee else "لا يمكن حذف الموظفين الأساسيين",
            on_click=lambda e, n=name: self.delete_employee(n) if not is_json_employee else None,
            disabled=is_json_employee
        )
        row_controls.append(delete_button)
        
        card = ft.Card(
            content=ft.Container(
                content=ft.Row(
                    controls=row_controls,
                    alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                ),
                padding=ft.padding.symmetric(horizontal=25, vertical=18),
                border_radius=14,
                gradient=ft.LinearGradient(
                    begin=ft.alignment.center_left,
                    end=ft.alignment.center_right,
                    colors=[
                        ft.Colors.GREY_900,
                        ft.Colors.GREY_800,
                    ]
                )
            ),
            elevation=6,
            shadow_color=ft.Colors.with_opacity(0.4, ft.Colors.BLACK),
            shape=ft.RoundedRectangleBorder(radius=14),
            margin=ft.margin.symmetric(vertical=6, horizontal=15),
            surface_tint_color=ft.Colors.BLUE_700
        )
        
        self.attendance_data[name] = {
            'card': card,
            'price_field': price_field,
            'present': is_present,
            'is_json_employee': is_json_employee
        }
        
        if self.employees_container is not None:
            self.employees_container.controls.append(card)
    
    def on_attendance_change(self, employee_name, is_present):
        """Handle attendance checkbox change"""
        if employee_name in self.attendance_data:
            self.attendance_data[employee_name]['present'] = is_present
    
    def save_to_excel(self, e):
        """Save attendance data to Excel file"""
        
        # التحقق من أن Excel مغلق
        if is_excel_running():
            self._show_excel_warning_dialog()
            return
        
        self._do_save()

    def _show_excel_warning_dialog(self):
        """Show Excel warning dialog with continue option"""
        def close_dlg(e=None):
            dlg.open = False
            self.page.update()

        def continue_save(e=None):
            dlg.open = False
            self.page.update()
            self._do_save()

        dlg = ft.AlertDialog(
            title=ft.Text("تحذير", color=ft.Colors.ORANGE_400, weight=ft.FontWeight.BOLD),
            content=ft.Text("برنامج Excel مفتوح حالياً.\nيرجى إغلاقه قبل الحفظ.", size=16, rtl=True),
            actions=[
                ft.TextButton(
                    "متابعة على أي حال",
                    on_click=continue_save,
                    style=ft.ButtonStyle(color=ft.Colors.ORANGE_400)
                ),
                ft.TextButton(
                    "إلغاء",
                    on_click=close_dlg,
                    style=ft.ButtonStyle(color=ft.Colors.GREY_400)
                ),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
            bgcolor=ft.Colors.BLUE_GREY_900
        )
        self.page.overlay.append(dlg)
        dlg.open = True
        self.page.update()

    def _do_save(self):
        """تنفيذ عملية الحفظ الفعلية"""
        documents_path = os.path.join(os.path.expanduser("~"), "Documents")
        alswaife_path = os.path.join(documents_path, "alswaife")
        attendance_path = os.path.join(alswaife_path, "حضور وانصراف")
        
        try:
            os.makedirs(attendance_path, exist_ok=True)
        except OSError as ex:
            self.show_message(f"فشل إنشاء المجلد: {ex}", error=True)
            return
        
        if self.shift_dropdown is None or not self.shift_dropdown.value:
            self.show_message("الرجاء اختيار الوردية", error=True)
            return
        
        try:
            filename = "سجل الحضور والانصراف.xlsx"
            filepath = os.path.join(attendance_path, filename)
        except Exception as ex:
            self.show_message(f"خطأ في إنشاء اسم الملف: {ex}", error=True)
            return
        
        existing_data = []
        if os.path.exists(filepath):
            success, data, error = load_attendance_data(filepath)
            if success and data:
                existing_data = data
        
        employees_data = []
        all_employee_names = set()
        
        # Normalize current date
        current_date = self.normalize_date(self.date_field.value if self.date_field else "")
        
        shift_key = self.get_shift_key()
        
        # Process employees from JSON
        for emp in self.employees_list:
            emp_name = emp['name']
            all_employee_names.add(emp_name)
            
            existing_record = None
            for record in existing_data:
                record_date = self.normalize_date(record.get('date', ''))
                if record['name'] == emp_name and record_date == current_date:
                    existing_record = record.copy()
                    break
            
            if existing_record:
                emp_record = existing_record
            else:
                emp_record = {
                    'name': emp_name,
                    'friday_shift1': 0, 'friday_shift2': 0,
                    'saturday_shift1': 0, 'saturday_shift2': 0,
                    'sunday_shift1': 0, 'sunday_shift2': 0,
                    'monday_shift1': 0, 'monday_shift2': 0,
                    'tuesday_shift1': 0, 'tuesday_shift2': 0,
                    'wednesday_shift1': 0, 'wednesday_shift2': 0,
                    'thursday_shift1': 0, 'thursday_shift2': 0,
                    'date': self.date_field.value if self.date_field is not None else "",
                    'advance': 0,
                    'price': emp.get('price', 0)
                }
            
            if emp_name in self.attendance_data:
                attendance_info = self.attendance_data[emp_name]
                is_present = attendance_info['present']
                price_field = attendance_info['price_field']
                
                try:
                    price = float(price_field.value) if price_field.value else emp.get('price', 0)
                except:
                    price = emp.get('price', 0)
                
                emp_record['price'] = price
                
                if shift_key:
                    emp_record[shift_key] = price if is_present else 0
            
            employees_data.append(emp_record)
        
        # Process manually added employees
        for emp_name, attendance_info in self.attendance_data.items():
            if emp_name not in all_employee_names:
                existing_record = None
                for record in existing_data:
                    record_date = self.normalize_date(record.get('date', ''))
                    if record['name'] == emp_name and record_date == current_date:
                        existing_record = record.copy()
                        break
                
                if existing_record:
                    emp_record = existing_record
                else:
                    emp_record = {
                        'name': emp_name,
                        'friday_shift1': 0, 'friday_shift2': 0,
                        'saturday_shift1': 0, 'saturday_shift2': 0,
                        'sunday_shift1': 0, 'sunday_shift2': 0,
                        'monday_shift1': 0, 'monday_shift2': 0,
                        'tuesday_shift1': 0, 'tuesday_shift2': 0,
                        'wednesday_shift1': 0, 'wednesday_shift2': 0,
                        'thursday_shift1': 0, 'thursday_shift2': 0,
                        'date': self.date_field.value if self.date_field is not None else "",
                        'advance': 0,
                        'price': 0
                    }
                
                is_present = attendance_info['present']
                price_field = attendance_info['price_field']
                
                try:
                    price = float(price_field.value) if price_field.value else 0
                except:
                    price = 0
                
                emp_record['price'] = price
                
                if shift_key:
                    emp_record[shift_key] = price if is_present else 0
                
                employees_data.append(emp_record)
        
        # Combine with existing data
        final_data = []
        employee_names = [emp['name'] for emp in employees_data]
        
        for record in existing_data:
            record_date = self.normalize_date(record.get('date', ''))
            if record_date != current_date or record['name'] not in employee_names:
                final_data.append(record)
        
        final_data.extend(employees_data)
        
        success, error = create_or_update_attendance(filepath, final_data)
        
        if success:
            self.current_file = filepath
            
            # Record attendance expenses to the expenses sheet
            self.record_attendance_expenses(employees_data, current_date)
            
            self.show_message(f"تم الحفظ بنجاح: {os.path.basename(filepath)}", filepath=filepath)
        else:
            if error == "file_locked":
                self.show_message("الملف مفتوح في برنامج آخر، الرجاء إغلاقه", error=True)
            else:
                self.show_message(f"خطأ في الحفظ: {error}", error=True)
    
    def record_attendance_expenses(self, employees_data, current_date):
        """Record attendance costs as expenses in the expenses sheet"""
        try:
            # Calculate total attendance cost and count for this date
            total_cost = 0
            present_count = 0
            shift_key = self.get_shift_key()
            
            for emp in employees_data:
                if shift_key:
                    cost = emp.get(shift_key, 0)
                    if cost:
                        total_cost += float(cost)
                        present_count += 1
            
            if total_cost > 0:
                # Get the expenses file path
                documents_path = os.path.join(os.path.expanduser("~"), "Documents", "alswaife")
                expenses_filepath = os.path.join(
                    documents_path, "ايرادات ومصروفات", "بيان مصروفات وايرادات مصنع جرانيت السويفى.xlsx"
                )
                
                # Create expense record with shift name including "الوردية"
                shift_name = self.shift_dropdown.value if self.shift_dropdown else ""
                item_name = f"حضور وانصراف - الوردية {shift_name}"
                
                expense_record = {
                    'quantity': present_count,
                    'item_name': item_name,
                    'total_price': total_cost,
                    'date': current_date
                }
                
                # Update or add expense record
                self.update_or_add_expense(expenses_filepath, expense_record)
                print(f"[DEBUG] Recorded attendance expense: {total_cost} for {present_count} employees on {current_date}")
        except Exception as e:
            print(f"[ERROR] Failed to record attendance expense: {e}")
    
    def update_or_add_expense(self, filepath, record):
        """Update existing expense record or add new one based on item_name and date"""
        try:
            import openpyxl
            from openpyxl.styles import Border, Side, Alignment, PatternFill
            
            if not os.path.exists(filepath):
                # File doesn't exist, create new
                export_purchases_to_excel([record], filepath)
                return
            
            workbook = openpyxl.load_workbook(filepath)
            
            if 'المصروفات' not in workbook.sheetnames:
                workbook.close()
                export_purchases_to_excel([record], filepath)
                return
            
            worksheet = workbook['المصروفات']
            
            # Search for existing record with same item_name and date
            found_row = None
            for row in range(3, worksheet.max_row + 1):
                item_name = worksheet.cell(row=row, column=2).value
                date_val = worksheet.cell(row=row, column=4).value
                
                if item_name == record['item_name'] and str(date_val) == str(record['date']):
                    found_row = row
                    break
            
            if found_row:
                # Update existing record
                worksheet.cell(row=found_row, column=1, value=record['quantity'])
                worksheet.cell(row=found_row, column=3, value=record['total_price'])
                workbook.save(filepath)
                workbook.close()
                print(f"[DEBUG] Updated existing expense record at row {found_row}")
            else:
                workbook.close()
                # Add new record
                export_purchases_to_excel([record], filepath)
                print(f"[DEBUG] Added new expense record")
        except Exception as e:
            print(f"[ERROR] update_or_add_expense failed: {e}")
            # Fallback to adding new record
            export_purchases_to_excel([record], filepath)
    
    def show_message(self, message, error=False, filepath=None):
        """Show status message with dialog notification"""
        if hasattr(self, 'dialog') and self.dialog:
            self.dialog.open = False
        
        if not error and filepath:
            self.dialog = ft.AlertDialog(
                title=ft.Row(
                    controls=[
                        ft.Icon(ft.Icons.CHECK_CIRCLE, color=ft.Colors.GREEN_400, size=30),
                        ft.Text("تم الحفظ بنجاح", color=ft.Colors.GREEN_300, weight=ft.FontWeight.BOLD, rtl=True),
                    ],
                    rtl=True,
                    spacing=10
                ),
                content=ft.Column(
                    rtl=True,
                    controls=[
                        ft.Text("تم حفظ ملف الحضور والانصراف بنجاح:", size=14, rtl=True),
                        ft.Container(
                            content=ft.Text(
                                os.path.basename(filepath),
                                size=13,
                                color=ft.Colors.BLUE_200,
                                weight=ft.FontWeight.W_500,
                                rtl=True
                            ),
                            bgcolor=ft.Colors.BLUE_GREY_800,
                            padding=10,
                            border_radius=8,
                            margin=ft.margin.only(top=10),
                            rtl=True
                        )
                    ],
                    tight=True
                ),
                actions=[
                    ft.TextButton(
                        "فتح الملف", 
                        on_click=lambda e: self.open_file(filepath),
                        style=ft.ButtonStyle(color=ft.Colors.BLUE_300)
                    ),
                    ft.TextButton(
                        "فتح المسار", 
                        on_click=lambda e: self.open_folder(filepath),
                        style=ft.ButtonStyle(color=ft.Colors.BLUE_300)
                    ),
                    ft.TextButton(
                        "إغلاق", 
                        on_click=lambda e: self.close_dialog(),
                        style=ft.ButtonStyle(color=ft.Colors.GREY_400)
                    ),
                ],
                actions_alignment=ft.MainAxisAlignment.END,
                bgcolor=ft.Colors.BLUE_GREY_900,
                shape=ft.RoundedRectangleBorder(radius=16)
            )
        elif not error:  # Success message without filepath
            self.dialog = ft.AlertDialog(
                title=ft.Row(
                    controls=[
                        ft.Icon(ft.Icons.CHECK_CIRCLE, color=ft.Colors.GREEN_400, size=30),
                        ft.Text("تمت العملية بنجاح", color=ft.Colors.GREEN_300, weight=ft.FontWeight.BOLD, rtl=True),
                    ],
                    rtl=True,
                    spacing=10
                ),
                content=ft.Text(message, size=16, rtl=True),
                actions=[
                    ft.TextButton(
                        "إغلاق", 
                        on_click=lambda e: self.close_dialog(),
                        style=ft.ButtonStyle(color=ft.Colors.GREY_400)
                    ),
                ],
                actions_alignment=ft.MainAxisAlignment.END,
                bgcolor=ft.Colors.GREEN_900,
                shape=ft.RoundedRectangleBorder(radius=16)
            )
        else:
            self.dialog = ft.AlertDialog(
                title=ft.Row(
                    controls=[
                        ft.Icon(ft.Icons.ERROR, color=ft.Colors.RED_400, size=30),
                        ft.Text("خطأ في الحفظ", color=ft.Colors.RED_300, weight=ft.FontWeight.BOLD, rtl=True),
                    ],
                    rtl=True,
                    spacing=10
                ),
                content=ft.Text(message, size=16, rtl=True),
                actions=[
                    ft.TextButton(
                        "إغلاق", 
                        on_click=lambda e: self.close_dialog(),
                        style=ft.ButtonStyle(color=ft.Colors.GREY_400)
                    ),
                ],
                actions_alignment=ft.MainAxisAlignment.END,
                bgcolor=ft.Colors.RED_900,
                shape=ft.RoundedRectangleBorder(radius=16)
            )
        
        self.page.overlay.append(self.dialog)
        self.dialog.open = True
        self.page.update()
    
    def close_dialog(self):
        """Close the dialog"""
        if hasattr(self, 'dialog') and self.dialog:
            self.dialog.open = False
            self.page.update()
    
    def open_file(self, filepath):
        """Open the saved Excel file"""
        try:
            if platform.system() == 'Windows':
                os.startfile(filepath)
            elif platform.system() == 'Darwin':
                subprocess.call(['open', filepath])
            else:
                subprocess.call(['xdg-open', filepath])
        except Exception as ex:
            self.show_message(f"فشل في فتح الملف: {ex}", error=True)
        finally:
            self.close_dialog()
    
    def open_folder(self, filepath):
        """Open the folder containing the saved Excel file"""
        try:
            folder_path = os.path.dirname(filepath)
            if platform.system() == 'Windows':
                os.startfile(folder_path)
            elif platform.system() == 'Darwin':
                subprocess.call(['open', folder_path])
            else:
                subprocess.call(['xdg-open', folder_path])
        except Exception as ex:
            self.show_message(f"فشل في فتح المسار: {ex}", error=True)
        finally:
            self.close_dialog()
    
    def add_new_employee(self, e):
        """Add a new employee not in the JSON file"""
        if not self.date_field or not self.date_field.value or not self.shift_dropdown or not self.shift_dropdown.value:
            self.show_message("الرجاء اختيار التاريخ والوردية أولاً", error=True)
            return
        
        self.name_field = ft.TextField(
            label="اسم الموظف",
            width=300,
            autofocus=True,
            border_radius=8,
            prefix_icon=ft.Icons.PERSON,
            border_color=ft.Colors.GREY_600,
            focused_border_color=ft.Colors.GREY_400,
            color=ft.Colors.WHITE,
            label_style=ft.TextStyle(color=ft.Colors.GREY_400),
            text_align=ft.TextAlign.RIGHT,
            rtl=True
        )
        
        self.price_field_new = ft.TextField(
            label="السعر",
            width=300,
            keyboard_type=ft.KeyboardType.NUMBER,
            value="400",
            border_radius=8,
            prefix_icon=ft.Icons.ATTACH_MONEY,
            border_color=ft.Colors.GREY_600,
            focused_border_color=ft.Colors.GREY_400,
            color=ft.Colors.WHITE,
            label_style=ft.TextStyle(color=ft.Colors.GREY_400),
            suffix_text="جنيه",
            text_align=ft.TextAlign.RIGHT,
            rtl=True
        )
        
        self.add_employee_dialog = ft.AlertDialog(
            title=ft.Text("إضافة موظف جديد", text_align=ft.TextAlign.CENTER, rtl=True),
            content=ft.Container(
                content=ft.Column(
                    controls=[
                        ft.Container(
                            content=self.name_field,
                            padding=ft.padding.symmetric(vertical=5)
                        ),
                        ft.Container(
                            content=self.price_field_new,
                            padding=ft.padding.symmetric(vertical=5)
                        )
                    ],
                    spacing=15,
                    tight=True
                ),
                padding=20,
                width=350
            ),
            actions=[
                ft.Container(
                    rtl=True,
                    content=ft.Row(
                        controls=[
                            ft.ElevatedButton(
                                "إضافة", 
                                on_click=self.confirm_add_employee,
                                bgcolor=ft.Colors.GREY_700,
                                color=ft.Colors.WHITE,
                                style=ft.ButtonStyle(
                                    shape=ft.RoundedRectangleBorder(radius=8)
                                )
                            ),
                            ft.TextButton(
                                "إلغاء", 
                                on_click=self.close_add_employee_dialog,
                                style=ft.ButtonStyle(
                                    color=ft.Colors.GREY_400,
                                    overlay_color=ft.Colors.GREY_700
                                )
                            ),
                        ],
                        alignment=ft.MainAxisAlignment.START,
                        spacing=10
                    ),
                    padding=ft.padding.only(top=10)
                )
            ],
            shape=ft.RoundedRectangleBorder(radius=16),
            bgcolor=ft.Colors.GREY_800,
            shadow_color=ft.Colors.with_opacity(0.3, ft.Colors.BLACK),
        )
        
        self.page.overlay.append(self.add_employee_dialog)
        self.add_employee_dialog.open = True
        self.page.update()
    
    def close_add_employee_dialog(self, e=None):
        """Close the add employee dialog"""
        if hasattr(self, 'add_employee_dialog') and self.add_employee_dialog:
            self.add_employee_dialog.open = False
            self.page.update()
    
    def confirm_add_employee(self, e):
        """Confirm adding the new employee"""
        if not self.name_field.value or not self.name_field.value.strip():
            self.show_message("الرجاء إدخال اسم الموظف", error=True)
            return
        
        employee_name = self.name_field.value.strip()
        
        if employee_name in self.attendance_data:
            self.show_message("الموظف موجود بالفعل", error=True)
            return
        
        try:
            price = float(self.price_field_new.value) if self.price_field_new.value else 400
        except:
            price = 400
        
        self.add_employee_row(employee_name, price, False)
        self.close_add_employee_dialog()
        self.page.update()
        self.show_message(f"تم إضافة الموظف: {employee_name}")
    
    def delete_employee(self, employee_name):
        """Delete an employee (only for manually added employees)"""
        if employee_name in self.attendance_data:
            employee_info = self.attendance_data[employee_name]
            
            # Check if this is a manually added employee (not from JSON)
            is_json_employee = employee_info.get('is_json_employee', True)
            
            if not is_json_employee:
                # This is a manually added employee, allow deletion
                card = employee_info['card']
                if self.employees_container and card in self.employees_container.controls:
                    self.employees_container.controls.remove(card)
                
                del self.attendance_data[employee_name]
                self.page.update()
                self.show_message(f"تم حذف الموظف: {employee_name}")
            else:
                # This is a JSON employee, show error message
                self.show_message("لا يمكن حذف الموظفين الأساسيين", error=True)
        else:
            self.show_message("الموظف غير موجود", error=True)
    
    def go_back(self, e):
        """Go back to dashboard"""
        from views.dashboard_view import DashboardView
        
        self.page.clean()
        dashboard = DashboardView(self.page)
        
        save_callback = getattr(self.page, '_save_callback', None)
        if save_callback is not None:
            dashboard.show(save_callback)
        else:
            try:
                from main import save_callback
                dashboard.show(save_callback)
            except:
                dashboard.show(None)
    
    def open_attendance_file(self, e):
        """Open the attendance Excel file directly"""
        documents_path = os.path.join(os.path.expanduser("~"), "Documents", "alswaife")
        attendance_path = os.path.join(documents_path, "حضور وانصراف")
        filepath = os.path.join(attendance_path, "سجل الحضور والانصراف.xlsx")
        
        if os.path.exists(filepath):
            try:
                if platform.system() == 'Windows':
                    os.startfile(filepath)
                elif platform.system() == 'Darwin':
                    subprocess.call(['open', filepath])
                else:
                    subprocess.call(['xdg-open', filepath])
            except Exception as ex:
                self.show_message(f"فشل في فتح الملف: {ex}", error=True)
        else:
            self.show_message("ملف الحضور غير موجود، قم بحفظ البيانات أولاً", error=True)