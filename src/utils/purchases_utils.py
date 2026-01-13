"""
Excel Utilities for Purchases Management
This module provides functions to generate and manage Excel files for income and expenses data.
Contains three sheets: Income (الإيرادات), Expenses (المصروفات), Summary (الإجمالي)
"""

import xlsxwriter
import openpyxl
import os
from typing import List, Dict
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font

from utils.log_utils import log_error, log_exception


def export_purchases_to_excel(records: List[Dict], filepath: str) -> str:
    """
    Export purchases data to the expenses sheet of the Excel file.
    """
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    if os.path.exists(filepath):
        append_to_expenses(filepath, records)
    else:
        create_purchases_excel_file(filepath, [], records)
    
    return filepath


def add_income_record(filepath: str, record: Dict) -> str:
    """
    Add an income record to the income sheet of the Excel file.
    Called from invoice saving to record income.
    If a record with the same invoice_number exists, it will be updated.
    """
    try:
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        if os.path.exists(filepath):
            if update_income_record(filepath, record):
                pass  # Updated existing record
            else:
                append_to_income(filepath, [record])
        else:
            create_purchases_excel_file(filepath, [record], [])
        
        return filepath
    except Exception as e:
        log_exception(f"add_income_record failed: {e}")
        raise


def update_income_record(filepath: str, record: Dict) -> bool:
    """
    Update an existing income record by invoice_number.
    Returns True if record was found and updated, False otherwise.
    """
    invoice_number = record.get('invoice_number', '')
    if not invoice_number:
        return False
    
    try:
        workbook = openpyxl.load_workbook(filepath)
        
        # Get income sheet
        if 'الإيرادات' not in workbook.sheetnames:
            workbook.close()
            return False
        
        worksheet = workbook['الإيرادات']
        
        # Search for the invoice number in column A (column 1)
        for row in range(3, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row, column=1).value
            if str(cell_value) == str(invoice_number):
                # Found the record, update it
                worksheet.cell(row=row, column=2, value=record.get('client', ''))
                worksheet.cell(row=row, column=3, value=record.get('amount', ''))
                worksheet.cell(row=row, column=4, value=record.get('date', ''))
                
                workbook.save(filepath)
                workbook.close()
                return True
        
        workbook.close()
        return False
    except Exception as e:
        log_error(f"update_income_record failed: {e}")
        return False


def create_purchases_excel_file(filepath: str, income_records: List[Dict] = None, expense_records: List[Dict] = None):
    """
    Create a new Excel file with single sheet containing income, expenses, and summary side by side.
    """
    if income_records is None:
        income_records = []
    if expense_records is None:
        expense_records = []
    
    workbook = xlsxwriter.Workbook(filepath)
    
    # ==================== FORMATS ====================
    title_format = workbook.add_format({
        'bold': True, 'align': 'center', 'valign': 'vcenter',
        'bg_color': '#1F4E78', 'font_color': 'white', 'font_size': 18, 'border': 2
    })
    
    section_title_format = workbook.add_format({
        'bold': True, 'align': 'center', 'valign': 'vcenter',
        'bg_color': '#4472C4', 'font_color': 'white', 'font_size': 14, 'border': 1
    })
    
    income_header_format = workbook.add_format({
        'bold': True, 'border': 1, 'align': 'center', 'valign': 'vcenter',
        'bg_color': '#00B050', 'font_color': 'white', 'font_size': 12
    })
    
    expenses_header_format = workbook.add_format({
        'bold': True, 'border': 1, 'align': 'center', 'valign': 'vcenter',
        'bg_color': '#C00000', 'font_color': 'white', 'font_size': 12
    })
    
    cell_format = workbook.add_format({
        'border': 1, 'align': 'center', 'valign': 'vcenter', 'font_size': 11
    })
    
    cell_format_alt = workbook.add_format({
        'border': 1, 'align': 'center', 'valign': 'vcenter', 'font_size': 11, 'bg_color': '#F2F2F2'
    })
    
    number_format = workbook.add_format({
        'border': 1, 'align': 'center', 'valign': 'vcenter', 'font_size': 11, 'num_format': '#,##0'
    })
    
    number_format_alt = workbook.add_format({
        'border': 1, 'align': 'center', 'valign': 'vcenter', 'font_size': 11,
        'bg_color': '#F2F2F2', 'num_format': '#,##0'
    })
    
    summary_label_format = workbook.add_format({
        'bold': True, 'align': 'center', 'valign': 'vcenter',
        'bg_color': '#D9E1F2', 'font_color': '#1F4E78', 'font_size': 14, 'border': 2
    })
    
    # Income-themed formats (Green)
    summary_income_format = workbook.add_format({
        'bold': True, 'align': 'center', 'valign': 'vcenter',
        'bg_color': '#C6EFCE', 'font_color': '#006100', 'font_size': 16, 'border': 2, 'num_format': '#,##0'
    })
    
    income_section_title_format = workbook.add_format({
        'bold': True, 'align': 'center', 'valign': 'vcenter',
        'bg_color': '#00B050', 'font_color': 'white', 'font_size': 14, 'border': 1
    })
    
    income_summary_label_format = workbook.add_format({
        'bold': True, 'align': 'center', 'valign': 'vcenter',
        'bg_color': '#C6EFCE', 'font_color': '#006100', 'font_size': 14, 'border': 2
    })
    
    # Expenses-themed formats (Red)
    summary_expenses_format = workbook.add_format({
        'bold': True, 'align': 'center', 'valign': 'vcenter',
        'bg_color': '#FFC7CE', 'font_color': '#9C0006', 'font_size': 16, 'border': 2, 'num_format': '#,##0'
    })
    
    expenses_section_title_format = workbook.add_format({
        'bold': True, 'align': 'center', 'valign': 'vcenter',
        'bg_color': '#C00000', 'font_color': 'white', 'font_size': 14, 'border': 1
    })
    
    expenses_summary_label_format = workbook.add_format({
        'bold': True, 'align': 'center', 'valign': 'vcenter',
        'bg_color': '#FFC7CE', 'font_color': '#9C0006', 'font_size': 14, 'border': 2
    })
    
    # Balance format (Blue - distinctive)
    summary_balance_format = workbook.add_format({
        'bold': True, 'align': 'center', 'valign': 'vcenter',
        'bg_color': '#BDD7EE', 'font_color': '#1F4E78', 'font_size': 16, 'border': 2, 'num_format': '#,##0'
    })
    
    balance_summary_label_format = workbook.add_format({
        'bold': True, 'align': 'center', 'valign': 'vcenter',
        'bg_color': '#BDD7EE', 'font_color': '#1F4E78', 'font_size': 14, 'border': 2
    })
    
    # ==================== SINGLE SHEET: الإيرادات والمصروفات ====================
    worksheet = workbook.add_worksheet('الإيرادات والمصروفات')
    worksheet.right_to_left()
    
    current_row = 0
    
    # ==================== MAIN TITLE ====================
    worksheet.merge_range(current_row, 0, current_row, 6, 'بيان الإيرادات والمصروفات', title_format)
    worksheet.set_row(current_row, 35)
    current_row += 1  # Remove empty row - go directly to summary
    
    # ==================== SUMMARY SECTION (SIDE BY SIDE) ====================
    # Summary headers - use themed colors matching their respective tables
    worksheet.merge_range(current_row, 0, current_row, 1, 'إجمالي الإيرادات', income_summary_label_format)  # Green theme
    worksheet.merge_range(current_row, 2, current_row, 4, 'الرصيد المتبقي', balance_summary_label_format)  # Blue theme
    worksheet.merge_range(current_row, 5, current_row, 6, 'إجمالي المصروفات', expenses_summary_label_format)  # Red theme
    current_row += 1
    
    # Calculate data ranges for formulas
    income_data_start = current_row + 2  # Updated - one less row
    expenses_data_start = current_row + 2  # Updated - one less row
    
    # Summary values - will be filled later with correct formulas
    summary_values_row = current_row  # Store this row number for later use
    current_row += 1  # Remove empty row - go directly to tables
    
    # ==================== TABLES SECTION (SIDE BY SIDE) ====================
    tables_start_row = current_row
    
    # Income section (Left side - columns A-C) - Green theme
    worksheet.merge_range(current_row, 0, current_row, 2, 'سجل الإيرادات', income_section_title_format)
    current_row += 1
    
    # Income headers
    income_headers = ["اسم العميل", "المبلغ", "التاريخ"]
    for col, header in enumerate(income_headers):
        worksheet.write(current_row, col, header, income_header_format)
    worksheet.set_row(current_row, 25)
    
    # Expenses section (Right side - columns D-G) - Red theme
    worksheet.merge_range(tables_start_row, 3, tables_start_row, 6, 'سجل المصروفات', expenses_section_title_format)
    
    # Expenses headers
    expenses_headers = ["العدد", "البيان", "المبلغ", "التاريخ"]
    for col, header in enumerate(expenses_headers):
        worksheet.write(current_row, col + 3, header, expenses_header_format)
    
    current_row += 1
    income_data_start = current_row
    expenses_data_start = current_row
    
    # Determine max rows needed
    max_rows = max(len(income_records), len(expense_records))
    
    # Income data
    for i in range(max_rows):
        if i < len(income_records):
            record = income_records[i]
            is_alt = i % 2 == 1
            cf = cell_format_alt if is_alt else cell_format
            nf = number_format_alt if is_alt else number_format
            
            worksheet.write(current_row + i, 0, record.get('client', ''), cf)
            worksheet.write(current_row + i, 1, record.get('amount', ''), nf)
            worksheet.write(current_row + i, 2, record.get('date', ''), cf)
        else:
            # Empty rows to maintain alignment
            for col in range(3):
                worksheet.write(current_row + i, col, '', cell_format)
    
    # Expenses data
    for i in range(max_rows):
        if i < len(expense_records):
            record = expense_records[i]
            is_alt = i % 2 == 1
            cf = cell_format_alt if is_alt else cell_format
            nf = number_format_alt if is_alt else number_format
            
            worksheet.write(current_row + i, 3, record.get('quantity', ''), cf)
            worksheet.write(current_row + i, 4, record.get('item_name', ''), cf)
            worksheet.write(current_row + i, 5, record.get('total_price', ''), nf)
            worksheet.write(current_row + i, 6, record.get('date', ''), cf)
        else:
            # Empty rows to maintain alignment
            for col in range(3, 7):
                worksheet.write(current_row + i, col, '', cell_format)
    
    # Update summary formulas with correct ranges
    if income_records:
        income_end_row = income_data_start + len(income_records) - 1
        worksheet.merge_range(summary_values_row, 0, summary_values_row, 1, f'=SUM(B{income_data_start + 1}:B{income_end_row + 1})', summary_income_format)
    else:
        worksheet.merge_range(summary_values_row, 0, summary_values_row, 1, 0, summary_income_format)
    
    # Balance formula (Income total - Expenses total) - now spans C-D-E for maximum visibility
    worksheet.merge_range(summary_values_row, 2, summary_values_row, 4, f'=A{summary_values_row + 1}-F{summary_values_row + 1}', summary_balance_format)
    
    if expense_records:
        expenses_end_row = expenses_data_start + len(expense_records) - 1
        worksheet.merge_range(summary_values_row, 5, summary_values_row, 6, f'=SUM(F{expenses_data_start + 1}:F{expenses_end_row + 1})', summary_expenses_format)
    else:
        worksheet.merge_range(summary_values_row, 5, summary_values_row, 6, 0, summary_expenses_format)
    
    # Column widths
    worksheet.set_column(0, 0, 20)  # اسم العميل
    worksheet.set_column(1, 1, 12)  # المبلغ (إيرادات)
    worksheet.set_column(2, 2, 12)  # التاريخ (إيرادات)
    worksheet.set_column(3, 3, 8)   # العدد
    worksheet.set_column(4, 4, 25)  # البيان
    worksheet.set_column(5, 5, 12)  # المبلغ (مصروفات)
    worksheet.set_column(6, 6, 12)  # التاريخ (مصروفات)
    
    try:
        workbook.close()
    except PermissionError as e:
        raise PermissionError("الملف مفتوح في Excel. أغلقه وحاول مرة أخرى.") from e


def append_to_income(filepath: str, new_records: List[Dict]):
    """
    Append new income records to the unified sheet (left side, columns A-C).
    """
    try:
        workbook = openpyxl.load_workbook(filepath)
        
        sheet_name = 'الإيرادات والمصروفات'
        if sheet_name not in workbook.sheetnames:
            workbook.close()
            raise ValueError("Unified sheet not found")
        
        worksheet = workbook[sheet_name]
        
        # Find the income section (left side, columns A-C)
        income_section_start = None
        
        # Search more thoroughly for the income headers
        for row in range(1, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row, column=1).value
            if cell_value and str(cell_value).strip() == "اسم العميل":
                income_section_start = row + 1  # Data starts after header
                break
        
        # If not found, try alternative search patterns
        if income_section_start is None:
            for row in range(1, worksheet.max_row + 1):
                cell_value = worksheet.cell(row=row, column=1).value
                if cell_value and str(cell_value).strip() in ["اسم العميل", "العميل", "Client", "اسم العميل"]:
                    income_section_start = row + 1
                    break
        
        # If still not found, look for the income table section title first
        if income_section_start is None:
            for row in range(1, worksheet.max_row + 1):
                # Check if this row contains the income section title
                for col in range(1, 4):  # Check columns A-C
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value and "سجل الإيرادات" in str(cell_value):
                        # Found section title, look for headers in next few rows
                        for header_row in range(row + 1, min(row + 5, worksheet.max_row + 1)):
                            header_cell = worksheet.cell(row=header_row, column=1).value
                            if header_cell and str(header_cell).strip() in ["اسم العميل", "العميل"]:
                                income_section_start = header_row + 1
                                break
                        break
                if income_section_start:
                    break
        
        # Last resort: look for any pattern that suggests income table structure
        if income_section_start is None:
            for row in range(1, worksheet.max_row + 1):
                # Look for a row where column A has client-like content and column B has amount-like content
                col_a = worksheet.cell(row=row, column=1).value
                col_b = worksheet.cell(row=row, column=2).value
                col_c = worksheet.cell(row=row, column=3).value
                
                # Check if this looks like a header row for income
                if (col_a and ("عميل" in str(col_a) or "اسم" in str(col_a)) and
                    col_b and ("مبلغ" in str(col_b) or "المبلغ" in str(col_b)) and
                    col_c and ("تاريخ" in str(col_c) or "التاريخ" in str(col_c))):
                    income_section_start = row + 1
                    break
        
        if income_section_start is None:
            # If we still can't find it, let's create a new file structure
            workbook.close()
            log_error("Income section not found - recreating file structure")
            # Create new file with current payment as first record
            create_purchases_excel_file(filepath, new_records, [])
            return
        
        # Find the last row with income data (columns A-C)
        start_row = income_section_start
        for row in range(income_section_start, worksheet.max_row + 2):
            # Check if all income columns are empty
            if (worksheet.cell(row=row, column=1).value is None and 
                worksheet.cell(row=row, column=2).value is None and
                worksheet.cell(row=row, column=3).value is None):
                start_row = row
                break
            start_row = row + 1
        
        # Styles
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        white_fill = PatternFill(fill_type=None)
        alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        for row_idx, record in enumerate(new_records, start=start_row):
            is_alt = (row_idx - income_section_start) % 2 == 1
            current_fill = alt_fill if is_alt else white_fill
            
            # Income data in columns A-C (no invoice number)
            cell = worksheet.cell(row=row_idx, column=1, value=record.get('client', ''))
            cell.border = thin_border
            cell.fill = current_fill
            cell.alignment = center_alignment
            
            cell = worksheet.cell(row=row_idx, column=2, value=record.get('amount', ''))
            cell.border = thin_border
            cell.fill = current_fill
            cell.alignment = center_alignment
            cell.number_format = '#,##0'
            
            cell = worksheet.cell(row=row_idx, column=3, value=record.get('date', ''))
            cell.border = thin_border
            cell.fill = current_fill
            cell.alignment = center_alignment
        
        workbook.save(filepath)
        workbook.close()
    except PermissionError as e:
        raise PermissionError("الملف مفتوح في Excel. أغلقه وحاول مرة أخرى.") from e


def append_to_expenses(filepath: str, new_records: List[Dict]):
    """
    Append new expense records to the unified sheet (right side, columns D-G).
    """
    try:
        workbook = openpyxl.load_workbook(filepath)
        
        sheet_name = 'الإيرادات والمصروفات'
        if sheet_name not in workbook.sheetnames:
            workbook.close()
            raise ValueError("Unified sheet not found")
        
        worksheet = workbook[sheet_name]
        
        # Find the expenses section (right side, columns D-G)
        expenses_section_start = None
        for row in range(1, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row, column=4).value  # Column D
            if cell_value == "العدد":  # Expenses header
                expenses_section_start = row + 1  # Data starts after header
                break
        
        # If not found, try alternative search
        if expenses_section_start is None:
            for row in range(1, worksheet.max_row + 1):
                cell_value = worksheet.cell(row=row, column=4).value
                if str(cell_value).strip() in ["العدد", "الكمية", "Quantity"]:
                    expenses_section_start = row + 1
                    break
        
        # If still not found, look for the expenses table section
        if expenses_section_start is None:
            for row in range(1, worksheet.max_row + 1):
                cell_value = worksheet.cell(row=row, column=4).value
                if cell_value == "سجل المصروفات":
                    # Found section title, look for headers in next few rows
                    for header_row in range(row + 1, min(row + 5, worksheet.max_row + 1)):
                        header_cell = worksheet.cell(row=header_row, column=4).value
                        if str(header_cell).strip() in ["العدد", "الكمية"]:
                            expenses_section_start = header_row + 1
                            break
                    break
        
        if expenses_section_start is None:
            workbook.close()
            raise ValueError("Expenses section not found")
        
        # Find the last row with expenses data (columns D-G)
        start_row = expenses_section_start
        for row in range(expenses_section_start, worksheet.max_row + 2):
            # Check if all expenses columns are empty
            if (worksheet.cell(row=row, column=4).value is None and 
                worksheet.cell(row=row, column=5).value is None and
                worksheet.cell(row=row, column=6).value is None and
                worksheet.cell(row=row, column=7).value is None):
                start_row = row
                break
            start_row = row + 1
        
        # Styles
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        white_fill = PatternFill(fill_type=None)
        alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        for row_idx, record in enumerate(new_records, start=start_row):
            is_alt = (row_idx - expenses_section_start) % 2 == 1
            current_fill = alt_fill if is_alt else white_fill
            
            # Expenses data in columns D-G
            cell = worksheet.cell(row=row_idx, column=4, value=record.get('quantity', ''))
            cell.border = thin_border
            cell.fill = current_fill
            cell.alignment = center_alignment
            
            cell = worksheet.cell(row=row_idx, column=5, value=record.get('item_name', ''))
            cell.border = thin_border
            cell.fill = current_fill
            cell.alignment = center_alignment
            
            cell = worksheet.cell(row=row_idx, column=6, value=record.get('total_price', ''))
            cell.border = thin_border
            cell.fill = current_fill
            cell.alignment = center_alignment
            cell.number_format = '#,##0'
            
            cell = worksheet.cell(row=row_idx, column=7, value=record.get('date', ''))
            cell.border = thin_border
            cell.fill = current_fill
            cell.alignment = center_alignment
        
        workbook.save(filepath)
        workbook.close()
    except PermissionError as e:
        raise PermissionError("الملف مفتوح في Excel. أغلقه وحاول مرة أخرى.") from e


def add_payment_to_income_file(client_name: str, payment_date: str, amount: float, notes: str = "") -> bool:
    """
    Add payment record to the income file (ايرادات ومصروفات).
    
    Args:
        client_name: Name of the client
        payment_date: Date of payment
        amount: Payment amount (positive value)
        notes: Additional notes
        
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        # Get the correct documents path
        documents_path = os.path.join(os.path.expanduser("~"), "Documents", "alswaife")
        
        # Create the income/expenses directory if it doesn't exist
        income_dir = os.path.join(documents_path, "ايرادات ومصروفات")
        os.makedirs(income_dir, exist_ok=True)
        
        # Path to income/expenses file
        income_file_path = os.path.join(income_dir, "بيان مصروفات وايرادات مصنع جرانيت السويفى.xlsx")
        
        # Create the record for income sheet
        payment_record = {
            'client': client_name,  # Changed from 'invoice_number' to 'client'
            'amount': abs(amount),  # Ensure positive amount for income
            'date': payment_date
        }
        
        # Add to income file
        if os.path.exists(income_file_path):
            append_to_income(income_file_path, [payment_record])
        else:
            # Create new file with this payment
            create_purchases_excel_file(income_file_path, [payment_record], [])
        
        return True
        
    except Exception as e:
        log_error(f"Error adding payment to income file: {e}")
        return False


def remove_payment_from_income_file(client_name: str, payment_date: str) -> bool:
    """
    Remove payment record from the income file.
    
    Args:
        client_name: Name of the client
        payment_date: Date of payment
        
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        import openpyxl
        
        # Get the correct documents path
        documents_path = os.path.join(os.path.expanduser("~"), "Documents", "alswaife")
        
        # Create the income/expenses directory path
        income_dir = os.path.join(documents_path, "ايرادات ومصروفات")
        if not os.path.exists(income_dir):
            return True  # Directory doesn't exist, nothing to remove
        
        # Path to income/expenses file
        income_file_path = os.path.join(income_dir, "بيان مصروفات وايرادات مصنع جرانيت السويفى.xlsx")
        
        if not os.path.exists(income_file_path):
            return True  # File doesn't exist, nothing to remove
        
        workbook = openpyxl.load_workbook(income_file_path)
        
        sheet_name = 'الإيرادات والمصروفات'
        if sheet_name not in workbook.sheetnames:
            workbook.close()
            return True
        
        worksheet = workbook[sheet_name]
        
        # Find the income section (left side, column A for client names)
        income_section_start = None
        for row in range(1, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row, column=1).value
            if cell_value == "اسم العميل":  # Income header
                income_section_start = row + 1  # Data starts after header
                break
        
        # If not found, try alternative search
        if income_section_start is None:
            for row in range(1, worksheet.max_row + 1):
                cell_value = worksheet.cell(row=row, column=1).value
                if str(cell_value).strip() in ["اسم العميل", "العميل", "Client"]:
                    income_section_start = row + 1
                    break
        
        if income_section_start is None:
            workbook.close()
            return True
        
        # Search for the payment record by client name and date
        rows_to_delete = []
        for row in range(income_section_start, worksheet.max_row + 1):
            client_cell = worksheet.cell(row=row, column=1).value  # Column A - client name
            date_cell = worksheet.cell(row=row, column=3).value    # Column C - date
            
            # Check if this is a payment record for this client and date
            if (str(client_cell) == client_name and 
                str(date_cell) == payment_date):
                rows_to_delete.append(row)
        
        # Delete rows in reverse order to maintain row indices
        for row in reversed(rows_to_delete):
            worksheet.delete_rows(row)
        
        workbook.save(income_file_path)
        workbook.close()
        return True
        
    except Exception as e:
        log_error(f"Error removing payment from income file: {e}")
        return False


def load_item_names_from_excel(filepath: str) -> List[str]:
    """
    Load existing item names from the unified sheet for auto-complete.
    """
    items = set()
    
    if not os.path.exists(filepath):
        return list(items)
    
    try:
        workbook = openpyxl.load_workbook(filepath)
        
        sheet_name = 'الإيرادات والمصروفات'
        if sheet_name not in workbook.sheetnames:
            workbook.close()
            return list(items)
        
        worksheet = workbook[sheet_name]
        
        # Find the expenses section (right side, column E for item names)
        expenses_section_start = None
        for row in range(1, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row, column=4).value  # Column D
            if cell_value == "العدد":  # Expenses header
                expenses_section_start = row + 1  # Data starts after header
                break
        
        # If not found, try alternative search
        if expenses_section_start is None:
            for row in range(1, worksheet.max_row + 1):
                cell_value = worksheet.cell(row=row, column=4).value
                if str(cell_value).strip() in ["العدد", "الكمية", "Quantity"]:
                    expenses_section_start = row + 1
                    break
        
        if expenses_section_start:
            # Read item names from column E (column 5) in expenses section
            for row in range(expenses_section_start, worksheet.max_row + 1):
                item_name = worksheet.cell(row=row, column=5).value  # Column E
                if item_name and str(item_name).strip():
                    items.add(str(item_name).strip())
                
        workbook.close()
    except PermissionError:
        return []
    except Exception as e:
        log_error(f"Error loading items from Excel: {e}")
    
    return list(items)
