"""
Excel Utilities for Invoice Creation
This module provides functions to generate Excel invoices from invoice data.
"""

import xlsxwriter
import openpyxl
from openpyxl.cell.cell import MergedCell
from typing import List, Tuple
import os

from utils.log_utils import log_error, log_exception


def save_invoice(filepath: str, op_num: str, client: str, driver: str,
                 items: List[Tuple], date_str: str = "", phone: str = ""):
    """
    Save invoice data to an Excel file with proper formatting.
    
    Args:
        filepath (str): Path to save the Excel file
        op_num (str): Operation/invoice number
        client (str): Client name
        driver (str): Driver name
        items (List[Tuple]): List of invoice items, each tuple contains:
            (description, block, thickness, material, count, length, height, price)
        date_str (str, optional): Date string. Defaults to "".
        phone (str, optional): Phone number. Defaults to "".
    """
    
    # If file exists, remove it to ensure we create a fresh file
    import os
    if os.path.exists(filepath):
        try:
            os.remove(filepath)
        except Exception as e:
            # If we can't remove the file (e.g. it's open in Excel), continue and let xlsxwriter handle it
            pass
    
    workbook = xlsxwriter.Workbook(filepath)
    worksheet = workbook.add_worksheet("فاتورة")

    # RTL + Page Break Preview
    worksheet.right_to_left()
    worksheet.set_pagebreak_view()
    worksheet.hide_gridlines(2)  # Hide screen and printed gridlines

    # ================
    #  PAGE SETTINGS
    # ================
    worksheet.set_paper(9)         # A4
    worksheet.set_landscape()
    worksheet.set_margins(0.7, 0.7, 0.75, 0.75)
    worksheet.fit_to_pages(1, 1)   # صفحة واحدة عرض + صفحة واحدة طول

    # ==========================
    #   FORMATS
    # ==========================

    header_fmt = workbook.add_format({
        'bold': True, 'border': 1, 'align': 'center', 'valign': 'vcenter',
        'font_name': 'Arial', 'font_size': 14, 'bg_color': '#2E75B6', 'font_color': '#FFFFFF'
    })

    label_fmt = workbook.add_format({
        'bold': True, 'border': 1,
        'align': 'center', 'valign': 'vcenter',
        'font_name': 'Arial', 'font_size': 12, 'bg_color': '#B4C6E7', 'font_color': '#000000'
    })

    border_fmt = workbook.add_format({
        'border': 1, 'align': 'center', 'valign': 'vcenter',
        'font_name': 'Arial', 'font_size': 10
    })

    # Format for integer numbers (no decimal places)
    integer_fmt = workbook.add_format({
        'border': 1, 'align': 'center', 'valign': 'vcenter',
        'font_name': 'Arial', 'font_size': 10,
        'num_format': '0'  # أعداد صحيحة فقط
    })

    # Format for decimal numbers (length, height - 2 decimal places)
    decimal_fmt = workbook.add_format({
        'border': 1, 'align': 'center', 'valign': 'vcenter',
        'font_name': 'Arial', 'font_size': 10,
        'num_format': '0.00'  # أعداد عشرية بمنزلتين
    })

    # Format for phone numbers (text format to prevent scientific notation)
    phone_fmt = workbook.add_format({
        'border': 1, 'align': 'center', 'valign': 'vcenter',
        'font_name': 'Arial', 'font_size': 10,
        'num_format': '@'  # Text format
    })

    # Format for area (always 2 decimal places, يحافظ على الصفر الأخير)
    area_fmt = workbook.add_format({
        'border': 1, 'align': 'center',
        'font_name': 'Arial', 'font_size': 10,
        'num_format': '0.00'  # يظهر دائمًا منزلتين عشريتين
    })

    # Format for money (integers without decimal places)
    money_fmt = workbook.add_format({
        'border': 1, 'align': 'center',
        'font_name': 'Arial', 'font_size': 10,
        'num_format': '0'
    })



    # ==========================
    #   SPACE
    # ==========================
    worksheet.set_row(0, 12)
    worksheet.set_row(1, 12)

    start_row = 2

    # ==========================
    #   عنوان — Merge
    # ==========================
    worksheet.merge_range(start_row, 3, start_row, 8,
                          f"فاتورة رقم ( {op_num} )", header_fmt)

    # ==========================
    #   جدول العميل / التاريخ
    # ==========================
    r = start_row + 2

    # العميل
    worksheet.write(r, 1, "العميل", label_fmt)
    worksheet.merge_range(r, 2, r, 3, client or "", border_fmt)

    # التاريخ
    worksheet.write(r, 4, "التاريخ", label_fmt)
    worksheet.merge_range(r, 5, r, 6, date_str or "", border_fmt)

    # عدد السيارات - expand label to span two columns, remove merge from value cell
    worksheet.merge_range(r, 7, r, 8, "عدد السيارات", label_fmt)
    worksheet.write(r, 9, "1", integer_fmt)

    r += 1

    # السائق
    worksheet.write(r, 1, "اسم السائق", label_fmt)
    worksheet.merge_range(r, 2, r, 3, driver or "", border_fmt)

    # تليفون
    worksheet.write(r, 4, "ت", label_fmt)
    worksheet.merge_range(r, 5, r, 6, phone or "", phone_fmt)

    # نوع السيارة - expand label to span two columns, remove merge from value cell
    worksheet.merge_range(r, 7, r, 8, "سيارة", label_fmt)
    worksheet.write(r, 9, "", border_fmt)

    r += 2

    # ==========================
    #   جدول الصنف مع المقاسات الفرعية
    # ==========================
    worksheet.merge_range(r, 1, r+1, 1, "البيان", header_fmt)
    worksheet.merge_range(r, 2, r+1, 2, "رقم البلوك", header_fmt)
    worksheet.merge_range(r, 3, r+1, 3, "السمك", header_fmt)
    worksheet.merge_range(r, 4, r+1, 4, "الخامة", header_fmt)
    
    # دمج 3 أعمدةللعنوان العام للمقاس
    worksheet.merge_range(r, 5, r, 7, "المقاس", header_fmt)
    worksheet.write(r+1, 5, "العدد", header_fmt)
    worksheet.write(r+1, 6, "الطول", header_fmt)
    worksheet.write(r+1, 7, "الارتفاع", header_fmt)
    
    worksheet.merge_range(r, 8, r+1, 8, "المسطح م٢", header_fmt)
    worksheet.merge_range(r, 9, r+1, 9, "السعر", header_fmt)
    worksheet.merge_range(r, 10, r+1, 10, "إجمالي السعر", header_fmt)

    first_item_row = None
    r += 2

    # ==========================
    #   العناصر
    # ==========================
    for item in items:
        try:
            desc = item[0]
            block = item[1]
            thickness = item[2]
            material = item[3]
            count = int(float(item[4])) if item[4] else 0
            length = float(item[5]) if item[5] else 0
            height = float(item[6]) if item[6] else 0
            price_val = int(float(item[7])) if item[7] else 0
        except (ValueError, IndexError) as e:
            log_error(f"Error parsing item {item}: {e}")
            continue

        if first_item_row is None:
            first_item_row = r

        worksheet.write(r, 1, "ش " + desc if desc else "", border_fmt)
        worksheet.write(r, 2, block, border_fmt)
        worksheet.write(r, 3, thickness, border_fmt)
        worksheet.write(r, 4, material, border_fmt)
        # القيم توضع مباشرة في الأعمدة الثلاثة
        worksheet.write_number(r, 5, count, integer_fmt)      # العدد
        worksheet.write_number(r, 6, length, decimal_fmt)     # الطول
        worksheet.write_number(r, 7, height, decimal_fmt)     # الارتفاع
        
        # الصيغ يحسبها Excel تلقائياً
        excel_row = r + 1
        worksheet.write_formula(r, 8, f"=F{excel_row}*G{excel_row}*H{excel_row}", area_fmt)
        worksheet.write_number(r, 9, price_val, money_fmt)
        # تقريب الناتج إلى أقرب عدد صحيح
        worksheet.write_formula(r, 10, f"=ROUND(I{excel_row}*J{excel_row},0)", money_fmt)

        r += 1

    # ==========================
    #   المجموع
    # ==========================
    if first_item_row is not None:
        total_row = r
        worksheet.merge_range(total_row, 1, total_row, 4, "المجموع", header_fmt)
        excel_first = first_item_row + 1
        excel_last = r
        
        # Sum for count (column 5)
        worksheet.write(total_row, 5, f"=SUM(F{excel_first}:F{excel_last})", integer_fmt)
        
        # Sum for area (column 8)
        worksheet.write(total_row, 8, f"=SUM(I{excel_first}:I{excel_last})", area_fmt)
        
        # Sum for total price (column 10)
        worksheet.write(total_row, 10, f"=SUM(K{excel_first}:K{excel_last})", money_fmt)

        # ==========================
        #   Aggregated Invoice Summary Table
        # ==========================
        summary_start_row = total_row + 2
        worksheet.merge_range(summary_start_row, 1, summary_start_row, 6, "اجمالي الفاتورة", header_fmt)

        # Write summary table headers
        summary_header_row = summary_start_row + 1
        worksheet.write(summary_header_row, 1, "البيان", header_fmt)
        worksheet.write(summary_header_row, 2, "النوع", header_fmt)
        worksheet.write(summary_header_row, 3, "المسطح م٢", header_fmt)
        worksheet.write(summary_header_row, 4, "إجمالي السعر", header_fmt)
        worksheet.write(summary_header_row, 5, "السمك", header_fmt)
        worksheet.write(summary_header_row, 6, "سعر المتر", header_fmt)

        # Process items to identify unique combinations and calculate totals
        aggregated_data = {}

        for item in items:
            try:
                desc = item[0] or ""
                material = item[3] or ""
                thickness = item[2] or ""
                count = int(float(item[4]))  # Convert to int (count should be whole number)
                length = float(item[5])
                height = float(item[6])
                price_val = float(item[7])
                
                # Calculate area and total for this item
                area = count * length * height
                total = area * price_val
                
                key = (desc, material, thickness)
                if key not in aggregated_data:
                    aggregated_data[key] = {
                        "description": desc,
                        "material": material,
                        "thickness": thickness,
                        "area": area,
                        "total": total,
                        "price_val": price_val
                    }
                else:
                    # Add to existing entry
                    aggregated_data[key]["area"] += area
                    aggregated_data[key]["total"] += total
            except (ValueError, IndexError):
                continue

        # Data rows start right after the header row
        summary_data_start_row = summary_header_row + 1
        row_counter = 0

        for (desc, material, thickness), data in aggregated_data.items():
            row_num = summary_data_start_row + row_counter

            # Static values
            worksheet.write(row_num, 1, "ش " + data["description"] if data["description"] else "", border_fmt)
            worksheet.write(row_num, 2, data["material"], border_fmt)
            worksheet.write(row_num, 5, data["thickness"], border_fmt)

            # Area and total values
            worksheet.write_number(row_num, 3, data["area"], area_fmt)
            worksheet.write_number(row_num, 4, data["total"], money_fmt)

            # Price per meter = total / area (avoid division by zero)
            if data["area"] > 0:
                price_per_meter = data["total"] / data["area"]
                worksheet.write_number(row_num, 6, price_per_meter, money_fmt)
            else:
                worksheet.write(row_num, 6, 0, money_fmt)

            row_counter += 1

        # If no aggregated data was found
        if row_counter == 0:
            worksheet.write(summary_data_start_row, 1, "لا توجد عناصر للتجميع", border_fmt)
            row_counter = 1

        # Summary TOTAL for the aggregated table
        summary_total_row = summary_data_start_row + row_counter
        worksheet.merge_range(summary_total_row, 1, summary_total_row, 2, "المجموع", header_fmt)

        # Total area - sum only the data rows, not the header or total row
        area_sum_formula = f"=SUM(D{summary_data_start_row+1}:D{summary_data_start_row+row_counter})"
        worksheet.write_formula(summary_total_row, 3, area_sum_formula, area_fmt)

        # Total price - sum only the data rows, not the header or total row
        price_sum_formula = f"=SUM(E{summary_data_start_row+1}:E{summary_data_start_row+row_counter})"
        worksheet.write_formula(summary_total_row, 4, price_sum_formula, money_fmt)

        # ==========================
        #   جدول المدفوعات
        # ==========================
        payments_start_row = summary_total_row - 3
        worksheet.merge_range(payments_start_row, 8, payments_start_row, 10, "المدفوعات", header_fmt)

        # Write payments table headers
        payments_header_row = payments_start_row + 1
        worksheet.write(payments_header_row, 8, "المبلغ", header_fmt)
        worksheet.write(payments_header_row, 9, "المدفوع", header_fmt)
        worksheet.write(payments_header_row, 10, "المتبقي", header_fmt)

        # Payments data row
        payments_data_row = payments_header_row + 1

        # Total amount (from invoice total)
        worksheet.write_formula(payments_data_row, 8, f"=E{summary_total_row+1}", money_fmt)

        # Paid amount (leave empty for user to fill)
        worksheet.write(payments_data_row, 9, "", money_fmt)

        # Remaining amount (Amount - Paid)
        worksheet.write_formula(payments_data_row, 10, f"=J{payments_data_row+1}-I{payments_data_row+1}", money_fmt)

        # ==========================
        #   Signature Section
        # ==========================
        signature_row = payments_data_row + 4  # Position after the payments table
        
        # Create a format for centered text (both vertically and horizontally)
        center_fmt = workbook.add_format({
            'align': 'center',
            'valign': 'vcenter',
            'font_name': 'Arial',
            'font_size': 10
        })
        
        # Merge cells for signature section on the right side (columns 8-10, three columns wide) with centering
        worksheet.merge_range(signature_row, 8, signature_row, 10, "التوقيع(_____________)", center_fmt)
        
        # Merge cells for name on the right side (columns 8-10, same position) with centering
        signature_name_row = signature_row + 1
        worksheet.merge_range(signature_name_row, 8, signature_name_row, 10, "أ/ مصطفي السويفي", center_fmt)
        
        # Merge cells for title on the right side (columns 8-10, same position) with centering
        signature_title_row = signature_name_row + 1
        worksheet.merge_range(signature_title_row, 8, signature_title_row, 10, "رئيس مجلس الإدارة", center_fmt)

    # ==========================
    #   عرض الأعمدة
    # ==========================
    worksheet.set_column(0, 0, 3)
    worksheet.set_column(1, 1, 16)  # البيان
    worksheet.set_column(2, 2, 10)  # رقم البلوك
    worksheet.set_column(3, 3, 10)  # السمك
    worksheet.set_column(4, 4, 12)  # الخامة
    worksheet.set_column(5, 5, 8)   # العدد
    worksheet.set_column(6, 6, 8)   # الطول
    worksheet.set_column(7, 7, 8)   # الارتفاع
    worksheet.set_column(8, 8, 10)  # المسطح م٢ / المبلغ (جدول المدفوعات)
    worksheet.set_column(9, 9, 10)  # السعر / المدفوع (جدول المدفوعات)
    worksheet.set_column(10, 10, 14)  # إجمالي السعر / المتبقي (جدول المدفوعات)

    try:
        workbook.close()
    except PermissionError as e:
        log_error(f"Permission error when closing workbook for invoice {op_num}: {e}")
        # Re-raise as a PermissionError so the calling code can handle it
        raise PermissionError("File is currently open in Excel. Please close the file and try again.") from e
    except Exception as e:
        log_error(f"Error when closing workbook for invoice {op_num}: {e}")
        raise


def update_payment_in_invoice(filepath: str, payment_amount: float) -> bool:
    """
    Update the payment amount in an existing invoice Excel file.
    
    Args:
        filepath (str): Path to the invoice Excel file
        payment_amount (float): The payment amount to set
        
    Returns:
        bool: True if successful, False otherwise
    """
    if not os.path.exists(filepath):
        log_error(f"Invoice file does not exist: {filepath}")
        return False
    
    try:
        import openpyxl
        from openpyxl.cell.cell import MergedCell
        
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
        
        if sheet is None:
            log_error("Could not access active sheet in invoice")
            workbook.close()
            return False
        
        # Find the payments table - look for "المدفوعات" header
        payments_row = None
        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value == "المدفوعات":
                    payments_row = row
                    break
            if payments_row:
                break
        
        if not payments_row:
            log_error("Could not find payments table in invoice")
            workbook.close()
            return False
        
        # The payment data row is 2 rows after the header (header + column labels + data)
        # Header row: "المدفوعات"
        # Labels row: "المبلغ" | "المدفوع" | "المتبقي"
        # Data row: value | payment | remaining
        payment_data_row = payments_row + 2
        
        # In xlsxwriter (0-indexed): column 8=I (المبلغ), column 9=J (المدفوع), column 10=K (المتبقي)
        # In openpyxl (1-indexed): column 9=I, column 10=J, column 11=K
        # So "المدفوع" is column 10 in openpyxl
        paid_column = 10  # Column J in openpyxl (1-indexed)
        
        # Update the payment cell
        cell = sheet.cell(row=payment_data_row, column=paid_column)
        if not isinstance(cell, MergedCell):
            cell.value = payment_amount
        else:
            workbook.close()
            return False
        
        # Save the workbook
        workbook.save(filepath)
        workbook.close()
        return True
        
    except PermissionError as e:
        log_error(f"Permission error updating payment in invoice: {e}")
        raise PermissionError("الملف مفتوح حالياً في برنامج Excel. يرجى إغلاق الملف والمحاولة مرة أخرى.")
    except Exception as e:
        log_error(f"Error updating payment in invoice: {e}")
        return False


def get_payment_from_invoice(filepath: str) -> float:
    """
    Read the payment amount from an existing invoice Excel file.
    
    Args:
        filepath (str): Path to the invoice Excel file
        
    Returns:
        float: The payment amount, or 0 if not found
    """
    if not os.path.exists(filepath):
        return 0
    
    try:
        workbook = openpyxl.load_workbook(filepath, data_only=True)
        sheet = workbook.active
        
        # Find the payments table - look for "المدفوعات" header
        payments_row = None
        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value == "المدفوعات":
                    payments_row = row
                    break
            if payments_row:
                break
        
        if not payments_row:
            workbook.close()
            return 0
        
        # The payment data row is 2 rows after the header
        payment_data_row = payments_row + 2
        paid_column = 10  # Column J in openpyxl (1-indexed) - المدفوع
        
        # Read the payment cell
        cell = sheet.cell(row=payment_data_row, column=paid_column)
        payment_value = cell.value
        
        workbook.close()
        
        if payment_value is not None:
            try:
                return float(payment_value)
            except (ValueError, TypeError):
                return 0
        return 0
        
    except Exception as e:
        log_error(f"Error reading payment from invoice: {e}")
        return 0


def update_payment_in_ledger(folder_path: str, op_num: str, payment_amount: float) -> bool:
    """
    Update the payment amount for a specific invoice in the client ledger.
    
    Args:
        folder_path (str): Path to the client's folder
        op_num (str): Invoice number to update
        payment_amount (float): The payment amount to set
        
    Returns:
        bool: True if successful, False otherwise
    """
    filename = "كشف حساب.xlsx"
    filepath = os.path.join(folder_path, filename)
    
    if not os.path.exists(filepath):
        log_error(f"Ledger file does not exist: {filepath}")
        return False
    
    try:
        import openpyxl
        from openpyxl.cell.cell import MergedCell
        
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
        
        if sheet is None:
            log_error("Could not access active sheet in ledger")
            workbook.close()
            return False
        
        # Find the row with the matching invoice number (column 1)
        found_row = None
        for row in range(3, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=1).value
            if str(cell_value) == str(op_num):
                found_row = row
                break
        
        if not found_row:
            workbook.close()
            return False
        
        # Column 9 is "الدفعات" (payments) in the ledger
        # Update the payment cell
        payment_cell = sheet.cell(row=found_row, column=9)
        if not isinstance(payment_cell, MergedCell):
            payment_cell.value = payment_amount
        else:
            # If merged, find the top-left cell of the merge
            for merged_range in sheet.merged_cells.ranges:
                if payment_cell.coordinate in merged_range:
                    top_left = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                    top_left.value = payment_amount
                    break
        
        # Save the workbook
        workbook.save(filepath)
        workbook.close()
        return True
        
    except PermissionError as e:
        log_error(f"Permission error updating payment in ledger: {e}")
        raise PermissionError("ملف كشف الحساب مفتوح حالياً في برنامج Excel. يرجى إغلاق الملف والمحاولة مرة أخرى.")
    except Exception as e:
        log_error(f"Error updating payment in ledger: {e}")
        return False


def delete_existing_invoice_file(filepath: str) -> bool:
    """
    Delete an existing invoice Excel file if it exists.
    
    Args:
        filepath (str): Path to the invoice file to delete
        
    Returns:
        bool: True if file was deleted or didn't exist, False if deletion failed
    """
    import os
    try:
        if os.path.exists(filepath):
            os.remove(filepath)
            return True
        return True  # File doesn't exist, so condition is satisfied
    except Exception as e:
        log_error(f"Error deleting existing invoice file {filepath}: {e}")
        return False


def remove_invoice_from_ledger(folder_path: str, op_num: str):
    """
    Remove an existing invoice entry from the client ledger before updating.
    
    Args:
        folder_path (str): Path to the client's folder
        op_num (str): Invoice number to remove
    """
    import os
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import MergedCell

    filename = f"كشف حساب.xlsx"
    filepath = os.path.join(folder_path, filename)
    if not os.path.exists(filepath):
        return False

    workbook = None
    try:
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
        if sheet is None:
            log_error(f"Could not access active sheet in ledger")
            return False

        # Find the row with the matching invoice number
        found_row = None
        for row in range(3, sheet.max_row + 1):  # Start from row 3 (after headers)
            cell_value = sheet.cell(row=row, column=1).value
            if cell_value == op_num:  # Column 1 is invoice number
                found_row = row
                break
            elif cell_value is not None:
                pass
        if not found_row:
            return False  # Invoice not found in ledger

        # Determine how many rows this invoice entry spans
        # Look for the next row that has a value in column A or the total row
        end_row = found_row
        for row_idx in range(found_row + 1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_idx, column=1).value
            if cell_value == "المجموع":  # This is the total row
                break
            elif cell_value is not None and cell_value != "":  # New invoice starts
                break
            else:
                end_row = row_idx
        # Also need to find the total row to update calculations
        total_row = None
        for row_idx in range(end_row + 1, sheet.max_row + 1):
            if sheet.cell(row=row_idx, column=1).value == "المجموع":
                total_row = row_idx
                break
        
        # Delete the rows for this invoice
        rows_to_delete = end_row - found_row + 1
        if rows_to_delete > 0:
            sheet.delete_rows(found_row, rows_to_delete)
            # If we had a total row, we need to adjust the formulas
            if total_row:
                # Adjust total row index since we deleted rows above it
                new_total_row = total_row - rows_to_delete
                
                # Update formulas in the total row to reflect the removed rows
                # Column E (quantity), F (total price), G (amount), I (payments)
                try:
                    # Protect all cell value assignments
                    try:
                        cell = sheet.cell(row=new_total_row, column=5)
                        if not isinstance(cell, MergedCell):
                            cell.value = f"=SUM(E3:E{new_total_row-1})"
                    except Exception:
                        pass
                    try:
                        cell = sheet.cell(row=new_total_row, column=6)
                        if not isinstance(cell, MergedCell):
                            cell.value = f"=SUM(F3:F{new_total_row-1})"
                    except Exception:
                        pass
                    try:
                        cell = sheet.cell(row=new_total_row, column=7)
                        if not isinstance(cell, MergedCell):
                            cell.value = f"=SUM(G3:G{new_total_row-1})"
                    except Exception:
                        pass
                    try:
                        cell = sheet.cell(row=new_total_row, column=9)
                        if not isinstance(cell, MergedCell):
                            cell.value = f"=SUM(I3:I{new_total_row-1})"
                    except Exception:
                        pass
                    # Update total debt cell
                    try:
                        j1_cell = sheet['J1']
                        if not isinstance(j1_cell, MergedCell):
                            j1_cell.value = f"=G{new_total_row}-I{new_total_row}"
                    except Exception:
                        pass
                except Exception as formula_ex:
                    log_error(f"Error updating formulas: {formula_ex}")

        # Save the updated workbook
        workbook.save(filepath)
        workbook.close()
        return True

    except Exception as e:
        log_error(f"Error removing invoice {op_num} from ledger at {folder_path}: {e}")
        return False
    finally:
        if workbook:
            try:
                workbook.close()
            except Exception:
                pass


def update_invoice_in_ledger(folder_path: str, op_num: str, client_name: str, date_str: str, 
                           total_amount: float, driver: str = "", invoice_items=None):
    """
    Update an existing invoice entry in the client ledger by removing and re-adding it.
    
    This function uses a simpler approach: remove the old invoice entry first,
    then add the new one using the standard update_client_ledger function.
    
    IMPORTANT: invoice_items format should be:
    [(description, material, thickness, area, total_price), ...]
    Where total_price = area * price_per_meter (already calculated)
    
    Args:
        folder_path (str): Path to the client's folder
        op_num (str): Invoice number to update
        client_name (str): Client name
        date_str (str): Date string
        total_amount (float): Total amount for the invoice
        driver (str): Driver name
        invoice_items: List of invoice items with pre-calculated totals
    """
    
    filename = f"كشف حساب.xlsx"
    filepath = os.path.join(folder_path, filename)
    if not os.path.exists(filepath):
        return update_client_ledger(folder_path, client_name, date_str, op_num, total_amount, driver, invoice_items)

    # Step 1: Remove the existing invoice entry from the ledger
    removal_result = remove_invoice_from_ledger(folder_path, op_num)
    # removal_result indicates if invoice was found and removed
    
    # Step 2: Add the updated invoice using the standard function
    return update_client_ledger(folder_path, client_name, date_str, op_num, total_amount, driver, invoice_items)


def update_client_ledger(folder_path: str, client_name: str, date_str: str, op_num: str,
                          total_amount: float, driver: str = "", invoice_items=None):
    import os
    from datetime import datetime
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import MergedCell

    filename = f"كشف حساب.xlsx"
    filepath = os.path.join(folder_path, filename)
    file_exists = os.path.exists(filepath)
    # تجهيز حدود الخلايا (Border)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    workbook = None
    try:
        if file_exists:
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active
            if sheet is None: 
                log_error(f"Could not access active sheet in ledger")
                return (False, "invalid_sheet")

            # Find if there's already a total row and remember its position
            total_row = None
            for row in range(sheet.max_row, 1, -1):
                if sheet.cell(row=row, column=1).value == "المجموع":
                    total_row = row
                    break
            
            # If there's a total row, we need to insert new data before it
            if total_row:
                # Insert new data before the total row
                next_row = total_row
            else:
                # No total row found, append at the end
                next_row = sheet.max_row
                if next_row < 3:  # Make sure we start after headers
                    next_row = 3
                else:
                    next_row += 1
            
            # Unmerge the total row if it exists, to make space for new entries
            if total_row:
                try:
                    sheet.unmerge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
                except:
                    pass  # Might not be merged or other error, continue anyway
                # Delete the old total row to recalculate later
                sheet.delete_rows(total_row, 1)
            
            # Continue with the normal process of adding the invoice data
            num_items = len(invoice_items) if invoice_items else 1
            start_row = next_row
            end_row = next_row + num_items - 1

            # IMPORTANT: Insert new rows to avoid conflicts with existing merged cells
            # This ensures we have fresh, unmerged cells to work with
            sheet.insert_rows(start_row, num_items)

            # Writing rows
            for idx, item in enumerate(invoice_items if invoice_items else [None]):
                current_row = start_row + idx

                # Clean up formatting for the current row
                for col in range(1, 11):
                    try:
                        cell = sheet.cell(row=current_row, column=col)
                        if not isinstance(cell, MergedCell):
                            cell.fill = PatternFill(fill_type=None)
                            cell.font = Font(name='Arial', size=11)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            # Set number format based on column
                            if col == 5:  # الكمية م٢ - مساحة بمنزلتين عشريتين
                                cell.number_format = '0.00'
                            elif col in [6, 7, 9]:  # إجمالي السعر، المبلغ، الدفعات - أعداد صحيحة
                                cell.number_format = '#,##0'
                    except Exception:
                        pass
                # Merged data (written only on the first row)
                if idx == 0:
                    # Set values BEFORE merging to avoid MergedCell issues
                    try:
                        # 1. Invoice number
                        cell1 = sheet.cell(row=start_row, column=1)
                        if not isinstance(cell1, MergedCell):
                            cell1.value = op_num
                        if num_items > 1:
                            sheet.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)

                        # 2. Driver name
                        cell2 = sheet.cell(row=start_row, column=2)
                        if not isinstance(cell2, MergedCell):
                            cell2.value = driver
                        if num_items > 1:
                            sheet.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)

                        # 3. Date
                        cell3 = sheet.cell(row=start_row, column=3)
                        if not isinstance(cell3, MergedCell):
                            cell3.value = date_str
                        if num_items > 1:
                            sheet.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)

                        # 7. Amount (with sum formula)
                        amount_cell = sheet.cell(row=start_row, column=7)
                        if not isinstance(amount_cell, MergedCell):
                            amount_cell.value = f"=SUM(F{start_row}:F{end_row})"
                            amount_cell.font = Font(name='Arial', size=11, bold=True)
                            amount_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                        if num_items > 1:
                            sheet.merge_cells(start_row=start_row, start_column=7, end_row=end_row, end_column=7)

                        # 8, 9, 10 empty columns or payments (merged)
                        cell8 = sheet.cell(row=start_row, column=8)
                        if not isinstance(cell8, MergedCell):
                            cell8.value = ""
                        if num_items > 1:
                            sheet.merge_cells(start_row=start_row, start_column=8, end_row=end_row, end_column=8)

                        cell9 = sheet.cell(row=start_row, column=9)
                        if not isinstance(cell9, MergedCell):
                            cell9.value = 0
                        if num_items > 1:
                            sheet.merge_cells(start_row=start_row, start_column=9, end_row=end_row, end_column=9)

                        cell10 = sheet.cell(row=start_row, column=10)
                        if not isinstance(cell10, MergedCell):
                            cell10.value = ""
                        if num_items > 1:
                            sheet.merge_cells(start_row=start_row, start_column=10, end_row=end_row, end_column=10)
                    except Exception as e:
                        log_error(f"Error setting merged data for invoice {op_num}: {e}")

                # Variable data (non-merged) - item details
                if item:
                    material = item[1] if len(item) > 1 else ""
                    thickness = item[2] if len(item) > 2 else ""
                    area = item[3] if len(item) > 3 else 0
                    price = item[4] if len(item) > 4 else 0

                    try:
                        cell4 = sheet.cell(row=current_row, column=4)
                        if not isinstance(cell4, MergedCell):
                            cell4.value = f"{material} - {thickness}"
                        cell5 = sheet.cell(row=current_row, column=5)
                        if not isinstance(cell5, MergedCell):
                            cell5.value = area
                        cell6 = sheet.cell(row=current_row, column=6)
                        if not isinstance(cell6, MergedCell):
                            cell6.value = price
                    except Exception:
                        pass
                else:
                    try:
                        cell4 = sheet.cell(row=current_row, column=4)
                        if not isinstance(cell4, MergedCell):
                            cell4.value = ""
                        cell5 = sheet.cell(row=current_row, column=5)
                        if not isinstance(cell5, MergedCell):
                            cell5.value = 0
                        cell6 = sheet.cell(row=current_row, column=6)
                        if not isinstance(cell6, MergedCell):
                            cell6.value = total_amount
                    except Exception:
                        pass
                sheet.row_dimensions[current_row].height = 22

            # --- CREATE TOTAL ROW AT THE END ---
            # Find the actual end of data to place the total row
            actual_end_row = end_row
            while actual_end_row < sheet.max_row and sheet.cell(row=actual_end_row+1, column=1).value not in [None, ""]:
                actual_end_row += 1
                if sheet.cell(row=actual_end_row, column=1).value == "المجموع":
                    # If there's another total row, delete it first
                    try:
                        sheet.delete_rows(actual_end_row, 1)
                        actual_end_row -= 1
                    except:
                        break
            
            total_row = actual_end_row + 1  # Position after the last data row
            
            # Set value and format BEFORE merging to avoid MergedCell issues
            try:
                total_label = sheet.cell(row=total_row, column=1)
                total_label.value = "المجموع"
                total_label.font = Font(name='Arial', size=12, bold=True, color="1F4E78")
                total_label.fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
                total_label.alignment = Alignment(horizontal='center', vertical='center')
                total_label.border = thin_border
            except Exception:
                pass
            # Now merge after setting value
            try:
                sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
            except Exception:
                pass
            # Format other cells in the total row
            for col in range(5, 11):
                try:
                    cell = sheet.cell(row=total_row, column=col)
                    if not isinstance(cell, MergedCell):
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.font = Font(name='Arial', size=12, bold=True, color="1F4E78")
                        cell.fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
                        # Set number format based on column
                        if col == 5:  # الكمية م٢ - مساحة بمنزلتين عشريتين
                            cell.number_format = '0.00'
                        elif col in [6, 7, 9]:  # إجمالي السعر، المبلغ، الدفعات - أعداد صحيحة
                            cell.number_format = '#,##0'
                except Exception:
                    pass
            # SUM formulas for the total row - sum all values from row 3 up to the row before the total
            # Column E (quantity area), F (total price), G (amount), I (payments)
            try:
                cell = sheet.cell(row=total_row, column=5)
                if not isinstance(cell, MergedCell):
                    cell.value = f"=SUM(E3:E{total_row-1})"
            except Exception:
                pass
            try:
                cell = sheet.cell(row=total_row, column=6)
                if not isinstance(cell, MergedCell):
                    cell.value = f"=SUM(F3:F{total_row-1})"
            except Exception:
                pass
            try:
                cell = sheet.cell(row=total_row, column=7)
                if not isinstance(cell, MergedCell):
                    cell.value = f"=SUM(G3:G{total_row-1})"
            except Exception:
                pass
            try:
                cell = sheet.cell(row=total_row, column=9)
                if not isinstance(cell, MergedCell):
                    cell.value = f"=SUM(I3:I{total_row-1})"
            except Exception:
                pass
            sheet.row_dimensions[total_row].height = 25

            # --- UPDATE TOTAL DEBT CELL AT THE TOP ---
            # Total debt = sum of amounts (G) - sum of payments (I)
            try:
                j1_cell = sheet['J1']
                if not isinstance(j1_cell, MergedCell):
                    j1_cell.value = f"=G{total_row}-I{total_row}"
            except Exception:
                pass
            # Adjust column widths (if needed)
            column_widths = [15, 15, 15, 25, 13, 16, 16, 16, 13, 22]
            for col, width in enumerate(column_widths, start=1):
                sheet.column_dimensions[get_column_letter(col)].width = width
            workbook.save(filepath)
            workbook.close()
            return (True, None)
        else:
            # Create a new ledger file
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "كشف حساب"
            sheet.sheet_view.rightToLeft = True
            # Page header
            sheet.merge_cells('A1:H1')
            title_cell = sheet['A1']
            title_cell.value = f"كشف حساب العميل / {client_name}"
            title_cell.font = Font(name='Arial', size=18, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            # Total debt header
            try:
                cell = sheet['I1']
                if not isinstance(cell, MergedCell):
                    cell.value = "إجمالي الديون"
                    cell.font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            except Exception as e:
                log_error(f"❌ ERROR setting I1 formatting: {e}", exc_info=True)
            
            try:
                cell = sheet['J1']
                if not isinstance(cell, MergedCell):
                    cell.font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            except Exception as e:
                pass
            sheet.row_dimensions[1].height = 35

            # Table headers
            headers = ["رقم الفاتورة", "اسم السائق", "تاريخ التحميل", "النوع", "الكمية م٢", "إجمالي السعر", "المبلغ", "تاريخ الدفعات", "الدفعات", "ملاحظات"]
            header_fill = PatternFill(start_color="8FAADC", end_color="8FAADC", fill_type="solid")
            
            for col, header in enumerate(headers, start=1):
                try:
                    cell = sheet.cell(row=2, column=col)
                    if not isinstance(cell, MergedCell):
                        cell.value = header
                        cell.font = Font(name='Arial', size=12, bold=True, color="1F4E78")
                        cell.fill = header_fill
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                except Exception as e:
                    pass
            sheet.row_dimensions[2].height = 35
            next_row = 3  # First data row in new file

            # --- Write invoice data ---
            num_items = len(invoice_items) if invoice_items else 1
            start_row = next_row
            end_row = next_row + num_items - 1

            # Write rows
            for idx, item in enumerate(invoice_items if invoice_items else [None]):
                current_row = start_row + idx

                # Clean formatting for current row
                for col in range(1, 11):
                    cell = sheet.cell(row=current_row, column=col)
                    cell.fill = PatternFill(fill_type=None)
                    cell.font = Font(name='Arial', size=11)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    # Set number format based on column
                    if col == 5:  # الكمية م٢ - مساحة بمنزلتين عشريتين
                        cell.number_format = '0.00'
                    elif col in [6, 7, 9]:  # إجمالي السعر، المبلغ، الدفعات - أعداد صحيحة
                        cell.number_format = '#,##0'

                # Merged data (first row only)
                if idx == 0:
                    # Set values BEFORE merging to avoid MergedCell issues
                    # 1. Invoice number
                    sheet.cell(row=start_row, column=1).value = op_num
                    if num_items > 1:
                        sheet.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)

                    # 2. Driver name
                    sheet.cell(row=start_row, column=2).value = driver
                    if num_items > 1:
                        sheet.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)

                    # 3. Date
                    sheet.cell(row=start_row, column=3).value = date_str
                    if num_items > 1:
                        sheet.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)

                    # 7. Amount (with sum formula)
                    amount_cell = sheet.cell(row=start_row, column=7)
                    amount_cell.value = f"=SUM(F{start_row}:F{end_row})"
                    amount_cell.font = Font(name='Arial', size=11, bold=True)
                    amount_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    if num_items > 1:
                        sheet.merge_cells(start_row=start_row, start_column=7, end_row=end_row, end_column=7)

                    # 8, 9, 10 empty columns or payments (merged)
                    sheet.cell(row=start_row, column=8).value = ""
                    if num_items > 1:
                        sheet.merge_cells(start_row=start_row, start_column=8, end_row=end_row, end_column=8)

                    sheet.cell(row=start_row, column=9).value = 0
                    if num_items > 1:
                        sheet.merge_cells(start_row=start_row, start_column=9, end_row=end_row, end_column=9)

                    sheet.cell(row=start_row, column=10).value = ""
                    if num_items > 1:
                        sheet.merge_cells(start_row=start_row, start_column=10, end_row=end_row, end_column=10)

                # Variable data (non-merged) - item details
                if item:
                    material = item[1] if len(item) > 1 else ""
                    thickness = item[2] if len(item) > 2 else ""
                    area = item[3] if len(item) > 3 else 0
                    price = item[4] if len(item) > 4 else 0

                    sheet.cell(row=current_row, column=4, value=f"{material} - {thickness}")
                    sheet.cell(row=current_row, column=5, value=area)
                    sheet.cell(row=current_row, column=6, value=price)
                else:
                    sheet.cell(row=current_row, column=4, value="")
                    sheet.cell(row=current_row, column=5, value=0)
                    sheet.cell(row=current_row, column=6, value=total_amount)

                sheet.row_dimensions[current_row].height = 22

            # --- CREATE TOTAL ROW ---
            total_row = end_row + 1

            # Set value and format BEFORE merging to avoid MergedCell issues
            total_label = sheet.cell(row=total_row, column=1)
            total_label.value = "المجموع"
            total_label.font = Font(name='Arial', size=12, bold=True, color="1F4E78")
            total_label.fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
            total_label.alignment = Alignment(horizontal='center', vertical='center')
            total_label.border = thin_border
            # Now merge after setting value
            sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)

            # Format other cells in total row
            for col in range(5, 11):
                cell = sheet.cell(row=total_row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(name='Arial', size=12, bold=True, color="1F4E78")
                cell.fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
                # Set number format based on column
                if col == 5:  # الكمية م٢ - مساحة بمنزلتين عشريتين
                    cell.number_format = '0.00'
                elif col in [6, 7, 9]:  # إجمالي السعر، المبلغ، الدفعات - أعداد صحيحة
                    cell.number_format = '#,##0'

            # SUM formulas for total row
            try:
                cell = sheet.cell(row=total_row, column=5)
                if not isinstance(cell, MergedCell):
                    cell.value = f"=SUM(E3:E{total_row-1})"
            except Exception as e:
                pass
            try:
                cell = sheet.cell(row=total_row, column=6)
                if not isinstance(cell, MergedCell):
                    cell.value = f"=SUM(F3:F{total_row-1})"
            except Exception as e:
                pass
            try:
                cell = sheet.cell(row=total_row, column=7)
                if not isinstance(cell, MergedCell):
                    cell.value = f"=SUM(G3:G{total_row-1})"
            except Exception as e:
                pass
            try:
                cell = sheet.cell(row=total_row, column=9)
                if not isinstance(cell, MergedCell):
                    cell.value = f"=SUM(I3:I{total_row-1})"
            except Exception as e:
                pass
            sheet.row_dimensions[total_row].height = 25

            # --- UPDATE TOTAL DEBT CELL AT THE TOP ---
            try:
                j1_cell = sheet['J1']
                if not isinstance(j1_cell, MergedCell):
                    j1_cell.value = f"=G{total_row}-I{total_row}"
            except Exception as e:
                pass
            # Adjust column widths
            column_widths = [15, 15, 15, 25, 13, 16, 16, 16, 13, 22]
            for col, width in enumerate(column_widths, start=1):
                sheet.column_dimensions[get_column_letter(col)].width = width
            workbook.save(filepath)
            workbook.close()
            return (True, None)

    except PermissionError:
        return (False, "file_locked")
    except ImportError:
        return (False, "openpyxl_missing")
    except Exception as e:
        return (False, f"error: {str(e)}")
    finally:
        if workbook:
            try:
                workbook.close()
            except Exception as e:
                pass
