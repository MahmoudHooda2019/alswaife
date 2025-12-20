"""
Excel Utilities for Invoice Creation
This module provides functions to generate Excel invoices from invoice data.
"""

import xlsxwriter
from typing import List, Tuple


def save_invoice(filepath: str, op_num: str, client: str, driver: str,
                 items: List[Tuple], date_str: str = "", phone: str = ""):
    """
    Save invoice data to an Excel file with proper formatting.
    
    Args:
        filepath (str): Path to save the Excel file
        op_num (str): Operation/invoice number
        client (str): Client name
        driver (str): Driver name
        items (List[Tuple]): List of invoice items, each tuple contains:
            (description, block, thickness, material, count, length, height, price)
        date_str (str, optional): Date string. Defaults to "".
        phone (str, optional): Phone number. Defaults to "".
    """
    
    workbook = xlsxwriter.Workbook(filepath)
    worksheet = workbook.add_worksheet("فاتورة")

    # RTL + Page Break Preview
    worksheet.right_to_left()
    worksheet.set_pagebreak_view()
    worksheet.hide_gridlines(2)  # Hide screen and printed gridlines

    # ================
    #  PAGE SETTINGS
    # ================
    worksheet.set_paper(9)         # A4
    worksheet.set_landscape()
    worksheet.set_margins(0.7, 0.7, 0.75, 0.75)
    worksheet.fit_to_pages(1, 1)   # صفحة واحدة عرض + صفحة واحدة طول

    # ==========================
    #   FORMATS
    # ==========================

    header_fmt = workbook.add_format({
        'bold': True, 'border': 1, 'align': 'center', 'valign': 'vcenter',
        'font_name': 'Arial', 'font_size': 14, 'bg_color': '#2E75B6', 'font_color': '#FFFFFF'
    })

    label_fmt = workbook.add_format({
        'bold': True, 'border': 1,
        'align': 'center', 'valign': 'vcenter',
        'font_name': 'Arial', 'font_size': 12, 'bg_color': '#B4C6E7', 'font_color': '#000000'
    })

    border_fmt = workbook.add_format({
        'border': 1, 'align': 'center', 'valign': 'vcenter',
        'font_name': 'Arial', 'font_size': 10
    })

    # Format for integer numbers (no decimal places)
    integer_fmt = workbook.add_format({
        'border': 1, 'align': 'center', 'valign': 'vcenter',
        'font_name': 'Arial', 'font_size': 10,
        'num_format': '0'  # أعداد صحيحة فقط
    })

    # Format for decimal numbers (length, height - 2 decimal places)
    decimal_fmt = workbook.add_format({
        'border': 1, 'align': 'center', 'valign': 'vcenter',
        'font_name': 'Arial', 'font_size': 10,
        'num_format': '0.00'  # أعداد عشرية بمنزلتين
    })

    # Format for phone numbers (text format to prevent scientific notation)
    phone_fmt = workbook.add_format({
        'border': 1, 'align': 'center', 'valign': 'vcenter',
        'font_name': 'Arial', 'font_size': 10,
        'num_format': '@'  # Text format
    })

    # Format for area (always 2 decimal places, يحافظ على الصفر الأخير)
    area_fmt = workbook.add_format({
        'border': 1, 'align': 'center',
        'font_name': 'Arial', 'font_size': 10,
        'num_format': '0.00'  # يظهر دائمًا منزلتين عشريتين
    })

    # Format for money (integers without decimal places)
    money_fmt = workbook.add_format({
        'border': 1, 'align': 'center',
        'font_name': 'Arial', 'font_size': 10,
        'num_format': '0'
    })



    # ==========================
    #   SPACE
    # ==========================
    worksheet.set_row(0, 12)
    worksheet.set_row(1, 12)

    start_row = 2

    # ==========================
    #   عنوان — Merge
    # ==========================
    worksheet.merge_range(start_row, 3, start_row, 8,
                          f"فاتورة رقم ( {op_num} )", header_fmt)

    # ==========================
    #   جدول العميل / التاريخ
    # ==========================
    r = start_row + 2

    # العميل
    worksheet.write(r, 1, "العميل", label_fmt)
    worksheet.merge_range(r, 2, r, 3, client or "", border_fmt)

    # التاريخ
    worksheet.write(r, 4, "التاريخ", label_fmt)
    worksheet.merge_range(r, 5, r, 6, date_str or "", border_fmt)

    # عدد السيارات - expand label to span two columns, remove merge from value cell
    worksheet.merge_range(r, 7, r, 8, "عدد السيارات", label_fmt)
    worksheet.write(r, 9, "1", integer_fmt)

    r += 1

    # السائق
    worksheet.write(r, 1, "اسم السائق", label_fmt)
    worksheet.merge_range(r, 2, r, 3, driver or "", border_fmt)

    # تليفون
    worksheet.write(r, 4, "ت", label_fmt)
    worksheet.merge_range(r, 5, r, 6, phone or "", phone_fmt)

    # نوع السيارة - expand label to span two columns, remove merge from value cell
    worksheet.merge_range(r, 7, r, 8, "سيارة", label_fmt)
    worksheet.write(r, 9, "", border_fmt)

    r += 2

    # ==========================
    #   جدول الصنف مع المقاسات الفرعية
    # ==========================
    worksheet.merge_range(r, 1, r+1, 1, "البيان", header_fmt)
    worksheet.merge_range(r, 2, r+1, 2, "رقم البلوك", header_fmt)
    worksheet.merge_range(r, 3, r+1, 3, "السمك", header_fmt)
    worksheet.merge_range(r, 4, r+1, 4, "الخامة", header_fmt)
    
    # دمج 3 أعمدةللعنوان العام للمقاس
    worksheet.merge_range(r, 5, r, 7, "المقاس", header_fmt)
    worksheet.write(r+1, 5, "العدد", header_fmt)
    worksheet.write(r+1, 6, "الطول", header_fmt)
    worksheet.write(r+1, 7, "الارتفاع", header_fmt)
    
    worksheet.merge_range(r, 8, r+1, 8, "المسطح م٢", header_fmt)
    worksheet.merge_range(r, 9, r+1, 9, "السعر", header_fmt)
    worksheet.merge_range(r, 10, r+1, 10, "إجمالي السعر", header_fmt)

    first_item_row = None
    r += 2

    # ==========================
    #   العناصر
    # ==========================
    for item in items:
        try:
            desc = item[0]
            block = item[1]
            thickness = item[2]
            material = item[3]
            count = float(item[4])
            length = float(item[5])
            height = float(item[6])
            price_val =  int(item[7])
        except (ValueError, IndexError):
            continue

        if first_item_row is None:
            first_item_row = r

        worksheet.write(r, 1, "ش " + desc if desc else "", border_fmt)
        worksheet.write(r, 2, block, border_fmt)
        worksheet.write(r, 3, thickness, border_fmt)
        worksheet.write(r, 4, material, border_fmt)
        # القيم توضع مباشرة في الأعمدة الثلاثة
        worksheet.write_number(r, 5, count, integer_fmt)      # العدد
        worksheet.write_number(r, 6, length, decimal_fmt)     # الطول
        worksheet.write_number(r, 7, height, decimal_fmt)     # الارتفاع
        
        # الصيغ يحسبها Excel تلقائياً
        excel_row = r + 1
        worksheet.write_formula(r, 8, f"=F{excel_row}*G{excel_row}*H{excel_row}", area_fmt)
        worksheet.write_number(r, 9, price_val, money_fmt)
        # تقريب الناتج إلى أقرب عدد صحيح
        worksheet.write_formula(r, 10, f"=ROUND(I{excel_row}*J{excel_row},0)", money_fmt)

        r += 1

    # ==========================
    #   المجموع
    # ==========================
    if first_item_row is not None:
        total_row = r
        worksheet.merge_range(total_row, 1, total_row, 4, "المجموع", header_fmt)
        excel_first = first_item_row + 1
        excel_last = r
        
        # Sum for count (column 5)
        worksheet.write(total_row, 5, f"=SUM(F{excel_first}:F{excel_last})", integer_fmt)
        
        # Sum for area (column 8)
        worksheet.write(total_row, 8, f"=SUM(I{excel_first}:I{excel_last})", area_fmt)
        
        # Sum for total price (column 10)
        worksheet.write(total_row, 10, f"=SUM(K{excel_first}:K{excel_last})", money_fmt)

        # ==========================
        #   Aggregated Invoice Summary Table
        # ==========================
        summary_start_row = total_row + 2
        worksheet.merge_range(summary_start_row, 1, summary_start_row, 6, "اجمالي الفاتورة", header_fmt)

        # Write summary table headers
        summary_header_row = summary_start_row + 1
        worksheet.write(summary_header_row, 1, "البيان", header_fmt)
        worksheet.write(summary_header_row, 2, "النوع", header_fmt)
        worksheet.write(summary_header_row, 3, "المسطح م٢", header_fmt)
        worksheet.write(summary_header_row, 4, "إجمالي السعر", header_fmt)
        worksheet.write(summary_header_row, 5, "السمك", header_fmt)
        worksheet.write(summary_header_row, 6, "سعر المتر", header_fmt)

        # Process items to identify unique combinations and calculate totals
        aggregated_data = {}

        for item in items:
            try:
                desc = item[0] or ""
                material = item[3] or ""
                thickness = item[2] or ""
                count = float(item[4])
                length = float(item[5])
                height = float(item[6])
                price_val = float(item[7])
                
                # Calculate area and total for this item
                area = count * length * height
                total = area * price_val
                
                key = (desc, material, thickness)
                if key not in aggregated_data:
                    aggregated_data[key] = {
                        "description": desc,
                        "material": material,
                        "thickness": thickness,
                        "area": area,
                        "total": total,
                        "price_val": price_val
                    }
                else:
                    # Add to existing entry
                    aggregated_data[key]["area"] += area
                    aggregated_data[key]["total"] += total
            except (ValueError, IndexError):
                continue

        # Data rows start right after the header row
        summary_data_start_row = summary_header_row + 1
        row_counter = 0

        for (desc, material, thickness), data in aggregated_data.items():
            row_num = summary_data_start_row + row_counter

            # Static values
            worksheet.write(row_num, 1, "ش " + data["description"] if data["description"] else "", border_fmt)
            worksheet.write(row_num, 2, data["material"], border_fmt)
            worksheet.write(row_num, 5, data["thickness"], border_fmt)

            # Area and total values
            worksheet.write_number(row_num, 3, data["area"], area_fmt)
            worksheet.write_number(row_num, 4, data["total"], money_fmt)

            # Price per meter = total / area (avoid division by zero)
            if data["area"] > 0:
                price_per_meter = data["total"] / data["area"]
                worksheet.write_number(row_num, 6, price_per_meter, money_fmt)
            else:
                worksheet.write(row_num, 6, 0, money_fmt)

            row_counter += 1

        # If no aggregated data was found
        if row_counter == 0:
            worksheet.write(summary_data_start_row, 1, "لا توجد عناصر للتجميع", border_fmt)
            row_counter = 1

        # Summary TOTAL for the aggregated table
        summary_total_row = summary_data_start_row + row_counter
        worksheet.merge_range(summary_total_row, 1, summary_total_row, 2, "المجموع", header_fmt)

        # Total area - sum only the data rows, not the header or total row
        # Data rows in 1-indexed Excel notation: from (summary_data_start_row+1) to (summary_total_row)
        # But we don't want to include the total row itself, so we go to (summary_total_row-1+1) = summary_total_row
        # Wait, that's confusing. Let's be explicit:
        # First data row (0-indexed): summary_data_start_row
        # Last data row (0-indexed): summary_data_start_row + row_counter - 1
        # First data row (1-indexed): summary_data_start_row + 1
        # Last data row (1-indexed): summary_data_start_row + row_counter - 1 + 1 = summary_data_start_row + row_counter
        # Total row (1-indexed): summary_total_row + 1
        # So we want =SUM(D[first_data_row_1indexed]:D[last_data_row_1indexed])
        # =SUM(D{summary_data_start_row+1}:D{summary_data_start_row+row_counter})
        area_sum_formula = f"=SUM(D{summary_data_start_row+1}:D{summary_data_start_row+row_counter})"
        worksheet.write_formula(summary_total_row, 3, area_sum_formula, area_fmt)

        # Total price - sum only the data rows, not the header or total row
        # Same logic as above
        price_sum_formula = f"=SUM(E{summary_data_start_row+1}:E{summary_data_start_row+row_counter})"
        worksheet.write_formula(summary_total_row, 4, price_sum_formula, money_fmt)

        # ==========================
        #   جدول المدفوعات
        # ==========================
        payments_start_row = summary_total_row - 3
        worksheet.merge_range(payments_start_row, 8, payments_start_row, 10, "المدفوعات", header_fmt)

        # Write payments table headers
        payments_header_row = payments_start_row + 1
        worksheet.write(payments_header_row, 8, "المبلغ", header_fmt)
        worksheet.write(payments_header_row, 9, "المدفوع", header_fmt)
        worksheet.write(payments_header_row, 10, "المتبقي", header_fmt)

        # Payments data row
        payments_data_row = payments_header_row + 1

        # Total amount (from invoice total)
        worksheet.write_formula(payments_data_row, 8, f"=E{summary_total_row+1}", money_fmt)

        # Paid amount (leave empty for user to fill)
        worksheet.write(payments_data_row, 9, "", money_fmt)

        # Remaining amount (Amount - Paid)
        worksheet.write_formula(payments_data_row, 10, f"=J{payments_data_row+1}-I{payments_data_row+1}", money_fmt)

        # ==========================
        #   Signature Section
        # ==========================
        signature_row = payments_data_row + 4  # Position after the payments table
        
        # Create a format for centered text (both vertically and horizontally)
        center_fmt = workbook.add_format({
            'align': 'center',
            'valign': 'vcenter',
            'font_name': 'Arial',
            'font_size': 10
        })
        
        # Merge cells for signature section on the right side (columns 8-10, three columns wide) with centering
        worksheet.merge_range(signature_row, 8, signature_row, 10, "التوقيع(_____________)", center_fmt)
        
        # Merge cells for name on the right side (columns 8-10, same position) with centering
        signature_name_row = signature_row + 1
        worksheet.merge_range(signature_name_row, 8, signature_name_row, 10, "أ/ مصطفي السويفي", center_fmt)
        
        # Merge cells for title on the right side (columns 8-10, same position) with centering
        signature_title_row = signature_name_row + 1
        worksheet.merge_range(signature_title_row, 8, signature_title_row, 10, "رئيس مجلس الإدارة", center_fmt)

    # ==========================
    #   عرض الأعمدة
    # ==========================
    worksheet.set_column(0, 0, 3)
    worksheet.set_column(1, 1, 16)  # البيان
    worksheet.set_column(2, 2, 10)  # رقم البلوك
    worksheet.set_column(3, 3, 10)  # السمك
    worksheet.set_column(4, 4, 12)  # الخامة
    worksheet.set_column(5, 5, 8)   # العدد
    worksheet.set_column(6, 6, 8)   # الطول
    worksheet.set_column(7, 7, 8)   # الارتفاع
    worksheet.set_column(8, 8, 10)  # المسطح م٢ / المبلغ (جدول المدفوعات)
    worksheet.set_column(9, 9, 10)  # السعر / المدفوع (جدول المدفوعات)
    worksheet.set_column(10, 10, 14)  # إجمالي السعر / المتبقي (جدول المدفوعات)

    try:
        workbook.close()
    except PermissionError as e:
        # Re-raise as a PermissionError so the calling code can handle it
        raise PermissionError("File is currently open in Excel. Please close the file and try again.") from e

def update_client_ledger(folder_path: str, client_name: str, date_str: str, op_num: str, 
                         total_amount: float, driver: str = "", invoice_items = None):
    """
    Update or create the client's cumulative ledger Excel file with detailed information.
    
    Args:
        folder_path (str): The directory where the ledger should be saved.
        client_name (str): The name of the client.
        date_str (str): The date of the invoice.
        op_num (str): The invoice number.
        total_amount (float): The total amount of the invoice.
        driver (str): Driver name.
        invoice_items (list): List of aggregated invoice items (desc, material, thickness, area, price)
    
    Returns:
        tuple: (success: bool, error_message: str or None)
    """
    import os
    from datetime import datetime
    
    filename = f"{client_name}.xlsx"
    filepath = os.path.join(folder_path, filename)
    
    # Check if file exists to determine if we need to create headers
    file_exists = os.path.exists(filepath)
    
    try:
        if file_exists:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active
            
                    # Check if sheet is valid
            if sheet is None:
                return (False, "invalid_sheet")
            
            # Find next empty row (skip header and total debt rows)
            next_row = sheet.max_row + 1
            
            # Add invoice items to the ledger
            if invoice_items:
                for item_data in invoice_items:
                    desc, material, thickness, area, price = item_data
                    
                    sheet.cell(row=next_row, column=1, value=op_num)
                    sheet.cell(row=next_row, column=2, value=driver)
                    sheet.cell(row=next_row, column=3, value=date_str)
                    sheet.cell(row=next_row, column=4, value=f"{material} - {thickness}")
                    sheet.cell(row=next_row, column=5, value=area)
                    sheet.cell(row=next_row, column=6, value=price)
                    sheet.cell(row=next_row, column=7, value=area * price)  # المبلغ
                    sheet.cell(row=next_row, column=8, value="")  # تاريخ الدفع
                    sheet.cell(row=next_row, column=9, value=0)  # الدفعات
                    
                    # الرصيد = المبلغ - الدفعات (معادلة)
                    sheet.cell(row=next_row, column=10, value=f"=G{next_row}-I{next_row}")
                    
                    sheet.cell(row=next_row, column=11, value="")  # ملاحظات
                    
                    # Apply borders
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    for col in range(1, 12):
                        cell = sheet.cell(row=next_row, column=col)
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    next_row += 1
            
            # Update total debt formula in cell I1 (merged cells I1:K1)
            # إجمالي الديون = مجموع عمود الرصيد (J)
            # The debt formula should be consistent with initial creation
            try:
                debt_value = sheet['I1']  # This represents the merged range I1:K1
                debt_value.value = f"=SUM(J4:J{next_row-1})"
            except:
                # Fallback: try to update just cell I1
                sheet.cell(row=1, column=9, value=f"=SUM(J4:J{next_row-1})")
            
            workbook.save(filepath)
            return (True, None)
            
        else:
            # Create new file with detailed structure
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
            
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            
            # Check if sheet is valid before using it
            if sheet is None:
                return (False, "invalid_sheet")
                
            sheet.title = "كشف حساب"
            
            # RTL
            sheet.sheet_view.rightToLeft = True
            
            # Title: كشف الحساب - [Client Name]
            sheet.merge_cells('A1:F1')  # Reduce merge range to make space for debt info
            title_cell = sheet['A1']
            title_cell.value = f"كشف حساب العميل / {client_name}"
            title_cell.font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            sheet.row_dimensions[1].height = 30
            
            # Merge empty row 2 for better visual separation
            sheet.merge_cells('A2:K2')
            empty_row = sheet['A2']
            empty_row.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            
            # Total debt - placing it next to client name instead of below
            sheet.merge_cells('G1:H1')  # Merge cells G1 and H1 for debt label
            debt_label = sheet['G1']
            debt_label.value = "إجمالي الديون:"
            debt_label.font = Font(name='Arial', size=12, bold=True)
            # Improved color scheme to match sheet design
            debt_label.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            debt_label.font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
            debt_label.alignment = Alignment(horizontal='center', vertical='center')
            
            # Formula for total debt
            sheet.merge_cells('I1:K1')  # Merge cells I1, J1, K1 for debt value
            debt_value = sheet['I1']
            # Calculate the last row for the SUM formula
            # Data starts at row 4, and we add len(invoice_items) rows
            last_row = 4 + len(invoice_items) - 1 if invoice_items else 4
            debt_value.value = f"=SUM(J4:J{last_row})"
            debt_value.font = Font(name='Arial', size=12, bold=True, color="4472C4")
            # Improved color scheme to match sheet design
            debt_value.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            debt_value.alignment = Alignment(horizontal='center', vertical='center')
            debt_value.number_format = '#,##0'
            
            # Headers
            headers = [
                "رقم الفاتورة",
                "اسم السائق", 
                "تاريخ التحميل",
                "النوع (الخامة)",
                "المسطح م٢",
                "إجمالي السعر",
                "المبلغ",
                "تاريخ الدفع",
                "الدفعات",
                "الرصيد",
                "ملاحظات"
            ]
            
            header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            header_font = Font(name='Arial', size=11, bold=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Set headers for row 3
            # Note: We avoid setting values for cells that might be part of merged ranges
            # In this case, row 3 has no merged cells, so we can safely set all values
            for col, header in enumerate(headers, start=1):
                cell = sheet.cell(row=3, column=col)
                # Using setattr to avoid static analysis issues with merged cells
                setattr(cell, 'value', header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            sheet.row_dimensions[3].height = 30
            
            # Set column widths
            column_widths = [12, 12, 12, 15, 10, 12, 12, 12, 10, 10, 15]
            for col, width in enumerate(column_widths, start=1):
                from openpyxl.utils import get_column_letter
                sheet.column_dimensions[get_column_letter(col)].width = width
            
            # Add first invoice data
            row = 4
            if invoice_items:
                for item_data in invoice_items:
                    desc, material, thickness, area, price = item_data
                    
                    sheet.cell(row=row, column=1, value=op_num)
                    sheet.cell(row=row, column=2, value=driver)
                    sheet.cell(row=row, column=3, value=date_str)
                    sheet.cell(row=row, column=4, value=f"{material} - {thickness}")
                    sheet.cell(row=row, column=5, value=area)
                    sheet.cell(row=row, column=6, value=price)
                    sheet.cell(row=row, column=7, value=area * price)
                    sheet.cell(row=row, column=8, value="")
                    sheet.cell(row=row, column=9, value=0)
                    
                    # الرصيد = المبلغ - الدفعات (معادلة)
                    sheet.cell(row=row, column=10, value=f"=G{row}-I{row}")
                    
                    sheet.cell(row=row, column=11, value="")
                    
                    # Apply borders and alignment
                    for col in range(1, 12):
                        cell = sheet.cell(row=row, column=col)
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    row += 1
            
            workbook.save(filepath)
            return (True, None)
            
    except PermissionError as e:
        # File is open in Excel
        return (False, "file_locked")
    except ImportError:
        return (False, "openpyxl_missing")
    except Exception as e:
        return (False, f"error: {str(e)}")
