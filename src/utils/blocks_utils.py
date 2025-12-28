import xlsxwriter
from typing import List, Dict
import os
import openpyxl
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter

from utils.log_utils import log_error, log_exception

# Table column definitions
TABLE1_COLUMNS = [
    "رقم النقله", "عدد النقله", "التاريخ", "المحجر", 
    "رقم البلوك", "الخامه", "الطول", 
    "العرض", "الارتفاع", "م3", "الوزن", 
    "وزن البلوك", "سعر الطن", "اجمالي السعر"
]

# Slides table columns (starts after TABLE1 + 1 gap column)
TABLE2_COLUMNS = [
    "تاريخ النشر", "رقم البلوك", "النوع", "رقم المكينه",
    "وقت الدخول", "وقت الخروج", "عدد الساعات",
    "السمك", "العدد", "الطول", "الخصم", "الطول بعد",
    "الارتفاع", "الكمية م2", "سعر المتر", "اجمالي السعر"
]

# Wastage and productivity sheet columns
TABLE3_COLUMNS = [
    "رقم البلوك", "الطول قبل", "العرض قبل", "الارتفاع قبل", "م3",
    "السمك بعد", "الانتاج الفعلي م2", "معدل الانتاجيه", "الاجمالي", "الفرق (الهالك)"
]

# Column width definitions for each table
TABLE1_WIDTH = [12, 10, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12]
TABLE2_WIDTH = [12, 12, 12, 10, 20, 20, 12, 8, 8, 10, 8, 12, 10, 12, 12, 14]
TABLE3_WIDTH = [12, 10, 10, 10, 12, 10, 12, 14, 12, 14]

# Starting column for slides table (directly after TABLE1, no gap)
SLIDES_START_COL = len(TABLE1_COLUMNS) + 1  # +1 to start right after blocks table

# الفضل table columns (same structure as slides)
TABLE_FADL_COLUMNS = [
    "تاريخ النشر", "رقم البلوك", "النوع", "رقم المكينه",
    "وقت الدخول", "وقت الخروج", "عدد الساعات",
    "السمك", "العدد", "الطول", "الخصم", "الطول بعد",
    "الارتفاع", "الكمية م2", "سعر المتر", "اجمالي السعر"
]

TABLE_FADL_WIDTH = [12, 12, 12, 10, 20, 20, 12, 8, 8, 10, 8, 12, 10, 12, 12, 14]

# Productivity rates based on thickness
PRODUCTIVITY_RATES = {
    "2سم": 29,
    "3سم": 21.6,
    "4سم": 15
}

def export_simple_blocks_excel(rows: List[Dict]) -> str:
    """إنشاء أو تحديث ملف Excel لحساب مخزون البلوكات"""
    documents_folder = os.path.join(
        os.path.expanduser("~"), "Documents", "alswaife", "البلوكات"
    )
    os.makedirs(documents_folder, exist_ok=True)
    
    filepath = os.path.join(documents_folder, "مخزون البلوكات.xlsx")
    
    if os.path.exists(filepath):
        result = append_to_existing_file(filepath, rows)
        if result is False:  # File is locked
            raise PermissionError("File is currently open in Excel. Please close the file and try again.")
    else:
        result = create_new_excel_file(filepath, rows)
        if result is False:  # File is locked
            raise PermissionError("File is currently open in Excel. Please close the file and try again.")
    
    # Update wastage sheet
    update_wastage_sheet(filepath)
    
    return filepath


def export_slides_to_blocks_excel(slides_data: List[Dict]) -> str:
    """إضافة بيانات الشرائح إلى ملف البلوكات في جدول منفصل"""
    documents_folder = os.path.join(
        os.path.expanduser("~"), "Documents", "alswaife", "البلوكات"
    )
    os.makedirs(documents_folder, exist_ok=True)
    
    filepath = os.path.join(documents_folder, "مخزون البلوكات.xlsx")
    
    if os.path.exists(filepath):
        result = append_slides_to_existing_file(filepath, slides_data)
        if result is False:
            raise PermissionError("File is currently open in Excel. Please close the file and try again.")
        # Update wastage sheet after adding slides
        update_wastage_sheet(filepath)
    else:
        # Create new file with slides table structure
        result = create_new_excel_file_with_slides(filepath, [], slides_data)
        if result is False:
            raise PermissionError("File is currently open in Excel. Please close the file and try again.")
    
    return filepath


def append_slides_to_existing_file(filepath: str, slides_data: List[Dict]):
    """
    إضافة بيانات الشرائح إلى ملف موجود
    الشرائح A تُسجل في الصف الأول للبلوك
    الشرائح B تُسجل في الصف الثاني للبلوك
    الشرائح F تُسجل في شيت الفضل (شيت منفصل)
    """
    try:
        workbook = openpyxl.load_workbook(filepath)
        worksheet = workbook["البلوكات"]
        
        # Define styles
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Slides table colors (different from blocks)
        slides_colors = [
            "C9FFE4",  # Light Aqua
            "FFDFBA",  # Light Orange
            "E4BAFF",  # Light Purple
            "BAE1FF",  # Light Blue
            "FFFFBA",  # Light Yellow
            "FFC9DE",  # Light Pink
            "BAFFC9",  # Light Mint Green
            "FFB3BA",  # Light Red-Pink
            "C9D6FF",  # Light Indigo
            "FFE8C9",  # Light Peach
            "D6FFC9",  # Light Lime
            "FFC9F3",  # Light Magenta
            "C9FFF3",  # Light Turquoise
            "F3C9FF",  # Light Lavender
            "BAFFC9",  # Light Mint Green
            "FFB3BA",  # Light Red-Pink
        ]
        
        # الفضل table colors (different shade)
        fadl_colors = [
            "FFE4C9",  # Light Peach
            "C9E4FF",  # Light Sky Blue
            "E4C9FF",  # Light Violet
            "C9FFE4",  # Light Mint
            "FFE4E4",  # Light Rose
            "E4FFE4",  # Light Green
            "E4E4FF",  # Light Periwinkle
            "FFF0C9",  # Light Gold
            "F0C9FF",  # Light Orchid
            "C9FFF0",  # Light Cyan
            "FFC9E4",  # Light Pink
            "E4FFC9",  # Light Lime
            "C9E4E4",  # Light Teal
            "E4C9E4",  # Light Mauve
            "FFE4F0",  # Light Blush
            "F0FFE4",  # Light Honeydew
        ]
        
        header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        title_fill = PatternFill(start_color="9933FF", end_color="9933FF", fill_type="solid")
        title_font = Font(color="FFFFFF", bold=True, size=14)
        
        # الفضل header styles (different color)
        fadl_header_fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
        fadl_title_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
        
        gap_fill = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
        
        # Check if slides headers exist, if not add them
        slides_title_cell = worksheet.cell(row=1, column=SLIDES_START_COL)
        if not slides_title_cell.value:
            # Add slides title
            worksheet.merge_cells(start_row=1, start_column=SLIDES_START_COL, 
                                end_row=1, end_column=SLIDES_START_COL + len(TABLE2_COLUMNS) - 1)
            title_cell = worksheet.cell(row=1, column=SLIDES_START_COL, value="الشرائح")
            title_cell.font = title_font
            title_cell.fill = title_fill
            title_cell.alignment = center_alignment
            title_cell.border = thin_border
            
            # Add slides headers
            for idx, col in enumerate(TABLE2_COLUMNS):
                cell = worksheet.cell(row=2, column=SLIDES_START_COL + idx, value=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            
            # Set column widths for slides
            for i, width in enumerate(TABLE2_WIDTH):
                worksheet.column_dimensions[get_column_letter(SLIDES_START_COL + i)].width = width
        
        # Create or get الفضل sheet (separate sheet)
        if "الفضل" in workbook.sheetnames:
            fadl_sheet = workbook["الفضل"]
        else:
            fadl_sheet = workbook.create_sheet("الفضل")
            fadl_sheet.sheet_view.rightToLeft = True
            
            # Add الفضل title
            fadl_sheet.merge_cells(start_row=1, start_column=1, 
                                end_row=1, end_column=len(TABLE_FADL_COLUMNS))
            title_cell = fadl_sheet.cell(row=1, column=1, value="الفضل")
            title_cell.font = title_font
            title_cell.fill = fadl_title_fill
            title_cell.alignment = center_alignment
            title_cell.border = thin_border
            fadl_sheet.row_dimensions[1].height = 30
            
            # Add الفضل headers
            for idx, col in enumerate(TABLE_FADL_COLUMNS):
                cell = fadl_sheet.cell(row=2, column=1 + idx, value=col)
                cell.font = header_font
                cell.fill = fadl_header_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            fadl_sheet.row_dimensions[2].height = 25
            
            # Set column widths for الفضل
            for i, width in enumerate(TABLE_FADL_WIDTH):
                fadl_sheet.column_dimensions[get_column_letter(1 + i)].width = width
            
            # Freeze panes
            fadl_sheet.freeze_panes = 'A3'
        
        # Build a map of block numbers to their row positions
        # Column 5 (E) contains block numbers in the blocks table
        block_row_map = {}  # {block_number: first_row_of_block}
        for row in range(3, worksheet.max_row + 1):
            block_cell = worksheet.cell(row=row, column=5)
            if block_cell.value:
                block_num = str(block_cell.value).strip()
                if block_num and block_num not in block_row_map:
                    block_row_map[block_num] = row
        
        # Default discount
        DEFAULT_DISCOUNT = 0.20
        
        # Separate slides data into A/B and F
        ab_slides = []
        f_slides = []
        
        for slide_data in slides_data:
            slide_block_number = str(slide_data.get('block_number', '')).strip().upper()
            if slide_block_number.startswith('F'):
                f_slides.append(slide_data)
            else:
                ab_slides.append(slide_data)
        
        # Add A/B slides data to الشرائح table
        for slide_data in ab_slides:
            # Get slide block number (e.g., "A11" or "B11")
            slide_block_number = str(slide_data.get('block_number', '')).strip().upper()
            
            if not slide_block_number:
                continue
            
            # Determine if it's A or B and extract the base block number
            is_side_a = slide_block_number.startswith('A')
            is_side_b = slide_block_number.startswith('B')
            
            if is_side_a:
                base_block_number = slide_block_number[1:]  # Remove 'A' prefix
                row_offset = 0  # A goes in first row
            elif is_side_b:
                base_block_number = slide_block_number[1:]  # Remove 'B' prefix
                row_offset = 1  # B goes in second row
            else:
                # If no A/B prefix, treat as A
                base_block_number = slide_block_number
                row_offset = 0
            
            # Find the block row
            if base_block_number not in block_row_map:
                continue
            
            block_first_row = block_row_map[base_block_number]
            excel_row = block_first_row + row_offset
            
            # Get values
            publishing_date = slide_data.get('publishing_date', '')
            material = slide_data.get('material', '')
            machine_number = slide_data.get('machine_number', '')
            entry_time = slide_data.get('entry_time', '')
            exit_time = slide_data.get('exit_time', '')
            # Add RTL mark (U+200F) for proper Arabic text display in Excel
            if entry_time:
                entry_time = '\u200F' + entry_time
            if exit_time:
                exit_time = '\u200F' + exit_time
            hours_count = float(slide_data.get('hours_count', 0)) if slide_data.get('hours_count') else 0
            thickness = slide_data.get('thickness', '')
            quantity = int(slide_data.get('quantity', 0)) if slide_data.get('quantity') else 0
            length = float(slide_data.get('length', 0)) if slide_data.get('length') else 0
            height = float(slide_data.get('height', 0)) if slide_data.get('height') else 0
            price_per_meter = float(slide_data.get('price_per_meter', 0)) if slide_data.get('price_per_meter') else 0
            
            # Column data (single row, no merge)
            col_data = [
                (publishing_date, False),
                (slide_block_number, False),  # Keep full block number (A11 or B11)
                (material, False),
                (machine_number, False),
                (entry_time, False),
                (exit_time, False),
                (hours_count, False),
                (thickness, False),
                (quantity, False),
                (length, False),
                (DEFAULT_DISCOUNT, False),
                (f"={get_column_letter(SLIDES_START_COL + 9)}{excel_row}-{get_column_letter(SLIDES_START_COL + 10)}{excel_row}", True),  # الطول بعد الخصم
                (height, False),
                (f"={get_column_letter(SLIDES_START_COL + 11)}{excel_row}*{get_column_letter(SLIDES_START_COL + 12)}{excel_row}*{get_column_letter(SLIDES_START_COL + 8)}{excel_row}", True),  # الكمية م2
                (price_per_meter, False),
                (f"={get_column_letter(SLIDES_START_COL + 13)}{excel_row}*{get_column_letter(SLIDES_START_COL + 14)}{excel_row}", True),  # اجمالي السعر
            ]
            
            for col_idx, (value, is_formula) in enumerate(col_data):
                col_num = SLIDES_START_COL + col_idx
                
                # Single row cell (no merge)
                cell = worksheet.cell(row=excel_row, column=col_num, value=value)
                cell.border = thin_border
                cell.alignment = center_alignment
                cell.fill = PatternFill(start_color=slides_colors[col_idx], 
                                       end_color=slides_colors[col_idx], fill_type="solid")
                
                # Number format for numeric columns
                if col_idx in [6, 9, 10, 11, 12, 13, 14, 15]:
                    cell.number_format = '#,##0.00'
                elif col_idx == 8:
                    cell.number_format = '#,##0'
        
        # Add F slides data to الفضل sheet (separate sheet)
        # Find the next available row for الفضل data
        fadl_next_row = 3  # Start after headers
        for row in range(3, fadl_sheet.max_row + 1):
            fadl_cell = fadl_sheet.cell(row=row, column=1)
            if fadl_cell.value:
                fadl_next_row = row + 1
        
        for slide_data in f_slides:
            slide_block_number = str(slide_data.get('block_number', '')).strip().upper()
            
            if not slide_block_number:
                continue
            
            excel_row = fadl_next_row
            fadl_next_row += 1
            
            # Get values
            publishing_date = slide_data.get('publishing_date', '')
            material = slide_data.get('material', '')
            machine_number = slide_data.get('machine_number', '')
            entry_time = slide_data.get('entry_time', '')
            exit_time = slide_data.get('exit_time', '')
            # Add RTL mark (U+200F) for proper Arabic text display in Excel
            if entry_time:
                entry_time = '\u200F' + entry_time
            if exit_time:
                exit_time = '\u200F' + exit_time
            hours_count = float(slide_data.get('hours_count', 0)) if slide_data.get('hours_count') else 0
            thickness = slide_data.get('thickness', '')
            quantity = int(slide_data.get('quantity', 0)) if slide_data.get('quantity') else 0
            length = float(slide_data.get('length', 0)) if slide_data.get('length') else 0
            height = float(slide_data.get('height', 0)) if slide_data.get('height') else 0
            price_per_meter = float(slide_data.get('price_per_meter', 0)) if slide_data.get('price_per_meter') else 0
            
            # Column data for الفضل (starts from column 1 in separate sheet)
            col_data = [
                (publishing_date, False),
                (slide_block_number, False),  # Keep full block number (F11)
                (material, False),
                (machine_number, False),
                (entry_time, False),
                (exit_time, False),
                (hours_count, False),
                (thickness, False),
                (quantity, False),
                (length, False),
                (DEFAULT_DISCOUNT, False),
                (f"={get_column_letter(10)}{excel_row}-{get_column_letter(11)}{excel_row}", True),  # الطول بعد الخصم
                (height, False),
                (f"={get_column_letter(12)}{excel_row}*{get_column_letter(13)}{excel_row}*{get_column_letter(9)}{excel_row}", True),  # الكمية م2
                (price_per_meter, False),
                (f"={get_column_letter(14)}{excel_row}*{get_column_letter(15)}{excel_row}", True),  # اجمالي السعر
            ]
            
            for col_idx, (value, is_formula) in enumerate(col_data):
                col_num = 1 + col_idx  # Start from column 1 in separate sheet
                
                cell = fadl_sheet.cell(row=excel_row, column=col_num, value=value)
                cell.border = thin_border
                cell.alignment = center_alignment
                cell.fill = PatternFill(start_color=fadl_colors[col_idx], 
                                       end_color=fadl_colors[col_idx], fill_type="solid")
                
                # Number format for numeric columns
                if col_idx in [6, 9, 10, 11, 12, 13, 14, 15]:
                    cell.number_format = '#,##0.00'
                elif col_idx == 8:
                    cell.number_format = '#,##0'
        
        workbook.save(filepath)
        return True
        
    except PermissionError as e:
        log_error(f"File is locked: {e}")
        return False
    except Exception as e:
        log_exception(f"Error adding slides data: {e}")
        return False


def create_new_excel_file_with_slides(filepath: str, blocks_rows: List[Dict], slides_rows: List[Dict]):
    """إنشاء ملف جديد مع جدولي البلوكات والشرائح وشيت الفضل المنفصل"""
    from openpyxl import Workbook as OpenpyxlWorkbook
    
    workbook = OpenpyxlWorkbook()
    worksheet = workbook.active
    worksheet.title = "البلوكات"
    worksheet.sheet_view.rightToLeft = True
    
    # Define styles
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Header styles for blocks
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    
    # Header styles for slides
    slides_header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    slides_title_fill = PatternFill(start_color="9933FF", end_color="9933FF", fill_type="solid")
    
    # Header styles for الفضل
    fadl_header_fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
    fadl_title_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    
    gap_fill = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
    
    # === BLOCKS TABLE ===
    # Write blocks title
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(TABLE1_COLUMNS))
    title_cell = worksheet.cell(row=1, column=1, value="مقاس البلوك علي الارضية")
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = center_alignment
    title_cell.border = thin_border
    worksheet.row_dimensions[1].height = 30
    
    # Write blocks headers
    for idx, col in enumerate(TABLE1_COLUMNS, 1):
        cell = worksheet.cell(row=2, column=idx, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    worksheet.row_dimensions[2].height = 25
    
    # Set blocks column widths
    for i, width in enumerate(TABLE1_WIDTH, 1):
        worksheet.column_dimensions[get_column_letter(i)].width = width
    
    # === SLIDES TABLE ===
    # Write slides title
    worksheet.merge_cells(start_row=1, start_column=SLIDES_START_COL, 
                         end_row=1, end_column=SLIDES_START_COL + len(TABLE2_COLUMNS) - 1)
    title_cell = worksheet.cell(row=1, column=SLIDES_START_COL, value="الشرائح")
    title_cell.font = title_font
    title_cell.fill = slides_title_fill
    title_cell.alignment = center_alignment
    title_cell.border = thin_border
    
    # Write slides headers
    for idx, col in enumerate(TABLE2_COLUMNS):
        cell = worksheet.cell(row=2, column=SLIDES_START_COL + idx, value=col)
        cell.font = header_font
        cell.fill = slides_header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Set slides column widths
    for i, width in enumerate(TABLE2_WIDTH):
        worksheet.column_dimensions[get_column_letter(SLIDES_START_COL + i)].width = width
    
    # === الفضل SHEET (separate sheet) ===
    fadl_sheet = workbook.create_sheet("الفضل")
    fadl_sheet.sheet_view.rightToLeft = True
    
    # Write الفضل title
    fadl_sheet.merge_cells(start_row=1, start_column=1, 
                         end_row=1, end_column=len(TABLE_FADL_COLUMNS))
    title_cell = fadl_sheet.cell(row=1, column=1, value="الفضل")
    title_cell.font = title_font
    title_cell.fill = fadl_title_fill
    title_cell.alignment = center_alignment
    title_cell.border = thin_border
    fadl_sheet.row_dimensions[1].height = 30
    
    # Write الفضل headers
    for idx, col in enumerate(TABLE_FADL_COLUMNS):
        cell = fadl_sheet.cell(row=2, column=1 + idx, value=col)
        cell.font = header_font
        cell.fill = fadl_header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    fadl_sheet.row_dimensions[2].height = 25
    
    # Set الفضل column widths
    for i, width in enumerate(TABLE_FADL_WIDTH):
        fadl_sheet.column_dimensions[get_column_letter(1 + i)].width = width
    
    # Freeze panes for الفضل sheet
    fadl_sheet.freeze_panes = 'A3'
    
    # Freeze panes for main sheet
    worksheet.freeze_panes = 'A3'
    
    try:
        workbook.save(filepath)
        
        # Add slides data if provided
        if slides_rows:
            append_slides_to_existing_file(filepath, slides_rows)
        
        return True
    except PermissionError as e:
        log_error(f"File is locked: {e}")
        return False


def append_to_existing_file(filepath: str, new_rows: List[Dict]):
    """
    Append new rows to an existing Excel file.
    
    This function adds new block data to an existing Excel file while maintaining
    the existing structure and formatting. It handles one table with specific
    column arrangements and formulas.
    
    Args:
        filepath (str): Path to the existing Excel file
        new_rows (List[Dict]): List of dictionaries containing block data
    
    Returns:
        bool: True if successful, False if file is locked
    """
    try:
        workbook = openpyxl.load_workbook(filepath)
        worksheet = workbook["البلوكات"]
        
        # Starting row for new data (after existing data)
        start_row = worksheet.max_row + 1
        
        # Define styles
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        gap_fill = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")  # Gray for gaps
        # Custom style for headers
        header_fill_table1 = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")  # Dark blue for Table 1 headers
        header_font = Font(color="FFFFFF", bold=True)  # White bold font for headers
        
        # Define different colors for columns with improved contrast
        column_colors = [
            "FFB3BA",  # Light Red-Pink
            "BAFFC9",  # Light Mint Green
            "BAE1FF",  # Light Blue
            "FFFFBA",  # Light Yellow
            "FFDFBA",  # Light Orange
            "E4BAFF",  # Light Purple
            "FFC9DE",  # Light Pink
            "C9FFE4",  # Light Aqua
            "C9D6FF",  # Light Indigo
            "FFE8C9",  # Light Peach
            "D6FFC9",  # Light Lime
            "FFC9F3",  # Light Magenta
            "C9FFF3",  # Light Turquoise
            "F3C9FF",  # Light Lavender
        ]
        

        # Add new data - each block takes 2 rows (merged vertically)
        for i, block_data in enumerate(new_rows):
            excel_row = start_row + (i * 2)  # Each block takes 2 rows
            
            # --- الجدول الأول ---
            # Write values and set formulas for calculated fields
            # All cells will be merged vertically (2 rows)
            
            # Column 1: رقم النقله (merged 2 rows)
            worksheet.merge_cells(start_row=excel_row, start_column=1, end_row=excel_row+1, end_column=1)
            cell = worksheet.cell(row=excel_row, column=1, value=block_data.get("trip_number", ""))
            cell.border = thin_border
            cell.alignment = center_alignment
            cell.fill = PatternFill(start_color=column_colors[0], end_color=column_colors[0], fill_type="solid")
            # Apply border to merged cell bottom
            worksheet.cell(row=excel_row+1, column=1).border = thin_border
            
            # Column 2: عدد النقله (merged 2 rows)
            trip_count = block_data.get("trip_count", "")
            if trip_count:
                try:
                    trip_count = int(float(trip_count))
                except (ValueError, TypeError):
                    trip_count = ""
            worksheet.merge_cells(start_row=excel_row, start_column=2, end_row=excel_row+1, end_column=2)
            cell = worksheet.cell(row=excel_row, column=2, value=trip_count)
            cell.border = thin_border
            cell.alignment = center_alignment
            cell.fill = PatternFill(start_color=column_colors[1], end_color=column_colors[1], fill_type="solid")
            worksheet.cell(row=excel_row+1, column=2).border = thin_border
            
            # Column 3: التاريخ (merged 2 rows)
            worksheet.merge_cells(start_row=excel_row, start_column=3, end_row=excel_row+1, end_column=3)
            cell = worksheet.cell(row=excel_row, column=3, value=block_data.get("date", ""))
            cell.border = thin_border
            cell.alignment = center_alignment
            cell.fill = PatternFill(start_color=column_colors[2], end_color=column_colors[2], fill_type="solid")
            worksheet.cell(row=excel_row+1, column=3).border = thin_border
            
            # Column 4: المحجر (merged 2 rows)
            worksheet.merge_cells(start_row=excel_row, start_column=4, end_row=excel_row+1, end_column=4)
            cell = worksheet.cell(row=excel_row, column=4, value=block_data.get("quarry", ""))
            cell.border = thin_border
            cell.alignment = center_alignment
            cell.fill = PatternFill(start_color=column_colors[3], end_color=column_colors[3], fill_type="solid")
            worksheet.cell(row=excel_row+1, column=4).border = thin_border
            
            # Column 5: رقم البلوك (merged 2 rows)
            worksheet.merge_cells(start_row=excel_row, start_column=5, end_row=excel_row+1, end_column=5)
            cell = worksheet.cell(row=excel_row, column=5, value=block_data.get("block_number", ""))
            cell.border = thin_border
            cell.alignment = center_alignment
            cell.fill = PatternFill(start_color=column_colors[4], end_color=column_colors[4], fill_type="solid")
            worksheet.cell(row=excel_row+1, column=5).border = thin_border
            
            # Column 6: الخامه (merged 2 rows)
            worksheet.merge_cells(start_row=excel_row, start_column=6, end_row=excel_row+1, end_column=6)
            cell = worksheet.cell(row=excel_row, column=6, value=block_data.get("material", ""))
            cell.border = thin_border
            cell.alignment = center_alignment
            cell.fill = PatternFill(start_color=column_colors[5], end_color=column_colors[5], fill_type="solid")
            worksheet.cell(row=excel_row+1, column=6).border = thin_border
            
            # Column 7: الطول (merged 2 rows)
            worksheet.merge_cells(start_row=excel_row, start_column=7, end_row=excel_row+1, end_column=7)
            cell = worksheet.cell(row=excel_row, column=7, value=block_data.get("length", ""))
            cell.border = thin_border
            cell.alignment = center_alignment
            cell.fill = PatternFill(start_color=column_colors[6], end_color=column_colors[6], fill_type="solid")
            worksheet.cell(row=excel_row+1, column=7).border = thin_border
            
            # Column 8: العرض (merged 2 rows)
            worksheet.merge_cells(start_row=excel_row, start_column=8, end_row=excel_row+1, end_column=8)
            cell = worksheet.cell(row=excel_row, column=8, value=block_data.get("width", ""))
            cell.border = thin_border
            cell.alignment = center_alignment
            cell.fill = PatternFill(start_color=column_colors[7], end_color=column_colors[7], fill_type="solid")
            worksheet.cell(row=excel_row+1, column=8).border = thin_border
            
            # Column 9: الارتفاع (merged 2 rows)
            worksheet.merge_cells(start_row=excel_row, start_column=9, end_row=excel_row+1, end_column=9)
            cell = worksheet.cell(row=excel_row, column=9, value=block_data.get("height", ""))
            cell.border = thin_border
            cell.alignment = center_alignment
            cell.fill = PatternFill(start_color=column_colors[8], end_color=column_colors[8], fill_type="solid")
            worksheet.cell(row=excel_row+1, column=9).border = thin_border
            
            # Column 10: م3 (volume) - Calculate using formula (merged 2 rows)
            length_cell = get_column_letter(7)
            width_cell = get_column_letter(8)
            height_cell = get_column_letter(9)
            formula_str = f'={length_cell}{excel_row}*{width_cell}{excel_row}*{height_cell}{excel_row}'
            worksheet.merge_cells(start_row=excel_row, start_column=10, end_row=excel_row+1, end_column=10)
            cell = worksheet.cell(row=excel_row, column=10)
            cell.value = formula_str
            cell.data_type = 'f'
            cell.border = thin_border
            cell.alignment = center_alignment
            cell.fill = PatternFill(start_color=column_colors[9], end_color=column_colors[9], fill_type="solid")
            worksheet.cell(row=excel_row+1, column=10).border = thin_border
            
            # Column 11: الوزن (merged 2 rows)
            worksheet.merge_cells(start_row=excel_row, start_column=11, end_row=excel_row+1, end_column=11)
            cell = worksheet.cell(row=excel_row, column=11, value=block_data.get("weight_per_m3", ""))
            cell.border = thin_border
            cell.alignment = center_alignment
            cell.fill = PatternFill(start_color=column_colors[10], end_color=column_colors[10], fill_type="solid")
            worksheet.cell(row=excel_row+1, column=11).border = thin_border
            
            # Column 12: وزن البلوك - Calculate using formula (merged 2 rows)
            volume_cell = get_column_letter(10)
            weight_per_m3_cell = get_column_letter(11)
            formula_str = f'={volume_cell}{excel_row}*{weight_per_m3_cell}{excel_row}'
            worksheet.merge_cells(start_row=excel_row, start_column=12, end_row=excel_row+1, end_column=12)
            cell = worksheet.cell(row=excel_row, column=12)
            cell.value = formula_str
            cell.data_type = 'f'
            cell.border = thin_border
            cell.alignment = center_alignment
            cell.fill = PatternFill(start_color=column_colors[11], end_color=column_colors[11], fill_type="solid")
            worksheet.cell(row=excel_row+1, column=12).border = thin_border
            
            # Column 13: سعر الطن (merged 2 rows)
            worksheet.merge_cells(start_row=excel_row, start_column=13, end_row=excel_row+1, end_column=13)
            cell = worksheet.cell(row=excel_row, column=13, value=block_data.get("price_per_ton", ""))
            cell.border = thin_border
            cell.alignment = center_alignment
            cell.fill = PatternFill(start_color=column_colors[12], end_color=column_colors[12], fill_type="solid")
            worksheet.cell(row=excel_row+1, column=13).border = thin_border
            
            # Column 14: اجمالي السعر - Calculate using formula (merged 2 rows)
            block_weight_cell = get_column_letter(12)
            price_per_ton_cell = get_column_letter(13)
            formula_str = f'={block_weight_cell}{excel_row}*{price_per_ton_cell}{excel_row}'
            worksheet.merge_cells(start_row=excel_row, start_column=14, end_row=excel_row+1, end_column=14)
            cell = worksheet.cell(row=excel_row, column=14)
            cell.value = formula_str
            cell.data_type = 'f'
            cell.border = thin_border
            cell.alignment = center_alignment
            cell.fill = PatternFill(start_color=column_colors[13], end_color=column_colors[13], fill_type="solid")
            worksheet.cell(row=excel_row+1, column=14).border = thin_border
            
            
        workbook.save(filepath)
        
    except PermissionError as e:
        log_error(f"File is locked: {e}")
        return False  # File is locked
    except Exception as e:
        log_error(f"Error adding data: {e}")
        # In case of severe error, can try to recreate, but cautiously
        # create_new_excel_file(filepath, new_rows)
    return True

def create_new_excel_file(filepath: str, rows: List[Dict]):
    """
    Create a new Excel file with block data.
    
    This function creates a new Excel file with one table containing block data.
    It sets up the proper structure, formatting, and formulas for the table.
    Each block takes 2 rows (merged vertically).
    
    Args:
        filepath (str): Path where the new Excel file should be created
        rows (List[Dict]): List of dictionaries containing block data
    
    Returns:
        bool: True if successful, False if file is locked
    """
    # Use openpyxl instead of xlsxwriter for better merge support
    from openpyxl import Workbook as OpenpyxlWorkbook
    
    workbook = OpenpyxlWorkbook()
    worksheet = workbook.active
    worksheet.title = "البلوكات"
    worksheet.sheet_view.rightToLeft = True

    # Define styles
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # تعريف ألوان مختلفة للأعمدة مع تحسين التباين
    column_colors = [
        "FFB3BA",  # Light Red-Pink
        "BAFFC9",  # Light Mint Green
        "BAE1FF",  # Light Blue
        "FFFFBA",  # Light Yellow
        "FFDFBA",  # Light Orange
        "E4BAFF",  # Light Purple
        "FFC9DE",  # Light Pink
        "C9FFE4",  # Light Aqua
        "C9D6FF",  # Light Indigo
        "FFE8C9",  # Light Peach
        "D6FFC9",  # Light Lime
        "FFC9F3",  # Light Magenta
        "C9FFF3",  # Light Turquoise
        "F3C9FF",  # Light Lavender
    ]

    # Header styles
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(color="FFFFFF", bold=True, size=14)

    # Write title (merged across all columns)
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(TABLE1_COLUMNS))
    title_cell = worksheet.cell(row=1, column=1, value="مقاس البلوك علي الارضية")
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = center_alignment
    title_cell.border = thin_border
    worksheet.row_dimensions[1].height = 30

    # Write column headers
    for idx, col in enumerate(TABLE1_COLUMNS, 1):
        cell = worksheet.cell(row=2, column=idx, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    worksheet.row_dimensions[2].height = 25

    start_row = 3
    
    for i, block_data in enumerate(rows):
        excel_row = start_row + (i * 2)  # Each block takes 2 rows
        
        # Column 1: رقم النقله (merged 2 rows)
        worksheet.merge_cells(start_row=excel_row, start_column=1, end_row=excel_row+1, end_column=1)
        cell = worksheet.cell(row=excel_row, column=1, value=block_data.get("trip_number", ""))
        cell.border = thin_border
        cell.alignment = center_alignment
        cell.fill = PatternFill(start_color=column_colors[0], end_color=column_colors[0], fill_type="solid")
        worksheet.cell(row=excel_row+1, column=1).border = thin_border
        
        # Column 2: عدد النقله (merged 2 rows)
        trip_count = block_data.get("trip_count", "")
        if trip_count:
            try:
                trip_count = int(float(trip_count))
            except (ValueError, TypeError):
                trip_count = ""
        worksheet.merge_cells(start_row=excel_row, start_column=2, end_row=excel_row+1, end_column=2)
        cell = worksheet.cell(row=excel_row, column=2, value=trip_count)
        cell.border = thin_border
        cell.alignment = center_alignment
        cell.fill = PatternFill(start_color=column_colors[1], end_color=column_colors[1], fill_type="solid")
        worksheet.cell(row=excel_row+1, column=2).border = thin_border
        
        # Column 3: التاريخ (merged 2 rows)
        worksheet.merge_cells(start_row=excel_row, start_column=3, end_row=excel_row+1, end_column=3)
        cell = worksheet.cell(row=excel_row, column=3, value=block_data.get("date", ""))
        cell.border = thin_border
        cell.alignment = center_alignment
        cell.fill = PatternFill(start_color=column_colors[2], end_color=column_colors[2], fill_type="solid")
        worksheet.cell(row=excel_row+1, column=3).border = thin_border
        
        # Column 4: المحجر (merged 2 rows)
        worksheet.merge_cells(start_row=excel_row, start_column=4, end_row=excel_row+1, end_column=4)
        cell = worksheet.cell(row=excel_row, column=4, value=block_data.get("quarry", ""))
        cell.border = thin_border
        cell.alignment = center_alignment
        cell.fill = PatternFill(start_color=column_colors[3], end_color=column_colors[3], fill_type="solid")
        worksheet.cell(row=excel_row+1, column=4).border = thin_border
        
        # Column 5: رقم البلوك (merged 2 rows)
        worksheet.merge_cells(start_row=excel_row, start_column=5, end_row=excel_row+1, end_column=5)
        cell = worksheet.cell(row=excel_row, column=5, value=block_data.get("block_number", ""))
        cell.border = thin_border
        cell.alignment = center_alignment
        cell.fill = PatternFill(start_color=column_colors[4], end_color=column_colors[4], fill_type="solid")
        worksheet.cell(row=excel_row+1, column=5).border = thin_border
        
        # Column 6: الخامه (merged 2 rows)
        worksheet.merge_cells(start_row=excel_row, start_column=6, end_row=excel_row+1, end_column=6)
        cell = worksheet.cell(row=excel_row, column=6, value=block_data.get("material", ""))
        cell.border = thin_border
        cell.alignment = center_alignment
        cell.fill = PatternFill(start_color=column_colors[5], end_color=column_colors[5], fill_type="solid")
        worksheet.cell(row=excel_row+1, column=6).border = thin_border
        
        # Column 7: الطول (merged 2 rows)
        worksheet.merge_cells(start_row=excel_row, start_column=7, end_row=excel_row+1, end_column=7)
        cell = worksheet.cell(row=excel_row, column=7, value=block_data.get("length", ""))
        cell.border = thin_border
        cell.alignment = center_alignment
        cell.fill = PatternFill(start_color=column_colors[6], end_color=column_colors[6], fill_type="solid")
        worksheet.cell(row=excel_row+1, column=7).border = thin_border
        
        # Column 8: العرض (merged 2 rows)
        worksheet.merge_cells(start_row=excel_row, start_column=8, end_row=excel_row+1, end_column=8)
        cell = worksheet.cell(row=excel_row, column=8, value=block_data.get("width", ""))
        cell.border = thin_border
        cell.alignment = center_alignment
        cell.fill = PatternFill(start_color=column_colors[7], end_color=column_colors[7], fill_type="solid")
        worksheet.cell(row=excel_row+1, column=8).border = thin_border
        
        # Column 9: الارتفاع (merged 2 rows)
        worksheet.merge_cells(start_row=excel_row, start_column=9, end_row=excel_row+1, end_column=9)
        cell = worksheet.cell(row=excel_row, column=9, value=block_data.get("height", ""))
        cell.border = thin_border
        cell.alignment = center_alignment
        cell.fill = PatternFill(start_color=column_colors[8], end_color=column_colors[8], fill_type="solid")
        worksheet.cell(row=excel_row+1, column=9).border = thin_border
        
        # Column 10: م3 (volume) - Formula (merged 2 rows)
        length_col = get_column_letter(7)
        width_col = get_column_letter(8)
        height_col = get_column_letter(9)
        volume_formula = f'={length_col}{excel_row}*{width_col}{excel_row}*{height_col}{excel_row}'
        worksheet.merge_cells(start_row=excel_row, start_column=10, end_row=excel_row+1, end_column=10)
        cell = worksheet.cell(row=excel_row, column=10, value=volume_formula)
        cell.border = thin_border
        cell.alignment = center_alignment
        cell.fill = PatternFill(start_color=column_colors[9], end_color=column_colors[9], fill_type="solid")
        worksheet.cell(row=excel_row+1, column=10).border = thin_border
        
        # Column 11: الوزن (merged 2 rows)
        worksheet.merge_cells(start_row=excel_row, start_column=11, end_row=excel_row+1, end_column=11)
        cell = worksheet.cell(row=excel_row, column=11, value=block_data.get("weight_per_m3", ""))
        cell.border = thin_border
        cell.alignment = center_alignment
        cell.fill = PatternFill(start_color=column_colors[10], end_color=column_colors[10], fill_type="solid")
        worksheet.cell(row=excel_row+1, column=11).border = thin_border
        
        # Column 12: وزن البلوك - Formula (merged 2 rows)
        volume_col = get_column_letter(10)
        weight_per_m3_col = get_column_letter(11)
        weight_formula = f'={volume_col}{excel_row}*{weight_per_m3_col}{excel_row}'
        worksheet.merge_cells(start_row=excel_row, start_column=12, end_row=excel_row+1, end_column=12)
        cell = worksheet.cell(row=excel_row, column=12, value=weight_formula)
        cell.border = thin_border
        cell.alignment = center_alignment
        cell.fill = PatternFill(start_color=column_colors[11], end_color=column_colors[11], fill_type="solid")
        worksheet.cell(row=excel_row+1, column=12).border = thin_border
        
        # Column 13: سعر الطن (merged 2 rows)
        worksheet.merge_cells(start_row=excel_row, start_column=13, end_row=excel_row+1, end_column=13)
        cell = worksheet.cell(row=excel_row, column=13, value=block_data.get("price_per_ton", ""))
        cell.border = thin_border
        cell.alignment = center_alignment
        cell.fill = PatternFill(start_color=column_colors[12], end_color=column_colors[12], fill_type="solid")
        worksheet.cell(row=excel_row+1, column=13).border = thin_border
        
        # Column 14: اجمالي السعر - Formula (merged 2 rows)
        block_weight_col = get_column_letter(12)
        price_per_ton_col = get_column_letter(13)
        total_price_formula = f'={block_weight_col}{excel_row}*{price_per_ton_col}{excel_row}'
        worksheet.merge_cells(start_row=excel_row, start_column=14, end_row=excel_row+1, end_column=14)
        cell = worksheet.cell(row=excel_row, column=14, value=total_price_formula)
        cell.border = thin_border
        cell.alignment = center_alignment
        cell.fill = PatternFill(start_color=column_colors[13], end_color=column_colors[13], fill_type="solid")
        worksheet.cell(row=excel_row+1, column=14).border = thin_border
    
    # Set column widths
    for i, width in enumerate(TABLE1_WIDTH, 1):
        worksheet.column_dimensions[get_column_letter(i)].width = width
    
    # === ADD SLIDES TABLE ===
    gap_fill = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
    slides_header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    slides_title_fill = PatternFill(start_color="9933FF", end_color="9933FF", fill_type="solid")
    
    # Slides title
    worksheet.merge_cells(start_row=1, start_column=SLIDES_START_COL, 
                         end_row=1, end_column=SLIDES_START_COL + len(TABLE2_COLUMNS) - 1)
    slides_title_cell = worksheet.cell(row=1, column=SLIDES_START_COL, value="الشرائح")
    slides_title_cell.font = title_font
    slides_title_cell.fill = slides_title_fill
    slides_title_cell.alignment = center_alignment
    slides_title_cell.border = thin_border
    
    # Slides headers
    for idx, col in enumerate(TABLE2_COLUMNS):
        cell = worksheet.cell(row=2, column=SLIDES_START_COL + idx, value=col)
        cell.font = header_font
        cell.fill = slides_header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Set slides column widths
    for i, width in enumerate(TABLE2_WIDTH):
        worksheet.column_dimensions[get_column_letter(SLIDES_START_COL + i)].width = width
    
    # Freeze panes
    worksheet.freeze_panes = 'A3'
    
    try:
        workbook.save(filepath)
        return True
    except PermissionError as e:
        log_error(f"File is locked: {e}")
        return False


def update_wastage_sheet(filepath: str):
    """
    إنشاء أو تحديث شيت الهالك والانتاجية
    يأخذ البيانات من شيت البلوكات ويحسب الهالك والانتاجية
    """
    try:
        workbook = openpyxl.load_workbook(filepath)
        
        # Create or get wastage sheet
        if "هالك وانتاجيه" in workbook.sheetnames:
            wastage_sheet = workbook["هالك وانتاجيه"]
            # Clear existing data (keep headers)
            for row in range(2, wastage_sheet.max_row + 1):
                for col in range(1, len(TABLE3_COLUMNS) + 1):
                    wastage_sheet.cell(row=row, column=col).value = None
        else:
            wastage_sheet = workbook.create_sheet("هالك وانتاجيه")
            wastage_sheet.sheet_view.rightToLeft = True
        
        # Define styles
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Header styles
        header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")  # Dark red
        header_font = Font(color="FFFFFF", bold=True, size=12)
        title_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
        title_font = Font(color="FFFFFF", bold=True, size=14)
        
        # Data colors
        data_colors = [
            "FFE6E6",  # Light red
            "FFE6CC",  # Light orange
            "FFFFCC",  # Light yellow
            "E6FFE6",  # Light green
            "E6FFFF",  # Light cyan
            "E6E6FF",  # Light blue
            "FFE6FF",  # Light magenta
            "FFCCCC",  # Salmon
            "CCFFCC",  # Mint
            "FFCCFF",  # Pink
        ]
        
        # Write title
        wastage_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(TABLE3_COLUMNS))
        title_cell = wastage_sheet.cell(row=1, column=1, value="هالك وانتاجيه")
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = center_alignment
        title_cell.border = thin_border
        wastage_sheet.row_dimensions[1].height = 30
        
        # Write headers
        for idx, col in enumerate(TABLE3_COLUMNS, 1):
            cell = wastage_sheet.cell(row=2, column=idx, value=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        wastage_sheet.row_dimensions[2].height = 25
        
        # Set column widths
        for i, width in enumerate(TABLE3_WIDTH, 1):
            wastage_sheet.column_dimensions[get_column_letter(i)].width = width
        
        # Get data from blocks sheet
        blocks_sheet = workbook["البلوكات"]
        
        # Collect data: find slides with their corresponding block data
        wastage_data = []
        
        # Iterate through blocks sheet to find slides data
        for row in range(3, blocks_sheet.max_row + 1):
            # Check if there's slide data in this row (column SLIDES_START_COL + 1 = رقم البلوك الشرائح)
            slide_block_num = blocks_sheet.cell(row=row, column=SLIDES_START_COL + 1).value
            
            if slide_block_num:
                # Get block data from blocks table
                # Column 5 = رقم البلوك, Column 7 = الطول, Column 8 = العرض, Column 9 = الارتفاع, Column 10 = م3
                block_number = slide_block_num
                length = blocks_sheet.cell(row=row, column=7).value  # الطول من البلوك
                width = blocks_sheet.cell(row=row, column=8).value   # العرض من البلوك
                height = blocks_sheet.cell(row=row, column=9).value  # الارتفاع من البلوك
                
                # Get slides data
                # SLIDES_START_COL + 7 = السمك, SLIDES_START_COL + 13 = الكمية م2
                thickness = blocks_sheet.cell(row=row, column=SLIDES_START_COL + 7).value  # السمك
                area_m2 = blocks_sheet.cell(row=row, column=SLIDES_START_COL + 13).value   # الكمية م2
                
                if thickness and area_m2:
                    wastage_data.append({
                        'row': row,
                        'block_number': block_number,
                        'length': length,
                        'width': width,
                        'height': height,
                        'thickness': thickness,
                        'area_m2': area_m2
                    })
        
        # Write data to wastage sheet with formulas
        for i, data in enumerate(wastage_data):
            excel_row = 3 + i
            source_row = data['row']
            
            # Column 1: رقم البلوك - reference from slides
            cell = wastage_sheet.cell(row=excel_row, column=1)
            cell.value = f"=البلوكات!{get_column_letter(SLIDES_START_COL + 1)}{source_row}"
            cell.border = thin_border
            cell.alignment = center_alignment
            cell.fill = PatternFill(start_color=data_colors[0], end_color=data_colors[0], fill_type="solid")
            
            # Column 2: الطول - reference from blocks
            cell = wastage_sheet.cell(row=excel_row, column=2)
            cell.value = f"=البلوكات!G{source_row}"
            cell.border = thin_border
            cell.alignment = center_alignment
            cell.fill = PatternFill(start_color=data_colors[1], end_color=data_colors[1], fill_type="solid")
            cell.number_format = '#,##0.00'
            
            # Column 3: العرض - reference from blocks
            cell = wastage_sheet.cell(row=excel_row, column=3)
            cell.value = f"=البلوكات!H{source_row}"
            cell.border = thin_border
            cell.alignment = center_alignment
            cell.fill = PatternFill(start_color=data_colors[2], end_color=data_colors[2], fill_type="solid")
            cell.number_format = '#,##0.00'
            
            # Column 4: الارتفاع - reference from blocks
            cell = wastage_sheet.cell(row=excel_row, column=4)
            cell.value = f"=البلوكات!I{source_row}"
            cell.border = thin_border
            cell.alignment = center_alignment
            cell.fill = PatternFill(start_color=data_colors[3], end_color=data_colors[3], fill_type="solid")
            cell.number_format = '#,##0.00'
            
            # Column 5: م3 - reference from blocks (formula)
            cell = wastage_sheet.cell(row=excel_row, column=5)
            cell.value = f"=البلوكات!J{source_row}"
            cell.border = thin_border
            cell.alignment = center_alignment
            cell.fill = PatternFill(start_color=data_colors[4], end_color=data_colors[4], fill_type="solid")
            cell.number_format = '#,##0.00'
            
            # Column 6: السمك - reference from slides
            thickness_col = get_column_letter(SLIDES_START_COL + 7)
            cell = wastage_sheet.cell(row=excel_row, column=6)
            cell.value = f"=البلوكات!{thickness_col}{source_row}"
            cell.border = thin_border
            cell.alignment = center_alignment
            cell.fill = PatternFill(start_color=data_colors[5], end_color=data_colors[5], fill_type="solid")
            
            # Column 7: الكمية م2 - reference from slides
            area_col = get_column_letter(SLIDES_START_COL + 13)
            cell = wastage_sheet.cell(row=excel_row, column=7)
            cell.value = f"=البلوكات!{area_col}{source_row}"
            cell.border = thin_border
            cell.alignment = center_alignment
            cell.fill = PatternFill(start_color=data_colors[6], end_color=data_colors[6], fill_type="solid")
            cell.number_format = '#,##0.00'
            
            # Column 8: معدل الانتاجيه - formula based on thickness
            # IF(F="2سم",29,IF(F="3سم",21.6,IF(F="4سم",15,0)))
            cell = wastage_sheet.cell(row=excel_row, column=8)
            cell.value = f'=IF(F{excel_row}="2سم",29,IF(F{excel_row}="3سم",21.6,IF(F{excel_row}="4سم",15,0)))'
            cell.border = thin_border
            cell.alignment = center_alignment
            cell.fill = PatternFill(start_color=data_colors[7], end_color=data_colors[7], fill_type="solid")
            cell.number_format = '#,##0.00'
            
            # Column 9: الاجمالي = الكمية م2 * معدل الانتاجيه
            cell = wastage_sheet.cell(row=excel_row, column=9)
            cell.value = f"=G{excel_row}*H{excel_row}"
            cell.border = thin_border
            cell.alignment = center_alignment
            cell.fill = PatternFill(start_color=data_colors[8], end_color=data_colors[8], fill_type="solid")
            cell.number_format = '#,##0.00'
            
            # Column 10: الفرق (الهالك) = الكمية م2 - الاجمالي
            cell = wastage_sheet.cell(row=excel_row, column=10)
            cell.value = f"=G{excel_row}-I{excel_row}"
            cell.border = thin_border
            cell.alignment = center_alignment
            cell.fill = PatternFill(start_color=data_colors[9], end_color=data_colors[9], fill_type="solid")
            cell.number_format = '#,##0.00'
        
        # Freeze panes
        wastage_sheet.freeze_panes = 'A3'
        
        workbook.save(filepath)
        return True
        
    except PermissionError as e:
        log_error(f"File is locked: {e}")
        return False
    except Exception as e:
        log_exception(f"Error updating wastage sheet: {e}")
        return False
