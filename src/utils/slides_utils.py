import os
import json
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

from utils.log_utils import log_error, log_exception


def convert_arabic_datetime_to_excel(datetime_str):
    """
    Convert Arabic datetime string (e.g., '12:30 ص 01/01/2025') to Excel datetime
    
    Args:
        datetime_str (str): DateTime string in format 'HH:MM ص/م DD/MM/YYYY'
        
    Returns:
        datetime: Python datetime object or None if parsing fails
    """
    if not datetime_str:
        return None
    
    try:
        # Expected format: "12:30 ص 01/01/2025"
        parts = datetime_str.strip().split(' ')
        if len(parts) < 3:
            return None
        
        time_part = parts[0]  # "12:30"
        period = parts[1]     # "ص" or "م"
        date_part = parts[2]  # "01/01/2025"
        
        # Parse time
        time_parts = time_part.split(':')
        if len(time_parts) != 2:
            return None
        
        hour = int(time_parts[0])
        minute = int(time_parts[1])
        
        # Convert to 24-hour format
        if period == 'م' and hour != 12:
            hour += 12
        elif period == 'ص' and hour == 12:
            hour = 0
        
        # Parse date
        date_parts = date_part.split('/')
        if len(date_parts) != 3:
            return None
        
        day = int(date_parts[0])
        month = int(date_parts[1])
        year = int(date_parts[2])
        
        return datetime(year, month, day, hour, minute)
    except (ValueError, IndexError):
        return None


def initialize_slides_inventory_excel(file_path):
    """
    Initialize the slides inventory Excel file with proper formatting and formulas
    
    Args:
        file_path (str): Path to the Excel file
    """
    try:
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # Create sheets with Arabic right-to-left orientation
        add_sheet = wb.create_sheet("اذن اضافة الشرائح", 0)
        disburse_sheet = wb.create_sheet("اذن صرف الشرائح", 1)
        inventory_sheet = wb.create_sheet("مخزون الشرائح", 2)
        
        # Set right-to-left for all sheets
        for sheet in wb.sheetnames:
            wb[sheet].sheet_view.rightToLeft = True
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add headers to add sheet - Full publishing data
        # Columns: تاريخ النشر، رقم البلوك، النوع، رقم المكينه، وقت الدخول، وقت الخروج، عدد الساعات،
        #          السمك، العدد، الطول، الخصم، الطول بعد الخصم، الارتفاع، الكمية م2، سعر المتر، اجمالي السعر
        add_headers = [
            "تاريخ النشر", "رقم البلوك", "النوع", "رقم المكينه", 
            "وقت الدخول", "وقت الخروج", "عدد الساعات",
            "السمك", "العدد", "الطول", "الخصم", "الطول بعد", 
            "الارتفاع", "الكمية م2", "سعر المتر", "اجمالي السعر"
        ]
        for col_num, header in enumerate(add_headers, 1):
            cell = add_sheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Set column widths for add sheet
        add_column_widths = [12, 12, 15, 10, 12, 12, 12, 10, 8, 10, 8, 12, 10, 12, 12, 15]
        for col_num, width in enumerate(add_column_widths, 1):
            add_sheet.column_dimensions[get_column_letter(col_num)].width = width
        
        # Add headers to disburse sheet - New format with merged cells like ledger
        disburse_headers = ["رقم الفاتورة", "تاريخ الصرف", "اسم العميل", "اسم الصنف", "رقم البلوك", "السمك", "العدد", "ثمن الوحدة", "الإجمالي", "ملاحظات"]
        for col_num, header in enumerate(disburse_headers, 1):
            cell = disburse_sheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Set column widths for disburse sheet
        disburse_column_widths = [15, 15, 20, 20, 12, 10, 10, 15, 15, 25]
        for col_num, width in enumerate(disburse_column_widths, 1):
            disburse_sheet.column_dimensions[get_column_letter(col_num)].width = width
        
        # Add headers to inventory sheet - with block number and thickness
        inventory_headers = ["اسم الصنف", "رقم البلوك", "السمك", "إجمالي الإضافات", "إجمالي الصرف", "الرصيد الحالي"]
        for col_num, header in enumerate(inventory_headers, 1):
            cell = inventory_sheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Set column widths for inventory sheet
        inventory_column_widths = [20, 15, 12, 18, 18, 18]
        for col_num, width in enumerate(inventory_column_widths, 1):
            inventory_sheet.column_dimensions[get_column_letter(col_num)].width = width
        
        # Save the workbook
        wb.save(file_path)
        return wb
    except Exception as e:
        log_exception(f"Failed to initialize slides Excel file: {e}")
        raise


def convert_existing_slides_inventory_to_formulas(file_path):
    """
    Convert an existing slides inventory file to use formulas instead of manual calculations
    Shows inventory grouped by item name (النوع), block number, and thickness
    
    Args:
        file_path (str): Path to the Excel file
    """
    try:
        if not os.path.exists(file_path):
            # If file doesn't exist, create a new one
            initialize_slides_inventory_excel(file_path)
            return
        
        # Load workbook
        wb = openpyxl.load_workbook(file_path)
        add_sheet = wb["اذن اضافة الشرائح"]
        disburse_sheet = wb["اذن صرف الشرائح"]
        inventory_sheet = wb["مخزون الشرائح"]
        
        # Get all unique combinations of (item_name, block_number, thickness) from both sheets
        item_combinations = set()
        
        # Get items from additions sheet (skip header row)
        # New structure: Column B(2) = رقم البلوك, Column C(3) = النوع, Column H(8) = السمك, Column I(9) = العدد
        for row_num in range(2, add_sheet.max_row + 1):
            item_name = add_sheet.cell(row=row_num, column=3).value  # النوع (Column C)
            block_number = add_sheet.cell(row=row_num, column=2).value or ""  # رقم البلوك (Column B)
            thickness = add_sheet.cell(row=row_num, column=8).value or ""  # السمك (Column H)
            if item_name:
                item_combinations.add((item_name, str(block_number), str(thickness)))
        
        # Get items from disbursements sheet (skip header row)
        # Column 4 (D) = اسم الصنف, Column 5 (E) = رقم البلوك, Column 6 (F) = السمك, Column 7 (G) = العدد
        for row_num in range(2, disburse_sheet.max_row + 1):
            item_name = disburse_sheet.cell(row=row_num, column=4).value
            block_number = disburse_sheet.cell(row=row_num, column=5).value or ""
            thickness = disburse_sheet.cell(row=row_num, column=6).value or ""
            if item_name:
                item_combinations.add((item_name, str(block_number), str(thickness)))
        
        # Clear existing data in inventory sheet (keep header)
        for row_num in range(2, inventory_sheet.max_row + 1):
            for col_num in range(1, 7):
                inventory_sheet.cell(row=row_num, column=col_num).value = None
        
        # Apply styles
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        alignment = Alignment(horizontal='center', vertical='center')
        
        # Sort combinations by item name, then block number, then thickness
        sorted_combinations = sorted(item_combinations, key=lambda x: (x[0], x[1], x[2]))
        
        # Add items to inventory sheet with formulas
        row_num = 2
        for item_name, block_number, thickness in sorted_combinations:
            # Column 1: اسم الصنف
            inventory_sheet.cell(row=row_num, column=1, value=item_name).border = border
            inventory_sheet.cell(row=row_num, column=1).alignment = alignment
            
            # Column 2: رقم البلوك
            inventory_sheet.cell(row=row_num, column=2, value=block_number).border = border
            inventory_sheet.cell(row=row_num, column=2).alignment = alignment
            
            # Column 3: السمك
            inventory_sheet.cell(row=row_num, column=3, value=thickness).border = border
            inventory_sheet.cell(row=row_num, column=3).alignment = alignment
            
            # Column 4: إجمالي الإضافات - SUMIFS based on item name, block number, and thickness
            # New additions structure: C = النوع, B = رقم البلوك, H = السمك, I = العدد
            if block_number and thickness:
                additions_formula = f'=SUMIFS(\'اذن اضافة الشرائح\'!I:I,\'اذن اضافة الشرائح\'!C:C,A{row_num},\'اذن اضافة الشرائح\'!B:B,B{row_num},\'اذن اضافة الشرائح\'!H:H,C{row_num})'
            elif block_number:
                additions_formula = f'=SUMIFS(\'اذن اضافة الشرائح\'!I:I,\'اذن اضافة الشرائح\'!C:C,A{row_num},\'اذن اضافة الشرائح\'!B:B,B{row_num})'
            elif thickness:
                additions_formula = f'=SUMIFS(\'اذن اضافة الشرائح\'!I:I,\'اذن اضافة الشرائح\'!C:C,A{row_num},\'اذن اضافة الشرائح\'!H:H,C{row_num})'
            else:
                additions_formula = f"=SUMIF('اذن اضافة الشرائح'!C:C,A{row_num},'اذن اضافة الشرائح'!I:I)"
            
            inventory_sheet.cell(row=row_num, column=4).value = additions_formula
            inventory_sheet.cell(row=row_num, column=4).border = border
            inventory_sheet.cell(row=row_num, column=4).alignment = alignment
            inventory_sheet.cell(row=row_num, column=4).number_format = '#,##0'
            
            # Column 5: إجمالي الصرف - SUMIFS based on item name, block number, and thickness
            if block_number and thickness:
                disbursements_formula = f'=SUMIFS(\'اذن صرف الشرائح\'!G:G,\'اذن صرف الشرائح\'!D:D,A{row_num},\'اذن صرف الشرائح\'!E:E,B{row_num},\'اذن صرف الشرائح\'!F:F,C{row_num})'
            elif block_number:
                disbursements_formula = f'=SUMIFS(\'اذن صرف الشرائح\'!G:G,\'اذن صرف الشرائح\'!D:D,A{row_num},\'اذن صرف الشرائح\'!E:E,B{row_num})'
            elif thickness:
                disbursements_formula = f'=SUMIFS(\'اذن صرف الشرائح\'!G:G,\'اذن صرف الشرائح\'!D:D,A{row_num},\'اذن صرف الشرائح\'!F:F,C{row_num})'
            else:
                disbursements_formula = f"=SUMIF('اذن صرف الشرائح'!D:D,A{row_num},'اذن صرف الشرائح'!G:G)"
            
            inventory_sheet.cell(row=row_num, column=5).value = disbursements_formula
            inventory_sheet.cell(row=row_num, column=5).border = border
            inventory_sheet.cell(row=row_num, column=5).alignment = alignment
            inventory_sheet.cell(row=row_num, column=5).number_format = '#,##0'
            
            # Column 6: الرصيد الحالي (additions - disbursements)
            balance_formula = f"=D{row_num}-E{row_num}"
            inventory_sheet.cell(row=row_num, column=6).value = balance_formula
            inventory_sheet.cell(row=row_num, column=6).border = border
            inventory_sheet.cell(row=row_num, column=6).alignment = alignment
            inventory_sheet.cell(row=row_num, column=6).number_format = '#,##0'
            
            row_num += 1
        
        # Save the workbook
        wb.save(file_path)
    except Exception as e:
        log_exception(f"Failed to convert slides to formulas: {e}")
        raise


def add_slides_inventory_entry(file_path, item_name, quantity, unit_price, notes="", entry_date=None):
    """
    Add a slides inventory entry to the additions sheet
    
    Args:
        file_path (str): Path to the Excel file
        item_name (str): Name of the item
        quantity (float): Quantity of the item
        unit_price (float): Price per unit
        notes (str): Additional notes
        entry_date (str): Date of entry (defaults to today)
        
    Returns:
        int: Entry number
    """
    try:
        # Load workbook
        wb = openpyxl.load_workbook(file_path)
        add_sheet = wb["اذن اضافة الشرائح"]
        
        # Determine the next entry number
        next_entry_number = add_sheet.max_row
        
        # Get today's date if not provided
        if entry_date is None:
            entry_date = datetime.now().strftime('%d/%m/%Y')
        
        # Calculate total price
        total_price = float(quantity) * float(unit_price)
        
        # Add data row
        row_data = [
            next_entry_number,  # Auto entry number
            entry_date,
            item_name,
            float(quantity),
            float(unit_price),
            total_price,
            notes
        ]
        
        # Add row to sheet
        add_sheet.append(row_data)
        
        # Apply styles to the new row
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        alignment = Alignment(horizontal='center', vertical='center')
        
        row_num = add_sheet.max_row
        for col_num, value in enumerate(row_data, 1):
            cell = add_sheet.cell(row=row_num, column=col_num)
            cell.border = border
            cell.alignment = alignment
            # Apply number formatting for numeric columns
            if col_num in [4, 5, 6]:  # Quantity, Unit Price, Total Price
                cell.number_format = '#,##0.00'
        
        # Save the workbook
        wb.save(file_path)
        
        # Update inventory sheet with formulas
        convert_existing_slides_inventory_to_formulas(file_path)
        
        return next_entry_number
    except Exception as e:
        log_exception(f"Failed to add slides inventory entry: {e}")
        raise


def add_slides_inventory_from_publishing(file_path, publishing_data):
    """
    Add slides inventory entries from publishing data (from slides_add_view)
    With formulas for calculated fields
    
    Args:
        file_path (str): Path to the Excel file
        publishing_data (list): List of dictionaries containing publishing data
        
    Returns:
        int: Number of entries added
    """
    try:
        # Load workbook
        wb = openpyxl.load_workbook(file_path)
        add_sheet = wb["اذن اضافة الشرائح"]
        
        # Apply styles
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        alignment = Alignment(horizontal='center', vertical='center')
        
        # Default discount value
        DEFAULT_DISCOUNT = 0.20
        
        entries_added = 0
        for entry in publishing_data:
            row_num = add_sheet.max_row + 1
            
            # Get values from entry
            publishing_date = entry.get('publishing_date', datetime.now().strftime('%Y-%m-%d'))
            block_number = entry.get('block_number', '')
            material = entry.get('material', '')
            machine_number = entry.get('machine_number', '')
            entry_time = entry.get('entry_time', '')
            exit_time = entry.get('exit_time', '')
            thickness = entry.get('thickness', '')
            quantity = int(entry.get('quantity', 0)) if entry.get('quantity') else 0
            length = float(entry.get('length', 0)) if entry.get('length') else 0
            height = float(entry.get('height', 0)) if entry.get('height') else 0
            price_per_meter = float(entry.get('price_per_meter', 0)) if entry.get('price_per_meter') else 0
            
            # Convert datetime strings to Excel datetime format
            entry_time_excel = convert_arabic_datetime_to_excel(entry_time)
            exit_time_excel = convert_arabic_datetime_to_excel(exit_time)
            
            # Column mapping (1-based):
            # A(1): تاريخ النشر, B(2): رقم البلوك, C(3): النوع, D(4): رقم المكينه
            # E(5): وقت الدخول, F(6): وقت الخروج, G(7): عدد الساعات (formula)
            # H(8): السمك, I(9): العدد, J(10): الطول, K(11): الخصم (0.20)
            # L(12): الطول بعد الخصم (formula), M(13): الارتفاع
            # N(14): الكمية م2 (formula), O(15): سعر المتر, P(16): اجمالي السعر (formula)
            
            # Set static values
            add_sheet.cell(row=row_num, column=1, value=publishing_date)  # تاريخ النشر
            add_sheet.cell(row=row_num, column=2, value=block_number)     # رقم البلوك
            add_sheet.cell(row=row_num, column=3, value=material)         # النوع
            add_sheet.cell(row=row_num, column=4, value=machine_number)   # رقم المكينه
            
            # وقت الدخول - Store as Excel datetime
            entry_cell = add_sheet.cell(row=row_num, column=5)
            if entry_time_excel:
                entry_cell.value = entry_time_excel
                entry_cell.number_format = 'DD/MM/YYYY HH:MM'
            else:
                entry_cell.value = entry_time
            
            # وقت الخروج - Store as Excel datetime
            exit_cell = add_sheet.cell(row=row_num, column=6)
            if exit_time_excel:
                exit_cell.value = exit_time_excel
                exit_cell.number_format = 'DD/MM/YYYY HH:MM'
            else:
                exit_cell.value = exit_time
            
            # عدد الساعات - Formula: Calculate hours between entry and exit time
            # Formula: (F - E) * 24 to get hours
            hours_formula = f"=IF(AND(E{row_num}<>\"\",F{row_num}<>\"\"),(F{row_num}-E{row_num})*24,0)"
            add_sheet.cell(row=row_num, column=7, value=hours_formula)    # عدد الساعات
            
            add_sheet.cell(row=row_num, column=8, value=thickness)        # السمك
            add_sheet.cell(row=row_num, column=9, value=quantity)         # العدد
            add_sheet.cell(row=row_num, column=10, value=length)          # الطول
            add_sheet.cell(row=row_num, column=11, value=DEFAULT_DISCOUNT) # الخصم (0.20)
            
            # الطول بعد الخصم - Formula: الطول - الخصم (J - K)
            length_after_formula = f"=J{row_num}-K{row_num}"
            add_sheet.cell(row=row_num, column=12, value=length_after_formula)  # الطول بعد الخصم
            
            add_sheet.cell(row=row_num, column=13, value=height)          # الارتفاع
            
            # الكمية م2 - Formula: الطول بعد الخصم * الارتفاع * العدد (L * M * I)
            area_formula = f"=L{row_num}*M{row_num}*I{row_num}"
            add_sheet.cell(row=row_num, column=14, value=area_formula)    # الكمية م2
            
            add_sheet.cell(row=row_num, column=15, value=price_per_meter) # سعر المتر
            
            # اجمالي السعر - Formula: الكمية م2 * سعر المتر (N * O)
            total_formula = f"=N{row_num}*O{row_num}"
            add_sheet.cell(row=row_num, column=16, value=total_formula)   # اجمالي السعر
            
            # Apply styles to all cells in the row
            for col_num in range(1, 17):
                cell = add_sheet.cell(row=row_num, column=col_num)
                cell.border = border
                cell.alignment = alignment
                
                # Apply number formatting
                if col_num in [7, 10, 11, 12, 13, 14, 15, 16]:  # Numeric columns
                    cell.number_format = '#,##0.00'
                elif col_num == 9:  # العدد - integer
                    cell.number_format = '#,##0'
            
            entries_added += 1
        
        # Save the workbook
        wb.save(file_path)
        
        # Update inventory sheet with formulas
        convert_existing_slides_inventory_to_formulas(file_path)
        
        # Also save to blocks Excel file
        try:
            from utils.blocks_utils import export_slides_to_blocks_excel
            export_slides_to_blocks_excel(publishing_data)
        except Exception as blocks_error:
            log_error(f"Could not save slides to blocks file: {blocks_error}")
            # Don't fail the main operation if blocks file fails
        
        return entries_added
    except Exception as e:
        log_exception(f"Failed to add slides inventory from publishing: {e}")
        raise


def disburse_slides_inventory_entry(
    file_path, 
    invoice_number,
    disburse_date,
    client_name,
    item_name,
    block_number,
    thickness,
    quantity, 
    unit_price, 
    notes=""
):
    """
    Add a slides inventory disbursement entry to the disbursements sheet
    
    Args:
        file_path (str): Path to the Excel file
        invoice_number (str): Invoice number
        disburse_date (str): Date of disbursement
        client_name (str): Client name
        item_name (str): Name of the item
        block_number (str): Block number
        thickness (str): Thickness of the item
        quantity (int): Quantity (count) of the item
        unit_price (float): Price per unit
        notes (str): Additional notes
        
    Returns:
        int: Row number of the entry
    """
    try:
        # Load workbook
        wb = openpyxl.load_workbook(file_path)
        disburse_sheet = wb["اذن صرف الشرائح"]
        
        # Get today's date if not provided
        if disburse_date is None:
            disburse_date = datetime.now().strftime('%d/%m/%Y')
        
        # Round values
        quantity_int = int(float(quantity))
        unit_price_rounded = round(float(unit_price), 0)
        total_price = round(quantity_int * unit_price_rounded, 0)
        
        # Add data row
        row_data = [
            invoice_number,      # رقم الفاتورة
            disburse_date,       # تاريخ الصرف
            client_name,         # اسم العميل
            item_name,           # اسم الصنف
            block_number,        # رقم البلوك
            thickness,           # السمك
            quantity_int,        # العدد
            unit_price_rounded,  # ثمن الوحدة
            total_price,         # الإجمالي
            notes                # ملاحظات
        ]
        
        # Add row to sheet
        disburse_sheet.append(row_data)
        
        # Apply styles to the new row
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        alignment = Alignment(horizontal='center', vertical='center')
        
        row_num = disburse_sheet.max_row
        for col_num, value in enumerate(row_data, 1):
            cell = disburse_sheet.cell(row=row_num, column=col_num)
            cell.border = border
            cell.alignment = alignment
            # Apply number formatting for numeric columns
            if col_num == 7:  # العدد
                cell.number_format = '#,##0'
            elif col_num in [8, 9]:  # ثمن الوحدة، الإجمالي
                cell.number_format = '#,##0'
        
        # Save the workbook
        wb.save(file_path)
        
        # Update inventory sheet with formulas
        convert_existing_slides_inventory_to_formulas(file_path)
        
        return row_num
    except Exception as e:
        log_exception(f"Failed to add slides disbursement entry: {e}")
        raise


def get_slides_inventory_summary(file_path):
    """
    Get slides inventory summary data by evaluating formulas
    
    Args:
        file_path (str): Path to the Excel file
        
    Returns:
        list: List of dictionaries containing inventory data
    """
    if not os.path.exists(file_path):
        return []
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)  # data_only=True to get calculated values
        inventory_sheet = wb["مخزون الشرائح"]
        
        inventory_data = []
        for row_num in range(2, inventory_sheet.max_row + 1):
            item_name = inventory_sheet.cell(row=row_num, column=1).value
            if item_name:  # Only process rows with item names
                block_number = inventory_sheet.cell(row=row_num, column=2).value or ""
                thickness = inventory_sheet.cell(row=row_num, column=3).value or ""
                total_additions = inventory_sheet.cell(row=row_num, column=4).value or 0
                total_disbursements = inventory_sheet.cell(row=row_num, column=5).value or 0
                current_balance = inventory_sheet.cell(row=row_num, column=6).value or 0
                
                # Convert to float and handle None values
                try:
                    total_additions = float(total_additions)
                except (ValueError, TypeError):
                    total_additions = 0
                    
                try:
                    total_disbursements = float(total_disbursements)
                except (ValueError, TypeError):
                    total_disbursements = 0
                    
                try:
                    current_balance = float(current_balance)
                except (ValueError, TypeError):
                    current_balance = 0
                
                item_data = {
                    'item_name': item_name,
                    'block_number': block_number,
                    'thickness': thickness,
                    'total_additions': total_additions,
                    'total_disbursements': total_disbursements,
                    'current_balance': current_balance
                }
                inventory_data.append(item_data)
        
        return inventory_data
    except Exception as e:
        log_exception(f"Error reading slides inventory summary: {e}")
        return []


def get_available_slides_items_with_prices(file_path):
    """
    Get available slides items with their average unit prices from additions
    
    Args:
        file_path (str): Path to the Excel file
        
    Returns:
        dict: Dictionary with item names as keys and average unit prices as values
    """
    if not os.path.exists(file_path):
        return {}
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)  # data_only=True to get calculated values
        add_sheet = wb["اذن اضافة الشرائح"]
        inventory_sheet = wb["مخزون الشرائح"]
        
        # Get current inventory balances (now column 6 is balance)
        inventory_balances = {}
        for row_num in range(2, inventory_sheet.max_row + 1):
            item_name = inventory_sheet.cell(row=row_num, column=1).value
            if item_name:  # Only process rows with item names
                balance = inventory_sheet.cell(row=row_num, column=6).value or 0
                try:
                    # Accumulate balance for same item name (different blocks/thicknesses)
                    if item_name not in inventory_balances:
                        inventory_balances[item_name] = 0
                    inventory_balances[item_name] += float(balance)
                except (ValueError, TypeError):
                    pass
        
        # Get item prices from additions
        item_prices = {}
        item_quantities = {}
        
        # Skip header row (row 1)
        for row_num in range(2, add_sheet.max_row + 1):
            item_name = add_sheet.cell(row=row_num, column=3).value  # Item name column
            quantity = add_sheet.cell(row=row_num, column=4).value or 0  # Quantity column
            unit_price = add_sheet.cell(row=row_num, column=5).value or 0  # Unit price column
            
            try:
                quantity = float(quantity)
                unit_price = float(unit_price)
            except (ValueError, TypeError):
                continue  # Skip invalid rows
            
            if item_name and quantity > 0:
                # Only include items that have positive balance in inventory
                if inventory_balances.get(item_name, 0) > 0:
                    if item_name not in item_prices:
                        item_prices[item_name] = 0
                        item_quantities[item_name] = 0
                    
                    # Accumulate weighted prices
                    item_prices[item_name] += unit_price * quantity
                    item_quantities[item_name] += quantity
        
        # Calculate average prices
        avg_prices = {}
        for item_name in item_prices:
            if item_quantities[item_name] > 0:
                avg_prices[item_name] = item_prices[item_name] / item_quantities[item_name]
            else:
                avg_prices[item_name] = 0
        
        return avg_prices
    except Exception as e:
        log_exception(f"Error getting available slides items with prices: {e}")
        return {}


def initialize_slides_publishing_excel(file_path):
    """
    Initialize the slides publishing Excel file with proper formatting
    
    Args:
        file_path (str): Path to the Excel file
    """
    try:
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # Create publishing sheet
        publish_sheet = wb.create_sheet("اذن اضافه", 0)
        
        # Set right-to-left
        publish_sheet.sheet_view.rightToLeft = True
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add headers to publishing sheet
        publish_headers = [
            "تاريخ النشر", "رقم البلوك", "النوع", "رقم المكينه", "عدد", 
            "الطول", "الارتفاع", "السمك", "م2", "سعر المتر", "اجمالي السعر", 
            "وقت الدخول", "وقت الخروج", "عدد الساعات"
        ]
        for col_num, header in enumerate(publish_headers, 1):
            cell = publish_sheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Set column widths for publishing sheet
        column_widths = [15, 12, 15, 12, 10, 12, 12, 12, 12, 15, 15, 15, 15, 15]
        for col_num, width in enumerate(column_widths, 1):
            publish_sheet.column_dimensions[get_column_letter(col_num)].width = width
        
        # Save the workbook
        wb.save(file_path)
        return wb
    except Exception as e:
        log_exception(f"Failed to initialize slides publishing Excel file: {e}")
        raise


def _remove_invoice_from_slides_disbursement(file_path, invoice_number):
    """
    Remove all disbursement entries for a specific invoice number.
    This is used when updating an existing invoice.
    
    Args:
        file_path (str): Path to the Excel file
        invoice_number (str): Invoice number to remove
    """
    try:
        if not os.path.exists(file_path):
            return
        
        wb = openpyxl.load_workbook(file_path)
        disburse_sheet = wb["اذن صرف الشرائح"]
        
        # Find all rows with this invoice number (Column A = رقم الفاتورة)
        rows_to_delete = []
        for row_num in range(2, disburse_sheet.max_row + 1):
            cell_value = disburse_sheet.cell(row=row_num, column=1).value
            if cell_value is not None and str(cell_value) == str(invoice_number):
                rows_to_delete.append(row_num)
        
        # Also check for merged cells - the invoice number might only be in the first row of a merged group
        # We need to find all rows that belong to this invoice
        # Check if there are rows after the invoice number that have empty first column (part of merge)
        if rows_to_delete:
            first_row = rows_to_delete[0]
            # Check subsequent rows until we find a non-empty cell or another invoice number
            for row_num in range(first_row + 1, disburse_sheet.max_row + 1):
                cell_value = disburse_sheet.cell(row=row_num, column=1).value
                if cell_value is None or str(cell_value).strip() == "":
                    # This row might be part of the merged group
                    # Check if it has data in other columns (like item name)
                    item_name = disburse_sheet.cell(row=row_num, column=4).value
                    if item_name:
                        rows_to_delete.append(row_num)
                else:
                    # Found another invoice number, stop
                    break
        
        # Delete rows in reverse order to maintain correct indices
        rows_to_delete = sorted(set(rows_to_delete), reverse=True)
        for row_num in rows_to_delete:
            # Unmerge any merged cells in this row first
            for merged_range in list(disburse_sheet.merged_cells.ranges):
                if merged_range.min_row <= row_num <= merged_range.max_row:
                    try:
                        disburse_sheet.unmerge_cells(str(merged_range))
                    except Exception:
                        pass
            disburse_sheet.delete_rows(row_num)
        
        if rows_to_delete:
            wb.save(file_path)
            # Update inventory formulas
            convert_existing_slides_inventory_to_formulas(file_path)
            
    except Exception as e:
        log_exception(f"Failed to remove invoice from slides disbursement: {e}")


def disburse_slides_from_invoice(invoice_number, invoice_date, items_data, client_name=""):
    """
    Disburse slides from inventory based on invoice items.
    This function is called when saving an invoice that contains slide products.
    
    Args:
        invoice_number (str): Invoice number for reference
        invoice_date (str): Invoice date
        items_data (list): List of tuples containing item data from invoice
                          Format: (description, block, thickness, material, count, length, height, price, ...)
        client_name (str): Client name
        
    Returns:
        tuple: (success: bool, message: str, disbursed_items: list)
    """
    try:
        # Get the slides inventory file path
        documents_path = os.path.join(os.path.expanduser("~"), "Documents", "alswaife")
        slides_path = os.path.join(documents_path, "الشرائح")
        inventory_file = os.path.join(slides_path, "مخزون الشرائح.xlsx")
        
        # Check if inventory file exists
        if not os.path.exists(inventory_file):
            os.makedirs(slides_path, exist_ok=True)
            initialize_slides_inventory_excel(inventory_file)
        
        # Remove old entries for this invoice first (for updates)
        _remove_invoice_from_slides_disbursement(inventory_file, invoice_number)
        
        # Load slides products to check if item is a slide product
        slides_products = []
        try:
            project_root = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
            json_path = os.path.join(project_root, "data", "slides_products.json")
            with open(json_path, 'r', encoding='utf-8') as f:
                products = json.load(f)
                slides_products = [p["name"] for p in products]
        except Exception:
            # Default slides products if JSON fails
            slides_products = ["نيو حلايب", "جندولا", "احمر اسوان"]
        
        # Collect all slide items first
        slide_items_to_add = []
        
        # Process each item in the invoice
        for item in items_data:
            try:
                description = item[0] if len(item) > 0 else ""
                
                if not description:
                    continue
                
                # Check if this is a slide product
                # The description could be just the name (e.g., "نيو حلايب") 
                # or with prefix (e.g., "ش نيو حلايب")
                material_name = description
                if description.startswith("ش "):
                    material_name = description[2:]
                
                # Check if material is in slides products list
                if material_name not in slides_products:
                    continue
                
                # Get item details
                block_number = item[1] if len(item) > 1 else ""
                thickness = item[2] if len(item) > 2 else ""
                count = int(float(item[4])) if len(item) > 4 and item[4] else 0
                price = float(item[7]) if len(item) > 7 and item[7] else 0
                
                if count <= 0:
                    continue
                
                # Item name is just the material name (without thickness)
                item_name = material_name
                
                # Calculate total price for this item
                total_price = count * price
                
                slide_items_to_add.append({
                    'item_name': item_name,
                    'block_number': block_number,
                    'thickness': thickness,
                    'quantity': count,
                    'unit_price': price,
                    'total_price': total_price
                })
                
            except Exception as item_ex:
                log_error(f"Error processing item for disbursement: {item_ex}")
                continue
        
        # If we have slide items, add them with merged cells
        if slide_items_to_add:
            _add_slides_disbursement_with_merge(
                inventory_file,
                invoice_number,
                invoice_date,
                client_name,
                slide_items_to_add
            )
        
        if slide_items_to_add:
            return True, f"تم صرف {len(slide_items_to_add)} صنف من الشرائح", slide_items_to_add
        else:
            return True, "لا توجد شرائح في الفاتورة", []
            
    except Exception as e:
        log_exception(f"Error disbursing slides from invoice: {e}")
        return False, f"خطأ في صرف الشرائح: {str(e)}", []


def _add_slides_disbursement_with_merge(file_path, invoice_number, invoice_date, client_name, items):
    """
    Add multiple slides disbursement entries with merged cells for invoice info
    
    Args:
        file_path (str): Path to the Excel file
        invoice_number (str): Invoice number
        invoice_date (str): Date of disbursement
        client_name (str): Client name
        items (list): List of item dictionaries
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        disburse_sheet = wb["اذن صرف الشرائح"]
        
        # Get the starting row
        start_row = disburse_sheet.max_row + 1
        num_items = len(items)
        
        # Define styles
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        alignment = Alignment(horizontal='center', vertical='center')
        
        # Add all item rows first
        for idx, item in enumerate(items):
            row_num = start_row + idx
            
            quantity_int = int(float(item['quantity']))
            unit_price_rounded = round(float(item['unit_price']), 0)
            total_price = round(quantity_int * unit_price_rounded, 0)
            
            # Set values for each column
            # Column 1: رقم الفاتورة (will be merged)
            disburse_sheet.cell(row=row_num, column=1, value=invoice_number if idx == 0 else "")
            # Column 2: تاريخ الصرف (will be merged)
            disburse_sheet.cell(row=row_num, column=2, value=invoice_date if idx == 0 else "")
            # Column 3: اسم العميل (will be merged)
            disburse_sheet.cell(row=row_num, column=3, value=client_name if idx == 0 else "")
            # Column 4: اسم الصنف
            disburse_sheet.cell(row=row_num, column=4, value=item['item_name'])
            # Column 5: رقم البلوك
            disburse_sheet.cell(row=row_num, column=5, value=item['block_number'])
            # Column 6: السمك
            disburse_sheet.cell(row=row_num, column=6, value=item['thickness'])
            # Column 7: العدد
            disburse_sheet.cell(row=row_num, column=7, value=quantity_int)
            # Column 8: ثمن الوحدة
            disburse_sheet.cell(row=row_num, column=8, value=unit_price_rounded)
            # Column 9: الإجمالي
            disburse_sheet.cell(row=row_num, column=9, value=total_price)
            # Column 10: ملاحظات
            disburse_sheet.cell(row=row_num, column=10, value="")
            
            # Apply styles to all cells
            for col_num in range(1, 11):
                cell = disburse_sheet.cell(row=row_num, column=col_num)
                cell.border = border
                cell.alignment = alignment
                if col_num == 7:  # العدد
                    cell.number_format = '#,##0'
                elif col_num in [8, 9]:  # ثمن الوحدة، الإجمالي
                    cell.number_format = '#,##0'
        
        # Merge cells for invoice info if more than one item
        if num_items > 1:
            end_row = start_row + num_items - 1
            # Merge رقم الفاتورة (Column A)
            disburse_sheet.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            # Merge تاريخ الصرف (Column B)
            disburse_sheet.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
            # Merge اسم العميل (Column C)
            disburse_sheet.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)
        
        # Save the workbook
        wb.save(file_path)
        
        # Update inventory sheet with formulas
        convert_existing_slides_inventory_to_formulas(file_path)
        
    except Exception as e:
        log_exception(f"Failed to add slides disbursement with merge: {e}")
        raise


def add_slides_publishing_entry(file_path, publishing_data):
    """
    Add a slides inventory entry to the publishing sheet
    
    Args:
        file_path (str): Path to the Excel file
        publishing_data (list): List of dictionaries containing publishing data
        
    Returns:
        int: Number of entries added
    """
    try:
        # Load workbook
        wb = openpyxl.load_workbook(file_path)
        
        # Check if the sheet exists, if not create it
        if "اذن اضافه" not in wb.sheetnames:
            # Create the sheet with proper formatting
            publish_sheet = wb.create_sheet("اذن اضافه", 0)
            publish_sheet.sheet_view.rightToLeft = True
            
            # Define styles
            header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Add headers
            publish_headers = [
                "تاريخ النشر", "رقم البلوك", "النوع", "رقم المكينه", "عدد", 
                "الطول", "الارتفاع", "السمك", "م2", "سعر المتر", "اجمالي السعر", 
                "وقت الدخول", "وقت الخروج", "عدد الساعات"
            ]
            for col_num, header in enumerate(publish_headers, 1):
                cell = publish_sheet.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Set column widths
            column_widths = [15, 12, 15, 12, 10, 12, 12, 12, 12, 15, 15, 15, 15, 15]
            for col_num, width in enumerate(column_widths, 1):
                publish_sheet.column_dimensions[get_column_letter(col_num)].width = width
        else:
            publish_sheet = wb["اذن اضافه"]
        
        # Add data rows
        for entry in publishing_data:
            row_data = [
                entry.get('publishing_date', ''),
                entry.get('block_number', ''),
                entry.get('material', ''),
                entry.get('machine_number', ''),
                int(entry.get('quantity', 0)),
                float(entry.get('length', 0)),
                float(entry.get('height', 0)),
                entry.get('thickness', ''),
                '',  # Area will be calculated as formula
                float(entry.get('price_per_meter', 0)),
                '',  # Total price will be calculated as formula
                entry.get('entry_time', ''),
                entry.get('exit_time', ''),
                float(entry.get('hours_count', 0))
            ]
            
            # Add row to sheet
            publish_sheet.append(row_data)
            
            # Apply styles to the new row
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            alignment = Alignment(horizontal='center', vertical='center')
            
            row_num = publish_sheet.max_row
            for col_num, value in enumerate(row_data, 1):
                cell = publish_sheet.cell(row=row_num, column=col_num)
                cell.border = border
                cell.alignment = alignment
                # Convert quantity (col 5) to int, keep others as float
                if col_num == 5:  # Quantity column should be integer
                    cell.value = int(float(value)) if value else 0
                # Apply number formatting for numeric columns
                if col_num in [5, 6, 7, 9, 10, 11, 14]:  # Quantity, Length, Height, Area, Price, Total, Hours
                    if col_num == 5:  # Quantity column should be integer
                        cell.number_format = '#,##0'
                    else:
                        cell.number_format = '#,##0.00'
            
            # Add formulas for Area (م2) and Total Price
            # Area = Quantity * Length * Height (Column 5 * Column 6 * Column 7)
            area_formula = f'={get_column_letter(5)}{row_num}*{get_column_letter(6)}{row_num}*{get_column_letter(7)}{row_num}'
            publish_sheet.cell(row=row_num, column=9, value=area_formula)  # Column I (9) is Area
            
            # Total Price = Area * Price per meter (Column 9 * Column 10)
            total_price_formula = f'={get_column_letter(9)}{row_num}*{get_column_letter(10)}{row_num}'
            publish_sheet.cell(row=row_num, column=11, value=total_price_formula)  # Column K (11) is Total Price
        
        # Save the workbook
        wb.save(file_path)
        
        return len(publishing_data)
    except Exception as e:
        log_exception(f"Failed to add slides publishing entries: {e}")
        raise