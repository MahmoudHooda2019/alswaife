import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

from utils.log_utils import log_error, log_exception


def initialize_inventory_excel(file_path):
    """
    Initialize the inventory Excel file with proper formatting and formulas
    
    Args:
        file_path (str): Path to the Excel file
    """
    try:
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # Create sheets with Arabic right-to-left orientation
        add_sheet = wb.create_sheet("اذن الاضافه", 0)
        disburse_sheet = wb.create_sheet("اذن الصرف", 1)
        inventory_sheet = wb.create_sheet("المخزون", 2)
        
        # Set right-to-left for all sheets
        for sheet in wb.sheetnames:
            wb[sheet].sheet_view.rightToLeft = True
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add headers to add sheet
        add_headers = ["رقم اذن الاضافه", "تاريخ الدخول", "اسم الصنف", "العدد", "ثمن الوحدة", "الإجمالي", "ملاحظات"]
        for col_num, header in enumerate(add_headers, 1):
            cell = add_sheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Set column widths for add sheet
        column_widths = [18, 15, 25, 12, 15, 15, 30]
        for col_num, width in enumerate(column_widths, 1):
            add_sheet.column_dimensions[get_column_letter(col_num)].width = width
        
        # Add headers to disburse sheet
        disburse_headers = ["رقم اذن الصرف", "تاريخ الصرف", "اسم الصنف", "العدد", "ثمن الوحدة", "الإجمالي", "ملاحظات"]
        for col_num, header in enumerate(disburse_headers, 1):
            cell = disburse_sheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Set column widths for disburse sheet
        for col_num, width in enumerate(column_widths, 1):
            disburse_sheet.column_dimensions[get_column_letter(col_num)].width = width
        
        # Add headers to inventory sheet
        inventory_headers = ["اسم الصنف", "إجمالي الإضافات", "إجمالي الصرف", "الرصيد الحالي"]
        for col_num, header in enumerate(inventory_headers, 1):
            cell = inventory_sheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Set column widths for inventory sheet
        inventory_column_widths = [30, 20, 20, 20]
        for col_num, width in enumerate(inventory_column_widths, 1):
            inventory_sheet.column_dimensions[get_column_letter(col_num)].width = width
        
        # Save the workbook
        wb.save(file_path)
        return wb
    except Exception as e:
        log_exception(f"Failed to initialize Excel file: {e}")
        raise


def convert_existing_inventory_to_formulas(file_path):
    """
    Convert an existing inventory file to use formulas instead of manual calculations
    
    Args:
        file_path (str): Path to the Excel file
    """
    try:
        if not os.path.exists(file_path):
            # If file doesn't exist, create a new one
            initialize_inventory_excel(file_path)
            return
        
        # Load workbook
        wb = openpyxl.load_workbook(file_path)
        add_sheet = wb["اذن الاضافه"]
        disburse_sheet = wb["اذن الصرف"]
        inventory_sheet = wb["المخزون"]
        
        # Get all unique item names from both sheets
        item_names = set()
        
        # Get items from additions sheet (skip header row)
        for row_num in range(2, add_sheet.max_row + 1):
            item_name = add_sheet.cell(row=row_num, column=3).value  # Item name column
            if item_name:
                item_names.add(item_name)
        
        # Get items from disbursements sheet (skip header row)
        for row_num in range(2, disburse_sheet.max_row + 1):
            item_name = disburse_sheet.cell(row=row_num, column=3).value  # Item name column
            if item_name:
                item_names.add(item_name)
        
        # Clear existing data in inventory sheet (keep header)
        for row_num in range(2, inventory_sheet.max_row + 1):
            for col_num in range(1, 5):
                inventory_sheet.cell(row=row_num, column=col_num).value = None
        
        # Apply styles
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        alignment = Alignment(horizontal='center', vertical='center')
        
        # Add items to inventory sheet with corrected formulas
        for row_num, item_name in enumerate(sorted(item_names), 2):
            # Item name
            inventory_sheet.cell(row=row_num, column=1, value=item_name).border = border
            inventory_sheet.cell(row=row_num, column=1).alignment = alignment
            
            # Formula for total additions (SUMIF from additions sheet)
            # Using single quotes around sheet names to handle spaces
            additions_formula = f"=SUMIF('اذن الاضافه'!C:C,\"{item_name}\",'اذن الاضافه'!D:D)"
            inventory_sheet.cell(row=row_num, column=2).value = additions_formula
            inventory_sheet.cell(row=row_num, column=2).border = border
            inventory_sheet.cell(row=row_num, column=2).alignment = alignment
            inventory_sheet.cell(row=row_num, column=2).number_format = '#,##0.00'
            
            # Formula for total disbursements (SUMIF from disbursements sheet)
            # Using single quotes around sheet names to handle spaces
            disbursements_formula = f"=SUMIF('اذن الصرف'!C:C,\"{item_name}\",'اذن الصرف'!D:D)"
            inventory_sheet.cell(row=row_num, column=3).value = disbursements_formula
            inventory_sheet.cell(row=row_num, column=3).border = border
            inventory_sheet.cell(row=row_num, column=3).alignment = alignment
            inventory_sheet.cell(row=row_num, column=3).number_format = '#,##0.00'
            
            # Formula for current balance (additions - disbursements)
            balance_formula = f"=B{row_num}-C{row_num}"
            inventory_sheet.cell(row=row_num, column=4).value = balance_formula
            inventory_sheet.cell(row=row_num, column=4).border = border
            inventory_sheet.cell(row=row_num, column=4).alignment = alignment
            inventory_sheet.cell(row=row_num, column=4).number_format = '#,##0.00'
        
        # Save the workbook
        wb.save(file_path)
    except Exception as e:
        log_exception(f"Failed to convert to formulas: {e}")
        raise


def add_inventory_entry(file_path, item_name, quantity, unit_price, notes="", entry_date=None):
    """
    Add an inventory entry to the additions sheet
    
    Args:
        file_path (str): Path to the Excel file
        item_name (str): Name of the item
        quantity (float): Quantity of the item
        unit_price (float): Price per unit
        notes (str): Additional notes
        entry_date (str): Date of entry (defaults to today)
        
    Returns:
        int: Entry number
    """
    try:
        # Load workbook
        wb = openpyxl.load_workbook(file_path)
        add_sheet = wb["اذن الاضافه"]
        
        # Determine the next entry number
        next_entry_number = add_sheet.max_row
        
        # Get today's date if not provided
        if entry_date is None:
            entry_date = datetime.now().strftime('%d/%m/%Y')
        
        # Calculate total price
        total_price = float(quantity) * float(unit_price)
        
        # Add data row
        row_data = [
            next_entry_number,  # Auto entry number
            entry_date,
            item_name,
            float(quantity),
            float(unit_price),
            total_price,
            notes
        ]
        
        # Add row to sheet
        add_sheet.append(row_data)
        
        # Apply styles to the new row
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        alignment = Alignment(horizontal='center', vertical='center')
        
        row_num = add_sheet.max_row
        for col_num, value in enumerate(row_data, 1):
            cell = add_sheet.cell(row=row_num, column=col_num)
            cell.border = border
            cell.alignment = alignment
            # Apply number formatting for numeric columns
            if col_num in [4, 5, 6]:  # Quantity, Unit Price, Total Price
                cell.number_format = '#,##0.00'
        
        # Save the workbook
        wb.save(file_path)
        
        # Update inventory sheet with formulas
        convert_existing_inventory_to_formulas(file_path)
        
        return next_entry_number
    except Exception as e:
        log_exception(f"Failed to add inventory entry: {e}")
        raise


def disburse_inventory_entry(file_path, item_name, quantity, unit_price, notes="", disburse_date=None):
    """
    Add an inventory disbursement entry to the disbursements sheet
    
    Args:
        file_path (str): Path to the Excel file
        item_name (str): Name of the item
        quantity (float): Quantity of the item
        unit_price (float): Price per unit
        notes (str): Additional notes
        disburse_date (str): Date of disbursement (defaults to today)
        
    Returns:
        int: Disbursement entry number
    """
    try:
        # Load workbook
        wb = openpyxl.load_workbook(file_path)
        disburse_sheet = wb["اذن الصرف"]
        
        # Determine the next entry number
        next_entry_number = disburse_sheet.max_row
        
        # Get today's date if not provided
        if disburse_date is None:
            disburse_date = datetime.now().strftime('%d/%m/%Y')
        
        # Calculate total price
        total_price = float(quantity) * float(unit_price)
        
        # Add data row
        row_data = [
            next_entry_number,  # Auto entry number
            disburse_date,
            item_name,
            float(quantity),
            float(unit_price),
            total_price,
            notes
        ]
        
        # Add row to sheet
        disburse_sheet.append(row_data)
        
        # Apply styles to the new row
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        alignment = Alignment(horizontal='center', vertical='center')
        
        row_num = disburse_sheet.max_row
        for col_num, value in enumerate(row_data, 1):
            cell = disburse_sheet.cell(row=row_num, column=col_num)
            cell.border = border
            cell.alignment = alignment
            # Apply number formatting for numeric columns
            if col_num in [4, 5, 6]:  # Quantity, Unit Price, Total Price
                cell.number_format = '#,##0.00'
        
        # Save the workbook
        wb.save(file_path)
        
        # Update inventory sheet with formulas
        convert_existing_inventory_to_formulas(file_path)
        
        return next_entry_number
    except Exception as e:
        log_exception(f"Failed to add disbursement entry: {e}")
        raise


def get_inventory_summary(file_path):
    """
    Get inventory summary data by calculating from additions and disbursements sheets
    
    Args:
        file_path (str): Path to the Excel file
        
    Returns:
        list: List of dictionaries containing inventory data
    """
    if not os.path.exists(file_path):
        return []
    
    try:
        wb = openpyxl.load_workbook(file_path)
        add_sheet = wb["اذن الاضافه"]
        disburse_sheet = wb["اذن الصرف"]
        
        # Calculate additions per item
        additions_by_item = {}
        for row_num in range(2, add_sheet.max_row + 1):
            item_name = add_sheet.cell(row=row_num, column=3).value
            quantity = add_sheet.cell(row=row_num, column=4).value or 0
            if item_name:
                try:
                    quantity = float(quantity)
                except (ValueError, TypeError):
                    quantity = 0
                additions_by_item[item_name] = additions_by_item.get(item_name, 0) + quantity
        
        # Calculate disbursements per item
        disbursements_by_item = {}
        for row_num in range(2, disburse_sheet.max_row + 1):
            item_name = disburse_sheet.cell(row=row_num, column=3).value
            quantity = disburse_sheet.cell(row=row_num, column=4).value or 0
            if item_name:
                try:
                    quantity = float(quantity)
                except (ValueError, TypeError):
                    quantity = 0
                disbursements_by_item[item_name] = disbursements_by_item.get(item_name, 0) + quantity
        
        # Get all unique item names
        all_items = set(additions_by_item.keys()) | set(disbursements_by_item.keys())
        
        inventory_data = []
        for item_name in sorted(all_items):
            total_additions = additions_by_item.get(item_name, 0)
            total_disbursements = disbursements_by_item.get(item_name, 0)
            current_balance = total_additions - total_disbursements
            
            item_data = {
                'item_name': item_name,
                'total_additions': total_additions,
                'total_disbursements': total_disbursements,
                'current_balance': current_balance
            }
            inventory_data.append(item_data)
        
        return inventory_data
    except Exception as e:
        log_exception(f"Error reading inventory summary: {e}")
        return []


def get_available_items_with_prices(file_path):
    """
    Get available items with their average unit prices from additions
    
    Args:
        file_path (str): Path to the Excel file
        
    Returns:
        dict: Dictionary with item names as keys and average unit prices as values
    """
    if not os.path.exists(file_path):
        return {}
    
    try:
        wb = openpyxl.load_workbook(file_path)
        add_sheet = wb["اذن الاضافه"]
        disburse_sheet = wb["اذن الصرف"]
        
        # Calculate additions per item
        additions_by_item = {}
        for row_num in range(2, add_sheet.max_row + 1):
            item_name = add_sheet.cell(row=row_num, column=3).value
            quantity = add_sheet.cell(row=row_num, column=4).value or 0
            if item_name:
                try:
                    quantity = float(quantity)
                except (ValueError, TypeError):
                    quantity = 0
                additions_by_item[item_name] = additions_by_item.get(item_name, 0) + quantity
        
        # Calculate disbursements per item
        disbursements_by_item = {}
        for row_num in range(2, disburse_sheet.max_row + 1):
            item_name = disburse_sheet.cell(row=row_num, column=3).value
            quantity = disburse_sheet.cell(row=row_num, column=4).value or 0
            if item_name:
                try:
                    quantity = float(quantity)
                except (ValueError, TypeError):
                    quantity = 0
                disbursements_by_item[item_name] = disbursements_by_item.get(item_name, 0) + quantity
        
        # Calculate current balances
        inventory_balances = {}
        all_items = set(additions_by_item.keys()) | set(disbursements_by_item.keys())
        for item_name in all_items:
            balance = additions_by_item.get(item_name, 0) - disbursements_by_item.get(item_name, 0)
            inventory_balances[item_name] = balance
        
        # Get item prices from additions
        item_prices = {}
        item_quantities = {}
        
        for row_num in range(2, add_sheet.max_row + 1):
            item_name = add_sheet.cell(row=row_num, column=3).value
            quantity = add_sheet.cell(row=row_num, column=4).value or 0
            unit_price = add_sheet.cell(row=row_num, column=5).value or 0
            
            try:
                quantity = float(quantity)
                unit_price = float(unit_price)
            except (ValueError, TypeError):
                continue
            
            if item_name and quantity > 0:
                # Only include items that have positive balance
                if inventory_balances.get(item_name, 0) > 0:
                    if item_name not in item_prices:
                        item_prices[item_name] = 0
                        item_quantities[item_name] = 0
                    
                    item_prices[item_name] += unit_price * quantity
                    item_quantities[item_name] += quantity
        
        # Calculate average prices
        avg_prices = {}
        for item_name in item_prices:
            if item_quantities[item_name] > 0:
                avg_prices[item_name] = item_prices[item_name] / item_quantities[item_name]
            else:
                avg_prices[item_name] = 0
        
        return avg_prices
    except Exception as e:
        log_exception(f"Error getting available items with prices: {e}")
        return {}