"""
Attendance Utilities for Employee Tracking
This module provides functions to create and manage attendance Excel files.
"""

import xlsxwriter
from xlsxwriter.utility import xl_col_to_name
import os
from typing import List, Dict, Optional, Tuple
from datetime import datetime, timedelta


def create_or_update_attendance(filepath: str, employees_data: List[Dict]) -> Tuple[bool, Optional[str]]:
    """
    Create or update attendance Excel file with weekly schedule.
    
    Args:
        filepath (str): Path to save the Excel file
        employees_data (list): List of employee dictionaries with structure:
            {
                'name': str,
                'friday_shift1': float,
                'friday_shift2': float,
                'saturday_shift1': float,
                'saturday_shift2': float,
                'sunday_shift1': float,
                'sunday_shift2': float,
                'monday_shift1': float,
                'monday_shift2': float,
                'tuesday_shift1': float,
                'tuesday_shift2': float,
                'wednesday_shift1': float,
                'wednesday_shift2': float,
                'thursday_shift1': float,
                'thursday_shift2': float,
                'date': str,
                'advance': float,
                'price': float
            }
    
    Returns:
        tuple: (success: bool, error_message: str or None)
    """
    try:
        # Always rebuild the workbook to keep all sections consistent
        return create_new_attendance_file(filepath, employees_data)
    except Exception as e:
        return (False, f"error: {str(e)}")


def create_new_attendance_file(filepath: str, employees_data: List[Dict]) -> Tuple[bool, Optional[str]]:
    """Create a new attendance Excel file"""
    try:
        workbook = xlsxwriter.Workbook(filepath)
        worksheet = workbook.add_worksheet("الحضور والانصراف")
        
        # RTL + Page settings
        worksheet.right_to_left()
        worksheet.hide_gridlines(2)
        
        # Page settings
        worksheet.set_paper(9)  # A4
        worksheet.set_landscape()
        worksheet.set_margins(0.3, 0.3, 0.5, 0.5)
        
        # Formats
        header_day_fmt = workbook.add_format({
            'bold': True,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'font_name': 'Arial',
            'font_size': 12,
            'bg_color': '#4472C4',
            'font_color': '#FFFFFF'
        })
        
        header_shift_fmt = workbook.add_format({
            'bold': True,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'font_name': 'Arial',
            'font_size': 10,
            'bg_color': '#D9E1F2',
            'font_color': '#000000'
        })
        
        header_name_fmt = workbook.add_format({
            'bold': True,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'font_name': 'Arial',
            'font_size': 12,
            'bg_color': '#4472C4',
            'font_color': '#FFFFFF'
        })
        
        name_fmt = workbook.add_format({
            'bold': True,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'font_name': 'Arial',
            'font_size': 11,
            'bg_color': '#E7E6E6'
        })
        
        data_fmt = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'font_name': 'Arial',
            'font_size': 10
        })
        
        total_fmt = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'font_name': 'Arial',
            'font_size': 10,
            'bold': True,
            'bg_color': '#FFF2CC',
            'num_format': '0'
        })
        
        weekly_total_fmt = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'font_name': 'Arial',
            'font_size': 10,
            'bold': True,
            'bg_color': '#FFF2CC',
            # Hide zeros: show positive, negative, blank for zero
            'num_format': '#;-#;;'
        })
        
        weekly_total_label_fmt = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'font_name': 'Arial',
            'font_size': 10,
            'bold': True,
            'bg_color': '#FFF2CC'
        })
        
        special_fmt = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'font_name': 'Arial',
            'font_size': 10,
            'bg_color': '#E2EFDA'
        })
        
        arabic_digit_map = str.maketrans('0123456789', '٠١٢٣٤٥٦٧٨٩')
        
        day_definitions = [
            ('الجمعة', 'friday', 4),
            ('السبت', 'saturday', 5),
            ('الأحد', 'sunday', 6),
            ('الاثنين', 'monday', 0),
            ('الثلاثاء', 'tuesday', 1),
            ('الأربعاء', 'wednesday', 2),
            ('الخميس', 'thursday', 3)
        ]
        days = [info[0] for info in day_definitions]
        day_offsets = [info[2] for info in day_definitions]
        shift_keys = [
            'friday_shift1', 'friday_shift2',
            'saturday_shift1', 'saturday_shift2',
            'sunday_shift1', 'sunday_shift2',
            'monday_shift1', 'monday_shift2',
            'tuesday_shift1', 'tuesday_shift2',
            'wednesday_shift1', 'wednesday_shift2',
            'thursday_shift1', 'thursday_shift2'
        ]
        
        def parse_date_value(value: str) -> Optional[datetime]:
            if not value:
                return None
            try:
                return datetime.strptime(str(value), '%d/%m/%Y')
            except Exception:
                return None
        
        def build_week_dates(week_start: Optional[datetime]) -> List[str]:
            if not week_start:
                return ['' for _ in day_offsets]
            return [
                (week_start + timedelta(days=offset)).strftime('%Y/%m/%d').translate(arabic_digit_map)
                for offset in day_offsets
            ]
        
        def merge_records_by_employee(records: List[Dict]) -> List[Dict]:
            merged: Dict[str, Dict] = {}
            for emp in records:
                name = str(emp.get('name', '') or '').strip()
                if not name:
                    name = 'بدون اسم'
                
                if name not in merged:
                    merged[name] = {
                        'name': name,
                        'date': emp.get('date', ''),
                        'advance': emp.get('advance', 0),
                        'price': emp.get('price', 0)
                    }
                    for key in shift_keys:
                        merged[name][key] = 0
                
                for key in shift_keys:
                    value = emp.get(key, 0)
                    if value:
                        merged[name][key] = value
                
                if emp.get('advance', 0):
                    merged[name]['advance'] = emp.get('advance', 0)
                if emp.get('price', 0):
                    merged[name]['price'] = emp.get('price', 0)
                if emp.get('date', '') and not merged[name].get('date'):
                    merged[name]['date'] = emp.get('date', '')
            return list(merged.values())
        
        def write_section(section_row: int, week_dates: List[str], records: List[Dict]) -> int:
            """Write a full attendance section starting from section_row and return next row index."""
            # No separate date title row; headers start directly at section_row
            header_top_row = section_row
            
            # Row: Employee name header + Day names
            worksheet.merge_range(header_top_row, 0, header_top_row + 1, 0, 'اسم الموظف', header_name_fmt)
            
            col = 1
            for idx, day in enumerate(days):
                day_label = day
                if week_dates and week_dates[idx]:
                    day_label = f"{day} {week_dates[idx]}"
                worksheet.merge_range(header_top_row, col, header_top_row, col + 1, day_label, header_day_fmt)
                col += 2
            
            # Total, Date, Advance headers (no separate price column header)
            worksheet.merge_range(header_top_row, 15, header_top_row + 1, 15, 'الإجمالي', header_day_fmt)
            worksheet.merge_range(header_top_row, 16, header_top_row + 1, 16, 'التاريخ', header_day_fmt)
            worksheet.merge_range(header_top_row, 17, header_top_row + 1, 17, 'السلفة', header_day_fmt)
            
            # Row: Shift labels
            col = 1
            shift_row = header_top_row + 1
            for _ in days:
                worksheet.write(shift_row, col, 'وردية 1', header_shift_fmt)
                worksheet.write(shift_row, col + 1, 'وردية 2', header_shift_fmt)
                col += 2
            
            # Data rows
            data_row = shift_row + 1
            current_row = data_row
            
            merged_records = merge_records_by_employee(records)
            sorted_records = sorted(merged_records, key=lambda emp: emp.get('name', ''))
            for emp in sorted_records:
                worksheet.write(current_row, 0, emp.get('name', ''), name_fmt)
                
                col = 1
                for key in shift_keys:
                    value = emp.get(key, 0)
                    if value:
                        worksheet.write_number(current_row, col, float(value), data_fmt)
                    else:
                        worksheet.write(current_row, col, '', data_fmt)
                    col += 1
                
                excel_row = current_row + 1
                worksheet.write_formula(
                    current_row,
                    15,
                    f"=SUM({xl_col_to_name(1)}{excel_row}:{xl_col_to_name(14)}{excel_row})-{xl_col_to_name(17)}{excel_row}",
                    total_fmt
                )
                
                worksheet.write(current_row, 16, emp.get('date', ''), special_fmt)
                
                advance = emp.get('advance', 0)
                if advance:
                    worksheet.write_number(current_row, 17, float(advance), special_fmt)
                else:
                    worksheet.write(current_row, 17, '', special_fmt)
                
                current_row += 1
            
            data_end_row = current_row - 1
            weekly_total_row = current_row
            worksheet.write(weekly_total_row, 0, 'الإجمالي الأسبوعي', weekly_total_label_fmt)
            
            if data_end_row >= data_row:
                start_excel_row = data_row + 1
                end_excel_row = data_end_row + 1
                
                for col_idx in range(1, 15):
                    col_letter = xl_col_to_name(col_idx)
                    worksheet.write_formula(
                        weekly_total_row,
                        col_idx,
                        f"=SUM({col_letter}{start_excel_row}:{col_letter}{end_excel_row})",
                        weekly_total_fmt
                    )
                
                worksheet.write_formula(
                    weekly_total_row,
                    15,
                    f"=SUM({xl_col_to_name(15)}{start_excel_row}:{xl_col_to_name(15)}{end_excel_row})",
                    weekly_total_fmt
                )
                worksheet.write_formula(
                    weekly_total_row,
                    17,
                    f"=SUM({xl_col_to_name(17)}{start_excel_row}:{xl_col_to_name(17)}{end_excel_row})",
                    weekly_total_fmt
                )
                # No overall price total column; only attendance, total, and advance are summed
            else:
                # No data rows; leave totals blank to avoid invalid formulas
                for col_idx in range(1, 19):
                    worksheet.write(weekly_total_row, col_idx, '', total_fmt)
            
            # Leave an empty row before the next section
            return weekly_total_row + 2
        
        if not employees_data:
            worksheet.merge_range(0, 0, 0, 5, 'لا توجد بيانات حضور', header_name_fmt)
            workbook.close()
            return (True, None)
        
        # Group records by week (so days of the same week share one table)
        grouped_by_week = {}
        week_sort_key = {}
        
        for emp in employees_data:
            raw_date = str(emp.get('date', '') or '')
            date_obj = parse_date_value(raw_date)
            
            if date_obj:
                week_start = date_obj - timedelta(days=date_obj.weekday())
                week_key = ('week', week_start)
                sort_date = week_start
            else:
                week_start = None
                week_key = ('no_date', raw_date)
                sort_date = datetime.max
            
            bucket = grouped_by_week.setdefault(week_key, {'records': [], 'week_start': week_start})
            bucket['records'].append(emp)
            
            if bucket['week_start'] is None and week_start:
                bucket['week_start'] = week_start
            
            if week_key not in week_sort_key or sort_date < week_sort_key[week_key]:
                week_sort_key[week_key] = sort_date
        
        # Sort sections by week start date
        sorted_sections = sorted(
            grouped_by_week.items(),
            key=lambda item: week_sort_key.get(item[0], datetime.max)
        )
        
        current_row = 0
        for _, bucket in sorted_sections:
            week_dates = build_week_dates(bucket.get('week_start'))
            current_row = write_section(current_row, week_dates, bucket['records'])
        
        # Column widths
        worksheet.set_column(0, 0, 20)
        worksheet.set_column(1, 14, 11)
        worksheet.set_column(15, 15, 10)
        worksheet.set_column(16, 16, 12, None, {'hidden': True})
        worksheet.set_column(17, 17, 10)
        worksheet.set_column(18, 18, 2, None, {'hidden': True})
        
        workbook.close()
        return (True, None)
        
    except PermissionError:
        return (False, "file_locked")
    except Exception as e:
        return (False, f"error: {str(e)}")


def append_to_existing_attendance(filepath: str, new_employees_data: List[Dict]) -> Tuple[bool, Optional[str]]:
    """Append new attendance data to existing Excel file with proper weekly table format"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Load existing workbook
        workbook = openpyxl.load_workbook(filepath)
        worksheet = workbook.active
        
        if worksheet is None:
            return (False, "invalid_sheet")
        
        # Find the last row with data
        last_row = worksheet.max_row
        
        # Check if we need to add a separator
        if last_row > 2:  # If there's existing data
            # Add a blank row as separator
            last_row += 1
        
        # Add new weekly table
        data_start_row = last_row + 1
        
        # Add new data with proper formatting
        for emp in new_employees_data:
            # Employee name
            cell = worksheet.cell(row=last_row, column=1, value=emp.get('name', ''))
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Attendance data
            shift_keys = [
                'friday_shift1', 'friday_shift2',
                'saturday_shift1', 'saturday_shift2',
                'sunday_shift1', 'sunday_shift2',
                'monday_shift1', 'monday_shift2',
                'tuesday_shift1', 'tuesday_shift2',
                'wednesday_shift1', 'wednesday_shift2',
                'thursday_shift1', 'thursday_shift2'
            ]
            
            col = 2  # Column B
            for key in shift_keys:
                value = emp.get(key, 0)
                cell = worksheet.cell(row=last_row, column=col, value=float(value) if value else 0)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                col += 1
            
            # Date
            cell = worksheet.cell(row=last_row, column=17, value=emp.get('date', ''))
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Advance
            advance = emp.get('advance', 0)
            cell = worksheet.cell(row=last_row, column=18, value=float(advance) if advance else 0)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Price
            price = emp.get('price', 0)
            cell = worksheet.cell(row=last_row, column=19, value=float(price) if price else 0)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            last_row += 1
        
        # Add weekly total row
        weekly_total_row = last_row
        cell = worksheet.cell(row=weekly_total_row, column=1, value='الإجمالي الأسبوعي')
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Calculate weekly totals for each column
        data_end_row = weekly_total_row - 1
        for col_idx in range(2, 16):  # Columns B to O (attendance data)
            col_letter = get_column_letter(col_idx)
            formula = f'=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})'
            cell = worksheet.cell(row=weekly_total_row, column=col_idx, value=formula)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Total formula (P column)
        cell = worksheet.cell(row=weekly_total_row, column=16, value=f'=SUM(P{data_start_row}:P{data_end_row})')
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Advance total (R column)
        cell = worksheet.cell(row=weekly_total_row, column=18, value=f'=SUM(R{data_start_row}:R{data_end_row})')
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Price total (S column)
        cell = worksheet.cell(row=weekly_total_row, column=19, value=f'=SUM(S{data_start_row}:S{data_end_row})')
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Save the workbook
        workbook.save(filepath)
        return (True, None)
        
    except ImportError:
        return (False, "openpyxl_missing")
    except PermissionError:
        return (False, "file_locked")
    except Exception as e:
        return (False, f"error: {str(e)}")


def load_attendance_data(filepath: str) -> Tuple[bool, Optional[List[Dict]], Optional[str]]:
    """
    Load attendance data from an existing Excel file.
    
    Args:
        filepath (str): Path to the Excel file
    
    Returns:
        tuple: (success: bool, data: List[Dict] or None, error_message: str or None)
    """
    if not os.path.exists(filepath):
        return (False, None, "file_not_found")
    
    try:
        import openpyxl
        
        workbook = openpyxl.load_workbook(filepath, data_only=True)
        sheet = workbook.active
        
        if sheet is None:
            return (False, None, "invalid_sheet")
        
        employees_data = []
        
        max_row = sheet.max_row
        row_idx = 1  # Excel rows are 1-based
        
        while row_idx <= max_row:
            cell_value = sheet.cell(row=row_idx, column=1).value
            
            # Skip empty rows
            if cell_value is None:
                row_idx += 1
                continue
            
            text_value = str(cell_value).strip()
            
            if not text_value:
                row_idx += 1
                continue
            
            # Skip section headers - look for date headers like "التاريخ:"
            if text_value.startswith('التاريخ:'):
                row_idx += 1
                continue
            
            # Skip table headers - "اسم الموظف" indicates start of table headers
            if text_value == 'اسم الموظف':
                # Skip the merged header row and the shift labels row
                row_idx += 2
                continue
            
            # Skip total rows and separators
            if text_value in ('الإجمالي الأسبوعي', '----', 'لا توجد بيانات حضور'):
                row_idx += 1
                continue
            
            # This should be an employee data row
            name = text_value
            
            # Read all the data from this row
            try:
                date_value = sheet.cell(row=row_idx, column=17).value
                
                emp_data = {
                    'name': str(name),
                    'friday_shift1': sheet.cell(row=row_idx, column=2).value or 0,
                    'friday_shift2': sheet.cell(row=row_idx, column=3).value or 0,
                    'saturday_shift1': sheet.cell(row=row_idx, column=4).value or 0,
                    'saturday_shift2': sheet.cell(row=row_idx, column=5).value or 0,
                    'sunday_shift1': sheet.cell(row=row_idx, column=6).value or 0,
                    'sunday_shift2': sheet.cell(row=row_idx, column=7).value or 0,
                    'monday_shift1': sheet.cell(row=row_idx, column=8).value or 0,
                    'monday_shift2': sheet.cell(row=row_idx, column=9).value or 0,
                    'tuesday_shift1': sheet.cell(row=row_idx, column=10).value or 0,
                    'tuesday_shift2': sheet.cell(row=row_idx, column=11).value or 0,
                    'wednesday_shift1': sheet.cell(row=row_idx, column=12).value or 0,
                    'wednesday_shift2': sheet.cell(row=row_idx, column=13).value or 0,
                    'thursday_shift1': sheet.cell(row=row_idx, column=14).value or 0,
                    'thursday_shift2': sheet.cell(row=row_idx, column=15).value or 0,
                    'date': str(date_value) if date_value else '',
                    'advance': sheet.cell(row=row_idx, column=18).value or 0,
                    'price': sheet.cell(row=row_idx, column=19).value or 0
                }
                
                employees_data.append(emp_data)
                
            except Exception as e:
                pass
            
            row_idx += 1
        
        return (True, employees_data, None)
        
    except ImportError:
        return (False, None, "openpyxl_missing")
    except Exception as e:
        return (False, None, f"error: {str(e)}")
        