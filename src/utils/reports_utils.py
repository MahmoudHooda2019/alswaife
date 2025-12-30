"""
Reports Utilities - AI-powered report generation
Uses Groq/Gemini to parse user requests and Pandas to process Excel data
Supports all sections: Income, Expenses, Inventory, Attendance, Blocks, Slides
"""

import os
import json
import pandas as pd
from datetime import datetime
from typing import Dict, List, Optional
import xlsxwriter

# Load environment variables from .env file
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# API configuration - Groq as primary, Gemini as fallback
GROQ_API_KEY = os.environ.get("GROQ_API_KEY", "")
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")

REPORT_PROMPT = """أنت مساعد لتحليل طلبات التقارير. قم بتحويل طلب المستخدم إلى JSON بالتنسيق التالي:
{
    "report_type": نوع التقرير (انظر القائمة أدناه),
    "date_from": "DD/MM/YYYY" أو null,
    "date_to": "DD/MM/YYYY" أو null,
    "client_name": "اسم العميل" أو null,
    "item_name": "اسم الصنف" أو null,
    "block_number": "رقم البلوك" أو null,
    "machine_number": "رقم الماكينة" أو null,
    "filters": {},
    "group_by": "day" أو "month" أو "client" أو "item" أو null,
    "sort_by": "date" أو "amount" أو null,
    "sort_order": "asc" أو "desc"
}

أنواع التقارير المتاحة:
- income: تقرير الإيرادات
- expenses: تقرير المصروفات
- inventory: تقرير المخزون (الرصيد الحالي)
- inventory_add: تقرير إضافات المخزون
- inventory_disburse: تقرير صرف المخزون
- attendance: تقرير الحضور والانصراف
- blocks: تقرير البلوكات
- slides: تقرير الشرائح
- machine_production: تقرير إنتاج ماكينة معينة (استخدمه عند ذكر مكنه أو ماكينه أو انتاج مكينه)
- clients: تقرير العملاء

ملاحظات مهمة:
- إذا ذكر المستخدم "مكنه" أو "ماكينه" أو "انتاج" استخدم report_type = "machine_production" واستخرج رقم الماكينة في machine_number
- السنة الحالية هي 2025. إذا لم يذكر المستخدم السنة، استخدم 2025.
- أرجع JSON فقط بدون أي نص إضافي."""


def parse_with_groq(user_request: str) -> Dict:
    """Parse user request using Groq API (fast and free)"""
    try:
        from groq import Groq
        client = Groq(api_key=GROQ_API_KEY)
        
        response = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[
                {"role": "system", "content": REPORT_PROMPT},
                {"role": "user", "content": user_request}
            ],
            temperature=0.1,
            max_tokens=500
        )
        
        result = response.choices[0].message.content.strip()
        print(f"[DEBUG] Groq response: {result}")
        
        # Extract JSON from response
        if "```json" in result:
            result = result.split("```json")[1].split("```")[0]
        elif "```" in result:
            result = result.split("```")[1].split("```")[0]
        
        # Find JSON object in response
        start = result.find("{")
        end = result.rfind("}") + 1
        if start != -1 and end > start:
            result = result[start:end]
        
        return json.loads(result)
    
    except Exception as e:
        print(f"[ERROR] Groq parsing failed: {e}")
        return None


def parse_with_gemini(user_request: str) -> Dict:
    """Parse user request using Google Gemini API"""
    try:
        import google.generativeai as genai
        genai.configure(api_key=GEMINI_API_KEY)
        
        model = genai.GenerativeModel('gemini-1.5-flash')
        response = model.generate_content(f"{REPORT_PROMPT}\n\nطلب المستخدم: {user_request}")
        
        result = response.text.strip()
        print(f"[DEBUG] Gemini response: {result}")
        
        # Extract JSON from response
        if "```json" in result:
            result = result.split("```json")[1].split("```")[0]
        elif "```" in result:
            result = result.split("```")[1].split("```")[0]
        
        # Find JSON object in response
        start = result.find("{")
        end = result.rfind("}") + 1
        if start != -1 and end > start:
            result = result[start:end]
        
        return json.loads(result)
    
    except Exception as e:
        print(f"[ERROR] Gemini parsing failed: {e}")
        return None


def parse_user_request_with_ai(user_request: str) -> Dict:
    """
    Use AI to parse user's natural language request into structured JSON.
    Tries Groq first (fastest), then Gemini, then falls back to simple parser.
    """
    # Try Groq first (very fast and free)
    if GROQ_API_KEY:
        result = parse_with_groq(user_request)
        if result:
            return result
    
    # Try Gemini as fallback
    if GEMINI_API_KEY:
        result = parse_with_gemini(user_request)
        if result:
            return result
    
    # Fall back to simple parser
    return parse_request_simple(user_request)


def parse_request_simple(user_request: str) -> Dict:
    """Simple fallback parser without AI"""
    import re
    
    result = {
        "report_type": "income",
        "date_from": None,
        "date_to": None,
        "client_name": None,
        "item_name": None,
        "block_number": None,
        "machine_number": None,
        "filters": {},
        "group_by": None,
        "sort_by": "date",
        "sort_order": "desc"
    }
    
    # Detect report type
    if "مكن" in user_request or "مكينه" in user_request or "ماكينه" in user_request or "انتاج" in user_request:
        result["report_type"] = "machine_production"
        # Extract machine number
        machine_pattern = r'(?:مكن|مكينه|ماكينه|machine)\s*(\d+)'
        machine_match = re.search(machine_pattern, user_request, re.IGNORECASE)
        if machine_match:
            result["machine_number"] = machine_match.group(1)
        else:
            # Try to find standalone number after machine keywords
            num_pattern = r'(\d+)'
            nums = re.findall(num_pattern, user_request)
            if nums:
                result["machine_number"] = nums[0]
    elif "مصروف" in user_request or "صرف" in user_request:
        if "مخزون" in user_request:
            result["report_type"] = "inventory_disburse"
        else:
            result["report_type"] = "expenses"
    elif "إيراد" in user_request or "ايراد" in user_request or "دخل" in user_request:
        result["report_type"] = "income"
    elif "عميل" in user_request or "عملاء" in user_request:
        result["report_type"] = "clients"
    elif "مخزون" in user_request or "مخزن" in user_request:
        if "اضاف" in user_request:
            result["report_type"] = "inventory_add"
        else:
            result["report_type"] = "inventory"
    elif "حضور" in user_request or "انصراف" in user_request:
        result["report_type"] = "attendance"
    elif "بلوك" in user_request or "بلوكات" in user_request:
        result["report_type"] = "blocks"
    elif "شرائح" in user_request or "شريحة" in user_request:
        result["report_type"] = "slides"
    
    # Extract dates
    date_pattern = r'(\d{1,2})[/\\-](\d{1,2})(?:[/\\-](\d{2,4}))?'
    dates = re.findall(date_pattern, user_request)
    
    current_year = datetime.now().year
    
    if len(dates) >= 1:
        d, m, y = dates[0][0], dates[0][1], dates[0][2] or str(current_year)
        if len(y) == 2:
            y = "20" + y
        result["date_from"] = f"{d.zfill(2)}/{m.zfill(2)}/{y}"
    
    if len(dates) >= 2:
        d, m, y = dates[1][0], dates[1][1], dates[1][2] or str(current_year)
        if len(y) == 2:
            y = "20" + y
        result["date_to"] = f"{d.zfill(2)}/{m.zfill(2)}/{y}"
    
    return result


# ============ DATA LOADING FUNCTIONS ============

def load_income_expenses_data(documents_path: str):
    """Load income and expenses data from Excel file (3 sheets version)"""
    filepath = os.path.join(documents_path, "ايرادات ومصروفات", "بيان مصروفات وايرادات مصنع جرانيت السويفى.xlsx")
    
    if not os.path.exists(filepath):
        return pd.DataFrame(), pd.DataFrame()
    
    try:
        df_income = pd.read_excel(filepath, sheet_name='الإيرادات', skiprows=1, header=None)
        df_income.columns = ["رقم الفاتورة", "العميل", "المبلغ", "التاريخ"]
        df_income["النوع"] = "إيراد"
        df_income = df_income.dropna(subset=["رقم الفاتورة"])
        
        df_expenses = pd.read_excel(filepath, sheet_name='المصروفات', skiprows=1, header=None)
        df_expenses.columns = ["العدد", "البيان", "المبلغ", "التاريخ"]
        df_expenses["النوع"] = "مصروف"
        df_expenses = df_expenses.dropna(subset=["البيان"])
        
        return df_income, df_expenses
    except Exception as e:
        print(f"[ERROR] Failed to load income/expenses data: {e}")
        return pd.DataFrame(), pd.DataFrame()


def load_inventory_data(documents_path: str):
    """Load inventory data from Excel file"""
    filepath = os.path.join(documents_path, "مخزون الادوات", "مخزون ادوات التشغيل.xlsx")
    
    if not os.path.exists(filepath):
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    
    try:
        # Load additions first
        df_add = pd.read_excel(filepath, sheet_name='اذن الاضافه', skiprows=0)
        if len(df_add.columns) >= 7:
            df_add.columns = ["رقم الإذن", "التاريخ", "اسم الصنف", "العدد", "سعر الوحدة", "الإجمالي", "ملاحظات"]
        df_add = df_add.dropna(subset=["اسم الصنف"])
        
        # Load disbursements
        df_disburse = pd.read_excel(filepath, sheet_name='اذن الصرف', skiprows=0)
        if len(df_disburse.columns) >= 7:
            df_disburse.columns = ["رقم الإذن", "التاريخ", "اسم الصنف", "العدد", "سعر الوحدة", "الإجمالي", "ملاحظات"]
        df_disburse = df_disburse.dropna(subset=["اسم الصنف"])
        
        # Calculate inventory from additions and disbursements
        # Get all unique item names
        all_items = set()
        if not df_add.empty and "اسم الصنف" in df_add.columns:
            all_items.update(df_add["اسم الصنف"].dropna().unique())
        if not df_disburse.empty and "اسم الصنف" in df_disburse.columns:
            all_items.update(df_disburse["اسم الصنف"].dropna().unique())
        
        # Calculate totals for each item
        inv_data = []
        for item in sorted(all_items):
            # Sum additions
            total_add = 0
            if not df_add.empty and "اسم الصنف" in df_add.columns and "العدد" in df_add.columns:
                item_adds = df_add[df_add["اسم الصنف"] == item]["العدد"]
                total_add = pd.to_numeric(item_adds, errors='coerce').sum()
            
            # Sum disbursements
            total_disburse = 0
            if not df_disburse.empty and "اسم الصنف" in df_disburse.columns and "العدد" in df_disburse.columns:
                item_disburse = df_disburse[df_disburse["اسم الصنف"] == item]["العدد"]
                total_disburse = pd.to_numeric(item_disburse, errors='coerce').sum()
            
            # Calculate balance
            balance = total_add - total_disburse
            
            # Round to integers if whole numbers, otherwise 2 decimal places
            total_add = int(total_add) if pd.notna(total_add) and total_add == int(total_add) else round(total_add, 2) if pd.notna(total_add) else 0
            total_disburse = int(total_disburse) if pd.notna(total_disburse) and total_disburse == int(total_disburse) else round(total_disburse, 2) if pd.notna(total_disburse) else 0
            balance = int(balance) if pd.notna(balance) and balance == int(balance) else round(balance, 2) if pd.notna(balance) else 0
            
            inv_data.append({
                "اسم الصنف": item,
                "إجمالي الإضافات": total_add,
                "إجمالي الصرف": total_disburse,
                "الرصيد الحالي": balance
            })
        
        df_inventory = pd.DataFrame(inv_data)
        
        return df_inventory, df_add, df_disburse
    except Exception as e:
        print(f"[ERROR] Failed to load inventory data: {e}")
        import traceback
        traceback.print_exc()
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()


def load_attendance_data(documents_path: str) -> pd.DataFrame:
    """Load attendance data from Excel file"""
    filepath = os.path.join(documents_path, "حضور وانصراف", "سجل الحضور والانصراف.xlsx")
    
    if not os.path.exists(filepath):
        return pd.DataFrame()
    
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet = wb.active
        
        data = []
        for row_idx in range(3, sheet.max_row + 1):
            name = sheet.cell(row=row_idx, column=1).value
            if not name or str(name).strip() in ['', 'اسم الموظف', 'الإجمالي الأسبوعي']:
                continue
            
            # Sum all shifts
            total_shifts = 0
            for col in range(2, 16):
                val = sheet.cell(row=row_idx, column=col).value
                if val and isinstance(val, (int, float)):
                    total_shifts += val
            
            total = sheet.cell(row=row_idx, column=16).value or 0
            date = sheet.cell(row=row_idx, column=17).value or ''
            advance = sheet.cell(row=row_idx, column=18).value or 0
            
            data.append({
                "الاسم": str(name).strip(),
                "إجمالي الورديات": total_shifts,
                "الإجمالي": total,
                "التاريخ": str(date),
                "السلفة": advance
            })
        
        return pd.DataFrame(data)
    except Exception as e:
        print(f"[ERROR] Failed to load attendance data: {e}")
        return pd.DataFrame()


def load_blocks_data(documents_path: str) -> pd.DataFrame:
    """Load blocks data from Excel file"""
    filepath = os.path.join(documents_path, "البلوكات", "مخزون البلوكات.xlsx")
    
    if not os.path.exists(filepath):
        return pd.DataFrame()
    
    try:
        df = pd.read_excel(filepath, sheet_name='البلوكات', skiprows=1)
        # Rename columns based on TABLE1_COLUMNS
        expected_cols = ["رقم النقله", "عدد النقله", "التاريخ", "المحجر", 
                        "رقم البلوك", "الخامه", "الطول", "العرض", "الارتفاع", 
                        "م3", "الوزن", "وزن البلوك", "سعر الطن", "اجمالي السعر"]
        if len(df.columns) >= len(expected_cols):
            df.columns = expected_cols + list(df.columns[len(expected_cols):])
        df = df.dropna(subset=["رقم البلوك"])
        return df
    except Exception as e:
        print(f"[ERROR] Failed to load blocks data: {e}")
        return pd.DataFrame()


def load_slides_data(documents_path: str) -> pd.DataFrame:
    """Load slides data from Excel file"""
    filepath = os.path.join(documents_path, "البلوكات", "مخزون البلوكات.xlsx")
    
    if not os.path.exists(filepath):
        return pd.DataFrame()
    
    try:
        df = pd.read_excel(filepath, sheet_name='البلوكات', skiprows=1)
        # Slides start from column 15 (index 14)
        slides_cols = ["تاريخ النشر", "رقم البلوك", "النوع", "رقم المكينه",
                      "وقت الدخول", "وقت الخروج", "عدد الساعات",
                      "السمك", "العدد", "الطول", "الخصم", "الطول بعد",
                      "الارتفاع", "الكمية م2", "سعر المتر", "اجمالي السعر"]
        
        if len(df.columns) >= 15 + len(slides_cols):
            slides_df = df.iloc[:, 15:15+len(slides_cols)].copy()
            slides_df.columns = slides_cols
            slides_df = slides_df.dropna(subset=["رقم البلوك"])
            return slides_df
        return pd.DataFrame()
    except Exception as e:
        print(f"[ERROR] Failed to load slides data: {e}")
        return pd.DataFrame()


def load_machine_production_data(documents_path: str, machine_number: str = None) -> pd.DataFrame:
    """Load machine production data from slides, optionally filtered by machine number"""
    # Try slides folder first (new location)
    slides_filepath = os.path.join(documents_path, "الشرائح", "مخزون الشرائح.xlsx")
    # Fallback to blocks folder (old location)
    blocks_filepath = os.path.join(documents_path, "البلوكات", "مخزون البلوكات.xlsx")
    
    print(f"[DEBUG] load_machine_production_data")
    print(f"[DEBUG] machine_number: {machine_number}")
    print(f"[DEBUG] slides_filepath: {slides_filepath}")
    print(f"[DEBUG] blocks_filepath: {blocks_filepath}")
    
    df = pd.DataFrame()
    
    # Try slides file first
    if os.path.exists(slides_filepath):
        try:
            df = pd.read_excel(slides_filepath, sheet_name='اذن اضافة الشرائح', skiprows=0)
            print(f"[DEBUG] Read from slides file - shape: {df.shape}")
            print(f"[DEBUG] Columns: {list(df.columns)}")
            
            # Expected columns: رقم الإذن, التاريخ, النوع, رقم البلوك, رقم المكينه, وقت الدخول, وقت الخروج, السمك, العدد, الطول, الخصم, الطول بعد, الارتفاع, الكمية م2, سعر المتر, اجمالي السعر
            # Machine column is typically column 5 (index 4) - رقم المكينه
            
            if not df.empty:
                # Find machine column
                machine_col = None
                for col in df.columns:
                    col_str = str(col).strip()
                    if 'مكينه' in col_str or 'مكينة' in col_str or 'ماكينه' in col_str or 'ماكينة' in col_str or 'المكينه' in col_str:
                        machine_col = col
                        break
                
                print(f"[DEBUG] Machine column found: {machine_col}")
                
                # Drop header row if it exists in data
                if len(df) > 0:
                    first_row = df.iloc[0]
                    if any('رقم' in str(v) or 'الإذن' in str(v) for v in first_row.values if pd.notna(v)):
                        df = df.iloc[1:]
                
                # Filter by machine number
                if machine_number and machine_col and machine_col in df.columns:
                    print(f"[DEBUG] Filtering by machine: {machine_number}")
                    print(f"[DEBUG] Machine column unique values: {df[machine_col].unique()}")
                    
                    df[machine_col] = df[machine_col].astype(str).str.strip()
                    machine_str = str(machine_number).strip()
                    
                    # Try exact match
                    filtered_df = df[df[machine_col] == machine_str]
                    
                    # If no results, try contains
                    if filtered_df.empty:
                        filtered_df = df[df[machine_col].str.contains(machine_str, na=False)]
                    
                    df = filtered_df
                    print(f"[DEBUG] After filter shape: {df.shape}")
                
                if not df.empty:
                    return df
        except Exception as e:
            print(f"[DEBUG] Error reading slides file: {e}")
    
    # Fallback to blocks file
    if os.path.exists(blocks_filepath):
        try:
            # Try to read the slides sheet directly first
            try:
                df = pd.read_excel(blocks_filepath, sheet_name='الشرائح', skiprows=1)
                print(f"[DEBUG] Read from 'الشرائح' sheet in blocks file")
            except:
                # Fall back to البلوكات sheet
                df = pd.read_excel(blocks_filepath, sheet_name='البلوكات', skiprows=1)
                print(f"[DEBUG] Read from 'البلوكات' sheet")
            
            print(f"[DEBUG] DataFrame shape: {df.shape}")
            print(f"[DEBUG] DataFrame columns: {list(df.columns)}")
            
            # Look for machine column
            machine_col = None
            for col in df.columns:
                col_str = str(col).strip()
                if 'مكينه' in col_str or 'مكينة' in col_str or 'ماكينه' in col_str or 'ماكينة' in col_str:
                    machine_col = col
                    break
            
            if machine_col is None:
                # Try to find slides data in columns 14+
                slides_cols = ["تاريخ النشر", "رقم البلوك", "النوع", "رقم المكينه",
                              "وقت الدخول", "وقت الخروج", "عدد الساعات",
                              "السمك", "العدد", "الطول", "الخصم", "الطول بعد",
                              "الارتفاع", "الكمية م2", "سعر المتر", "اجمالي السعر"]
                
                slides_start = 14
                
                if len(df.columns) >= slides_start + len(slides_cols):
                    slides_df = df.iloc[:, slides_start:slides_start+len(slides_cols)].copy()
                    slides_df.columns = slides_cols
                    df = slides_df
                    machine_col = "رقم المكينه"
                else:
                    print(f"[DEBUG] Could not find machine column or slides data")
                    return pd.DataFrame()
            
            print(f"[DEBUG] Machine column: {machine_col}")
            
            # Drop rows where essential data is missing
            block_col = None
            for col in df.columns:
                if 'بلوك' in str(col).lower():
                    block_col = col
                    break
            
            if block_col:
                df = df.dropna(subset=[block_col])
            
            # Filter by machine number if provided
            if machine_number and machine_col in df.columns:
                print(f"[DEBUG] Filtering by machine: {machine_number}")
                
                df[machine_col] = df[machine_col].astype(str).str.strip()
                machine_str = str(machine_number).strip()
                
                filtered_df = df[df[machine_col] == machine_str]
                
                if filtered_df.empty:
                    filtered_df = df[df[machine_col].str.contains(machine_str, na=False)]
                
                df = filtered_df
                print(f"[DEBUG] After filter shape: {df.shape}")
            
        except Exception as e:
            print(f"[ERROR] Failed to load machine production data: {e}")
            import traceback
            traceback.print_exc()
            return pd.DataFrame()
    else:
        print(f"[DEBUG] No files found")
        return pd.DataFrame()
    
    return df


def load_invoices_data(documents_path: str) -> pd.DataFrame:
    """Load all invoices data from client folders"""
    invoices_path = os.path.join(documents_path, "الفواتير")
    
    if not os.path.exists(invoices_path):
        return pd.DataFrame()
    
    all_data = []
    
    try:
        for client_folder in os.listdir(invoices_path):
            client_path = os.path.join(invoices_path, client_folder)
            if not os.path.isdir(client_path):
                continue
            
            ledger_file = os.path.join(client_path, "كشف حساب.xlsx")
            if os.path.exists(ledger_file):
                try:
                    df = pd.read_excel(ledger_file, skiprows=2)
                    df["العميل"] = client_folder
                    all_data.append(df)
                except:
                    pass
        
        if all_data:
            return pd.concat(all_data, ignore_index=True)
        return pd.DataFrame()
    except Exception as e:
        print(f"[ERROR] Failed to load invoices data: {e}")
        return pd.DataFrame()


# ============ REPORT EXECUTION ============

def execute_report(query: Dict, documents_path: str) -> Optional[str]:
    """Execute the report based on parsed query and return the output file path."""
    report_type = query.get("report_type", "income")
    date_from = query.get("date_from")
    date_to = query.get("date_to")
    
    df = pd.DataFrame()
    
    # Load data based on report type
    if report_type == "income":
        df_income, _ = load_income_expenses_data(documents_path)
        df = df_income
    
    elif report_type == "expenses":
        _, df_expenses = load_income_expenses_data(documents_path)
        df = df_expenses
    
    elif report_type == "inventory":
        df_inv, _, _ = load_inventory_data(documents_path)
        df = df_inv
    
    elif report_type == "inventory_add":
        _, df_add, _ = load_inventory_data(documents_path)
        df = df_add
    
    elif report_type == "inventory_disburse":
        _, _, df_disburse = load_inventory_data(documents_path)
        df = df_disburse
    
    elif report_type == "attendance":
        df = load_attendance_data(documents_path)
    
    elif report_type == "blocks":
        df = load_blocks_data(documents_path)
    
    elif report_type == "slides":
        df = load_slides_data(documents_path)
    
    elif report_type == "machine_production":
        machine_number = query.get("machine_number")
        # If no specific machine, generate combined report for all machines
        if machine_number is None:
            return generate_all_machines_report(query, documents_path)
        df = load_machine_production_data(documents_path, machine_number)
    
    elif report_type == "clients":
        df = load_invoices_data(documents_path)
    
    # New report types
    elif report_type == "income_expenses_summary":
        df_income, df_expenses = load_income_expenses_data(documents_path)
        
        # Apply date filters first
        if date_from or date_to:
            if not df_income.empty:
                df_income = apply_date_filter(df_income, date_from, date_to)
            if not df_expenses.empty:
                df_expenses = apply_date_filter(df_expenses, date_from, date_to)
        
        # Calculate totals
        total_income = 0
        total_expenses = 0
        
        if not df_income.empty and "المبلغ" in df_income.columns:
            total_income = pd.to_numeric(df_income["المبلغ"], errors='coerce').sum()
        
        if not df_expenses.empty and "المبلغ" in df_expenses.columns:
            total_expenses = pd.to_numeric(df_expenses["المبلغ"], errors='coerce').sum()
        
        net_profit = total_income - total_expenses
        
        # Create summary dataframe
        df = pd.DataFrame({
            "البيان": ["إجمالي الإيرادات", "إجمالي المصروفات", "صافي الربح"],
            "المبلغ": [total_income, total_expenses, net_profit]
        })
        
        # Skip date filter since we already applied it
        date_from = None
        date_to = None
    
    elif report_type == "attendance_summary":
        df = load_attendance_data(documents_path)
        # Group by employee for summary
        if not df.empty and "الاسم" in df.columns:
            df = df.groupby("الاسم").agg({
                "إجمالي الورديات": "sum",
                "الإجمالي": "sum",
                "السلفة": "sum"
            }).reset_index()
    
    elif report_type == "blocks_by_material":
        df = load_blocks_data(documents_path)
        if not df.empty and "الخامه" in df.columns:
            df = df.sort_values("الخامه")
    
    elif report_type == "blocks_by_quarry":
        df = load_blocks_data(documents_path)
        if not df.empty and "المحجر" in df.columns:
            df = df.sort_values("المحجر")
    
    elif report_type == "slides_by_block":
        df = load_slides_data(documents_path)
        if not df.empty and "رقم البلوك" in df.columns:
            df = df.sort_values("رقم البلوك")
    
    elif report_type == "clients_balances":
        df = load_invoices_data(documents_path)
        # Group by client for balances
        if not df.empty and "العميل" in df.columns:
            # Try to calculate balances if columns exist
            pass  # Keep original data for now
    
    if df.empty:
        print(f"[DEBUG] DataFrame is empty after loading")
        return None
    
    print(f"[DEBUG] DataFrame shape before date filter: {df.shape}")
    
    # Apply date filters
    if date_from or date_to:
        df = apply_date_filter(df, date_from, date_to)
        print(f"[DEBUG] DataFrame shape after date filter: {df.shape}")
    
    # Apply item filter for inventory
    item_name = query.get("item_name")
    if item_name and "اسم الصنف" in df.columns:
        df = df[df["اسم الصنف"].str.contains(item_name, na=False)]
    
    # Apply block filter
    block_number = query.get("block_number")
    if block_number and "رقم البلوك" in df.columns:
        df = df[df["رقم البلوك"].str.contains(str(block_number), na=False)]
    
    # Apply client filter
    client_name = query.get("client_name")
    if client_name:
        if "العميل" in df.columns:
            df = df[df["العميل"].str.contains(client_name, na=False)]
    
    return generate_report_excel(df, query, documents_path)


def apply_date_filter(df: pd.DataFrame, date_from: str, date_to: str) -> pd.DataFrame:
    """Apply date range filter to dataframe"""
    date_col = None
    for col in ["التاريخ", "تاريخ النشر", "تاريخ الدخول", "date"]:
        if col in df.columns:
            date_col = col
            break
    
    if date_col is None:
        print(f"[DEBUG] No date column found in: {list(df.columns)}")
        return df
    
    print(f"[DEBUG] Using date column: {date_col}")
    print(f"[DEBUG] Date values sample: {df[date_col].head()}")
    
    try:
        df = df.copy()
        # Try multiple date formats
        df["date_parsed"] = pd.to_datetime(df[date_col], format="%d/%m/%Y", errors="coerce")
        
        # If parsing failed, try other formats
        if df["date_parsed"].isna().all():
            df["date_parsed"] = pd.to_datetime(df[date_col], errors="coerce")
        
        print(f"[DEBUG] Parsed dates sample: {df['date_parsed'].head()}")
        
        if date_from:
            from_dt = datetime.strptime(date_from, "%d/%m/%Y")
            print(f"[DEBUG] Filtering from: {from_dt}")
            df = df[df["date_parsed"] >= from_dt]
            print(f"[DEBUG] After from filter: {len(df)} rows")
        
        if date_to:
            to_dt = datetime.strptime(date_to, "%d/%m/%Y")
            print(f"[DEBUG] Filtering to: {to_dt}")
            df = df[df["date_parsed"] <= to_dt]
            print(f"[DEBUG] After to filter: {len(df)} rows")
        
        df = df.drop(columns=["date_parsed"])
    except Exception as e:
        print(f"[ERROR] Date filter failed: {e}")
        import traceback
        traceback.print_exc()
    
    return df


def generate_all_machines_report(query: Dict, documents_path: str) -> str:
    """Generate a combined report for all machines with separate tables"""
    reports_path = os.path.join(documents_path, "التقارير")
    os.makedirs(reports_path, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"تقرير إنتاج جميع الماكينات_{timestamp}.xlsx"
    filepath = os.path.join(reports_path, filename)
    
    date_from = query.get("date_from")
    date_to = query.get("date_to")
    
    workbook = xlsxwriter.Workbook(filepath)
    worksheet = workbook.add_worksheet("إنتاج الماكينات")
    worksheet.right_to_left()
    
    # Formats
    main_title_fmt = workbook.add_format({
        'bold': True, 'align': 'center', 'valign': 'vcenter',
        'bg_color': '#1F4E78', 'font_color': 'white', 'font_size': 18, 'border': 2
    })
    
    machine_title_fmt = workbook.add_format({
        'bold': True, 'align': 'center', 'valign': 'vcenter',
        'bg_color': '#2E7D32', 'font_color': 'white', 'font_size': 14, 'border': 2
    })
    
    header_fmt = workbook.add_format({
        'bold': True, 'align': 'center', 'valign': 'vcenter',
        'bg_color': '#4472C4', 'font_color': 'white', 'font_size': 11, 'border': 1
    })
    
    cell_fmt = workbook.add_format({
        'align': 'center', 'valign': 'vcenter', 'font_size': 10, 'border': 1
    })
    
    cell_fmt_alt = workbook.add_format({
        'align': 'center', 'valign': 'vcenter', 'font_size': 10, 'border': 1,
        'bg_color': '#F2F2F2'
    })
    
    number_fmt = workbook.add_format({
        'align': 'center', 'valign': 'vcenter', 'font_size': 10, 'border': 1,
        'num_format': '#,##0.00'
    })
    
    subtotal_label_fmt = workbook.add_format({
        'bold': True, 'align': 'center', 'valign': 'vcenter',
        'bg_color': '#FFC107', 'font_color': '#000000', 'font_size': 11, 'border': 2
    })
    
    subtotal_value_fmt = workbook.add_format({
        'bold': True, 'align': 'center', 'valign': 'vcenter',
        'bg_color': '#FFC107', 'font_color': '#000000', 'font_size': 11, 'border': 2,
        'num_format': '#,##0.00'
    })
    
    grand_total_label_fmt = workbook.add_format({
        'bold': True, 'align': 'center', 'valign': 'vcenter',
        'bg_color': '#1F4E78', 'font_color': 'white', 'font_size': 14, 'border': 2
    })
    
    grand_total_value_fmt = workbook.add_format({
        'bold': True, 'align': 'center', 'valign': 'vcenter',
        'bg_color': '#C6EFCE', 'font_color': '#006100', 'font_size': 14, 'border': 2,
        'num_format': '#,##0.00'
    })
    
    # Main title
    date_range = ""
    if date_from and date_to:
        date_range = f" من {date_from} إلى {date_to}"
    elif date_from:
        date_range = f" من {date_from}"
    elif date_to:
        date_range = f" حتى {date_to}"
    
    current_row = 0
    worksheet.merge_range(current_row, 0, current_row, 9, f"تقرير إنتاج جميع الماكينات{date_range}", main_title_fmt)
    worksheet.set_row(current_row, 35)
    current_row += 2
    
    # Load all machines data
    machines = ["1", "2", "3"]
    machine_totals = {}
    grand_total_quantity = 0
    grand_total_area = 0
    grand_total_price = 0
    
    for machine_num in machines:
        df = load_machine_production_data(documents_path, machine_num)
        
        if df.empty:
            continue
        
        # Apply date filter
        if date_from or date_to:
            df = apply_date_filter(df, date_from, date_to)
        
        if df.empty:
            continue
        
        # Machine title
        worksheet.merge_range(current_row, 0, current_row, 9, f"ماكينة {machine_num}", machine_title_fmt)
        worksheet.set_row(current_row, 25)
        current_row += 1
        
        # Select important columns for display
        display_cols = []
        col_mapping = {}
        
        for col in df.columns:
            col_str = str(col)
            if any(k in col_str for k in ['تاريخ', 'بلوك', 'النوع', 'السمك', 'العدد', 'الطول', 'الارتفاع', 'الكمية', 'سعر', 'اجمالي']):
                display_cols.append(col)
        
        if not display_cols:
            display_cols = list(df.columns)[:10]
        
        # Headers
        for col_idx, col_name in enumerate(display_cols):
            worksheet.write(current_row, col_idx, str(col_name), header_fmt)
        worksheet.set_row(current_row, 22)
        current_row += 1
        
        # Data rows
        for i, (_, row) in enumerate(df.iterrows()):
            is_alt = i % 2 == 1
            for col_idx, col_name in enumerate(display_cols):
                value = row[col_name] if col_name in row.index else ""
                if isinstance(value, (int, float)) and not pd.isna(value):
                    worksheet.write(current_row, col_idx, value, number_fmt)
                else:
                    fmt = cell_fmt_alt if is_alt else cell_fmt
                    worksheet.write(current_row, col_idx, str(value) if not pd.isna(value) else "", fmt)
            current_row += 1
        
        # Calculate machine totals
        machine_qty = 0
        machine_area = 0
        machine_price = 0
        
        for col in df.columns:
            col_str = str(col)
            if 'العدد' in col_str:
                machine_qty = pd.to_numeric(df[col], errors='coerce').sum()
            elif 'الكمية' in col_str and 'م2' in col_str:
                machine_area = pd.to_numeric(df[col], errors='coerce').sum()
            elif 'اجمالي' in col_str and 'سعر' in col_str.lower():
                machine_price = pd.to_numeric(df[col], errors='coerce').sum()
        
        machine_totals[machine_num] = {
            'qty': machine_qty,
            'area': machine_area,
            'price': machine_price
        }
        
        grand_total_quantity += machine_qty if pd.notna(machine_qty) else 0
        grand_total_area += machine_area if pd.notna(machine_area) else 0
        grand_total_price += machine_price if pd.notna(machine_price) else 0
        
        # Machine subtotal row
        num_display_cols = len(display_cols)
        if num_display_cols >= 4:
            worksheet.merge_range(current_row, 0, current_row, 1, f"إجمالي ماكينة {machine_num}", subtotal_label_fmt)
            worksheet.write(current_row, 2, machine_area if pd.notna(machine_area) else 0, subtotal_value_fmt)
            worksheet.merge_range(current_row, 3, current_row, num_display_cols - 1, machine_price if pd.notna(machine_price) else 0, subtotal_value_fmt)
        else:
            worksheet.write(current_row, 0, f"إجمالي ماكينة {machine_num}", subtotal_label_fmt)
            worksheet.write(current_row, 1, machine_price if pd.notna(machine_price) else 0, subtotal_value_fmt)
        
        worksheet.set_row(current_row, 22)
        current_row += 2  # Add space between machines
    
    # Grand total section
    if machine_totals:
        current_row += 1
        worksheet.merge_range(current_row, 0, current_row, 9, "ملخص الإنتاجية الكلية", grand_total_label_fmt)
        worksheet.set_row(current_row, 30)
        current_row += 1
        
        # Summary table headers
        summary_headers = ["الماكينة", "الكمية م2", "إجمالي السعر"]
        for col_idx, header in enumerate(summary_headers):
            worksheet.write(current_row, col_idx, header, header_fmt)
        current_row += 1
        
        # Summary data for each machine
        for machine_num, totals in machine_totals.items():
            worksheet.write(current_row, 0, f"ماكينة {machine_num}", cell_fmt)
            worksheet.write(current_row, 1, totals['area'] if pd.notna(totals['area']) else 0, number_fmt)
            worksheet.write(current_row, 2, totals['price'] if pd.notna(totals['price']) else 0, number_fmt)
            current_row += 1
        
        # Grand total row
        worksheet.write(current_row, 0, "الإجمالي الكلي", grand_total_label_fmt)
        worksheet.write(current_row, 1, grand_total_area, grand_total_value_fmt)
        worksheet.write(current_row, 2, grand_total_price, grand_total_value_fmt)
        worksheet.set_row(current_row, 25)
    
    # Set column widths
    worksheet.set_column(0, 0, 14)  # Date/Machine
    worksheet.set_column(1, 1, 14)  # Block
    worksheet.set_column(2, 2, 14)  # Type
    worksheet.set_column(3, 3, 12)  # Thickness
    worksheet.set_column(4, 4, 10)  # Count
    worksheet.set_column(5, 5, 12)  # Length
    worksheet.set_column(6, 6, 12)  # Height
    worksheet.set_column(7, 7, 14)  # Area
    worksheet.set_column(8, 8, 14)  # Price
    worksheet.set_column(9, 9, 16)  # Total
    
    workbook.close()
    return filepath


def generate_report_excel(df: pd.DataFrame, query: Dict, documents_path: str) -> str:
    """Generate Excel report from dataframe"""
    reports_path = os.path.join(documents_path, "التقارير")
    os.makedirs(reports_path, exist_ok=True)
    
    report_type = query.get("report_type", "report")
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    
    type_names = {
        "income": "الإيرادات",
        "expenses": "المصروفات",
        "income_expenses_summary": "ملخص الإيرادات والمصروفات",
        "inventory": "المخزون",
        "inventory_add": "إضافات المخزون",
        "inventory_disburse": "صرف المخزون",
        "attendance": "الحضور والانصراف",
        "attendance_summary": "ملخص الحضور",
        "blocks": "البلوكات",
        "blocks_by_material": "البلوكات حسب الخامة",
        "blocks_by_quarry": "البلوكات حسب المحجر",
        "slides": "الشرائح",
        "slides_by_block": "الشرائح حسب البلوك",
        "machine_production": "إنتاج الماكينات",
        "clients": "العملاء",
        "clients_balances": "أرصدة العملاء"
    }
    
    report_name = type_names.get(report_type, "تقرير")
    
    # Add machine number to report name if available
    machine_number = query.get("machine_number")
    if report_type == "machine_production" and machine_number:
        report_name = f"إنتاج ماكينة {machine_number}"
    
    filename = f"تقرير {report_name}_{timestamp}.xlsx"
    filepath = os.path.join(reports_path, filename)
    
    workbook = xlsxwriter.Workbook(filepath)
    worksheet = workbook.add_worksheet("التقرير")
    worksheet.right_to_left()
    
    # Formats
    title_fmt = workbook.add_format({
        'bold': True, 'align': 'center', 'valign': 'vcenter',
        'bg_color': '#1F4E78', 'font_color': 'white', 'font_size': 16, 'border': 2
    })
    
    header_fmt = workbook.add_format({
        'bold': True, 'align': 'center', 'valign': 'vcenter',
        'bg_color': '#4472C4', 'font_color': 'white', 'font_size': 12, 'border': 1
    })
    
    cell_fmt = workbook.add_format({
        'align': 'center', 'valign': 'vcenter', 'font_size': 11, 'border': 1
    })
    
    cell_fmt_alt = workbook.add_format({
        'align': 'center', 'valign': 'vcenter', 'font_size': 11, 'border': 1,
        'bg_color': '#F2F2F2'
    })
    
    number_fmt = workbook.add_format({
        'align': 'center', 'valign': 'vcenter', 'font_size': 11, 'border': 1,
        'num_format': '#,##0'
    })
    
    number_fmt_alt = workbook.add_format({
        'align': 'center', 'valign': 'vcenter', 'font_size': 11, 'border': 1,
        'num_format': '#,##0', 'bg_color': '#F2F2F2'
    })
    
    summary_label_fmt = workbook.add_format({
        'bold': True, 'align': 'center', 'valign': 'vcenter',
        'bg_color': '#1F4E78', 'font_color': 'white', 'font_size': 13, 'border': 2
    })
    
    summary_value_fmt = workbook.add_format({
        'bold': True, 'align': 'center', 'valign': 'vcenter',
        'bg_color': '#C6EFCE', 'font_color': '#006100', 'font_size': 13, 'border': 2,
        'num_format': '#,##0'
    })
    
    # Special format for net profit row (golden/yellow)
    net_profit_label_fmt = workbook.add_format({
        'bold': True, 'align': 'center', 'valign': 'vcenter',
        'bg_color': '#FFD700', 'font_color': '#000000', 'font_size': 14, 'border': 2
    })
    
    net_profit_value_fmt = workbook.add_format({
        'bold': True, 'align': 'center', 'valign': 'vcenter',
        'bg_color': '#FFD700', 'font_color': '#006100', 'font_size': 14, 'border': 2,
        'num_format': '#,##0'
    })
    
    # Title with date range
    date_range = ""
    if query.get("date_from") and query.get("date_to"):
        date_range = f" من {query['date_from']} إلى {query['date_to']}"
    elif query.get("date_from"):
        date_range = f" من {query['date_from']}"
    elif query.get("date_to"):
        date_range = f" حتى {query['date_to']}"
    
    title = f"تقرير {report_name}{date_range}"
    num_cols = len(df.columns) if not df.empty else 4
    
    worksheet.merge_range(0, 0, 0, num_cols - 1, title, title_fmt)
    worksheet.set_row(0, 35)
    
    if not df.empty:
        # Headers
        for col_idx, col_name in enumerate(df.columns):
            worksheet.write(1, col_idx, col_name, header_fmt)
        worksheet.set_row(1, 25)
        
        # Data rows
        data_start_row = 2
        for i, (_, row) in enumerate(df.iterrows()):
            current_row = data_start_row + i
            is_alt = i % 2 == 1
            
            for col_idx, value in enumerate(row):
                if isinstance(value, (int, float)) and not pd.isna(value):
                    fmt = number_fmt_alt if is_alt else number_fmt
                    worksheet.write(current_row, col_idx, value, fmt)
                else:
                    fmt = cell_fmt_alt if is_alt else cell_fmt
                    worksheet.write(current_row, col_idx, str(value) if not pd.isna(value) else "", fmt)
        
        # Column widths - smart sizing based on content
        for col_idx, col_name in enumerate(df.columns):
            col_name_str = str(col_name)
            header_len = len(col_name_str) * 1.8  # Arabic characters need more space
            
            # Calculate max data length
            data_len = 8
            if len(df) > 0:
                try:
                    data_len = df.iloc[:, col_idx].astype(str).str.len().max()
                except:
                    data_len = 8
            
            # Smart width based on column type
            if any(keyword in col_name_str for keyword in ['تاريخ', 'وقت']):
                # Date/time columns - fixed width
                col_width = 14
            elif any(keyword in col_name_str for keyword in ['رقم', 'العدد', 'السمك', 'الخصم']):
                # Number columns - smaller width
                col_width = 12
            elif any(keyword in col_name_str for keyword in ['النوع', 'الخامه']):
                # Type columns - medium width
                col_width = 18
            elif any(keyword in col_name_str for keyword in ['اسم', 'البيان', 'ملاحظات']):
                # Text columns - larger width
                col_width = max(header_len, data_len, 25)
            elif any(keyword in col_name_str for keyword in ['سعر', 'اجمالي', 'الإجمالي', 'المبلغ', 'الكمية']):
                # Price/total columns - medium width
                col_width = 16
            else:
                # Default - based on content
                col_width = max(header_len, data_len, 14)
            
            # Cap maximum width
            col_width = min(col_width, 40)
            worksheet.set_column(col_idx, col_idx, col_width)
        
        # Skip summary row for income_expenses_summary (already has net profit)
        if report_type != "income_expenses_summary":
            # Summary row - directly after data (no gap)
            summary_row = data_start_row + len(df)
            
            # Find amount/total column and calculate sum
            amount_cols = ["المبلغ", "الإجمالي", "اجمالي السعر", "الرصيد الحالي", "الكمية م2"]
            total_value = 0
            total_col_idx = num_cols - 1  # Default to last column
            
            for col_name in amount_cols:
                if col_name in df.columns:
                    col_idx = df.columns.get_loc(col_name)
                    total_col_idx = col_idx
                    # Convert column to numeric and sum
                    try:
                        numeric_col = pd.to_numeric(df.iloc[:, col_idx], errors='coerce')
                        total_value = numeric_col.sum()
                        if pd.notna(total_value):
                            break
                    except:
                        pass
            
            # Determine label and format based on report type
            if report_type == "income":
                summary_label = "صافي الربح"
                label_fmt = net_profit_label_fmt
                value_fmt = net_profit_value_fmt
            elif report_type == "expenses":
                summary_label = "إجمالي المصروفات"
                label_fmt = summary_label_fmt
                value_fmt = summary_value_fmt
            else:
                summary_label = "الإجمالي"
                label_fmt = summary_label_fmt
                value_fmt = summary_value_fmt
            
            # Write summary row
            if num_cols >= 5:
                worksheet.merge_range(summary_row, 0, summary_row, 2, summary_label, label_fmt)
                worksheet.merge_range(summary_row, 3, summary_row, num_cols - 1, total_value if pd.notna(total_value) else 0, value_fmt)
            elif num_cols >= 3:
                mid = num_cols // 2
                worksheet.merge_range(summary_row, 0, summary_row, mid - 1, summary_label, label_fmt)
                worksheet.merge_range(summary_row, mid, summary_row, num_cols - 1, total_value if pd.notna(total_value) else 0, value_fmt)
            elif num_cols == 2:
                worksheet.write(summary_row, 0, summary_label, label_fmt)
                worksheet.write(summary_row, 1, total_value if pd.notna(total_value) else 0, value_fmt)
            else:
                worksheet.write(summary_row, 0, f"{summary_label}: {total_value if pd.notna(total_value) else 0}", label_fmt)
            
            # Set row height for summary
            worksheet.set_row(summary_row, 25)
    else:
        worksheet.merge_range(1, 0, 1, 3, "لا توجد بيانات متاحة", cell_fmt)
    
    workbook.close()
    return filepath
