import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from utils.blocks_utils import append_to_existing_file
import openpyxl
from openpyxl.styles import PatternFill

# Test color definitions
try:
    # Test with a simple color
    fill = PatternFill(start_color="FFB3BA", end_color="FFB3BA", fill_type="solid")
    print("Basic color test passed")
    
    # Test with hash prefix
    fill = PatternFill(start_color="#FFB3BA", end_color="#FFB3BA", fill_type="solid")
    print("Hash prefix color test passed")
    
    # Test creating a workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    cell = ws.cell(row=1, column=1, value="test")
    cell.fill = PatternFill(start_color="FFB3BA", end_color="FFB3BA", fill_type="solid")
    wb.save("color_test.xlsx")
    print("Workbook creation test passed")
    
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()